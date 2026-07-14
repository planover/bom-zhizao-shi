#!/usr/bin/env python3
"""
generate_bom.py - BOM智造师 配套脚本（V7 增量）

读取包含产品信息、物料与工艺工序信息的 JSON，生成格式化的 BOM 表 Excel (.xlsx)。
依赖 openpyxl；若运行环境缺失则自动安装。

V7 增量（相对 V6）：
- 双语开关 --bilingual：主 sheet「BOM表」保持纯中文（与 V6 字节级一致），
  开启时追加第二个 sheet「BOM表(英)」，区块标题「中文 (English)」合并行、
  表头中英双行（中文在上、英文在下）；逆向恒读中文 sheet，零变化。
- B1 批量空白模板：--blank-templates + --out-dir [--industries]，每行业产
  template_<行业>.xlsx（含区块标题+表头+空物料/工序区，不触发校验）。
- B2 批量生成 BOM：--batch-dir <dir> 或 --batch f1,f2,... + --out-dir，
  逐条生成 BOM_<产品名>_<日期>.xlsx，单文件失败记录错误继续，结束汇总。
- 翻译字典承载于 bom_constants.I18N（eng 为键），仅正向渲染引用。

既有能力（V2–V6）全部沿用不变：物料区 8 列 A–H、industry 8 值枚举、7 大专属
视图、成本视图双编号、37 唯一键、行业模板预设、完整逆向闭环。

用法:
    # 单条生成（V6 行为零变化）
    python3 generate_bom.py --data bom.json --out BOM_2026-07-07.xlsx
    # 双语导出
    python3 generate_bom.py --data bom.json --out BOM_x.xlsx --bilingual
    # B1 批量空白模板
    python3 generate_bom.py --blank-templates --out-dir ./templates
    # B2 批量生成
    python3 generate_bom.py --batch-dir ./inputs --out-dir ./outputs

输入 JSON 结构见 references/bom-spec.md。
"""
import argparse
import datetime
import glob
import json
import os
import sys
import subprocess
from types import SimpleNamespace

from bom_constants import (
    INDUSTRIES,
    CATEGORY_TO_INDUSTRY,
    COMPONENT_EXCLUDE,
    FORMULA_EXCLUDE,
    EDIBLE,
    ALLERGEN_SET,
    ALLERGEN_HINTS,
    TEXTILE_TYPES,
    TEXTILE_EXCLUDE,
    FURNITURE_TYPES,
    FURNITURE_EXCLUDE,
    MECHANICAL_TYPES,
    MECHANICAL_EXCLUDE,
    PACKAGING_TYPES,
    PACKAGING_EXCLUDE,
    EDIBLE_LIST,
    INDUSTRY_TEMPLATES,
    I18N,
    ZH2EN,
)

# V7: 8 行业固定输出顺序（B1 模板 / 默认全行业）
ALL_INDUSTRIES = ["食品", "电子", "化工", "机械", "纺织", "家具", "包装", "通用"]

# 产品类别枚举（R1 / R4）
CATEGORIES = {"食品", "工业品", "日化化妆品", "医药", "其他"}

# V5 成本币种默认值（currency 选填，缺省为人民币(CNY)）
DEFAULT_CURRENCY = "人民币(CNY)"
# 存在「三、」行业派生视图的行业集合（用于成本视图双编号判定）
# V6：扩至含机械/包装共 7 行业（机械/包装带成本时为「四、成本明细」）
INDUSTRY_VIEW_SET = {"食品", "电子", "化工", "纺织", "家具", "机械", "包装"}


def ensure_openpyxl():
    """确保 openpyxl 可用，缺失时自动安装。"""
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"]
        )


def load_data(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def infer_industry(data):
    """推断 industry：显式 > category 推断。返回 (industry, warnings)。

    - industry 已显式设置且合法 → 使用该值，warnings=[]
    - industry 已显式设置但非法 → V8 WARNING，回退为推断值
    - industry 未设置 → 按 category 推断（食品→食品，日化/医药→化工，其他→通用）

    Args:
        data: BOM JSON dict。

    Returns:
        (industry_str, warnings_list)
    """
    industry = str(data.get("industry") or "").strip()
    category = str(data.get("category") or "其他").strip()

    if industry:
        if industry in INDUSTRIES:
            return industry, []
        # V8: 非法值 WARNING（非阻断），回退推断
        warning = (
            "WARNING: industry 值『%s』不在枚举内"
            "（食品/电子/化工/机械/纺织/家具/包装/通用），已回退为推断值"
            % industry
        )
        return CATEGORY_TO_INDUSTRY.get(category, "通用"), [warning]

    # 未填 → 按 category 推断
    return CATEGORY_TO_INDUSTRY.get(category, "通用"), []


def validate(data):
    """返回错误列表；为空表示校验通过。

    校验项（V1–V7）：
    - V1(R1): product_name 必填非空
    - V2(R1): category 必填且枚举
    - V3(R2): output_rate 必填且 >0（允许 >100）
    - V4(R2): 物料 yield_rate 须 0<值≤100
    - V5(R3): 工序流转链完整性（仅 len(processes)>=2 触发，阻断级）
    - V6(沿用): 工序编号唯一 / 名称非空 / 工时≥0
    - V7(沿用): 物料 name/unit 非空、usage>0

    说明：V3 的 `approver`/`effective_date`/`standard`/`allergen` 与 V4 的
    `industry`/专属物料字段均为可选，不纳入阻断级校验。
    V8（industry 枚举软校验）由 infer_industry() 返回，非阻断。
    """
    errors = []
    if not isinstance(data, dict):
        return ["输入数据必须是 JSON 对象"]

    # V1(R1): product_name 必填非空
    product_name = str(data.get("product_name") or "").strip()
    if not product_name:
        errors.append("产品名称为必填，且不可为空")

    # V2(R1): category 必填且枚举
    category = str(data.get("category") or "").strip()
    if not category or category not in CATEGORIES:
        errors.append("产品类别为必填，且须为：食品/工业品/日化化妆品/医药/其他")

    # V3(R2): output_rate 必填且 >0（允许 >100）
    raw_rate = data.get("output_rate")
    try:
        rate = float(raw_rate)
        if rate <= 0:
            errors.append(
                "全产品出品率(output_rate)为必填，且须为正数（可大于100，如干香菇泡发增重）"
            )
    except (TypeError, ValueError):
        errors.append(
            "全产品出品率(output_rate)为必填，且须为正数（可大于100，如干香菇泡发增重）"
        )

    materials = data.get("materials", [])
    if not isinstance(materials, list) or len(materials) == 0:
        errors.append("至少需要一条物料记录")

    for i, m in enumerate(materials, 1):
        name = str(m.get("name") or "").strip()
        unit = str(m.get("unit") or "").strip()
        usage = m.get("usage")
        yield_rate = m.get("yield_rate")

        if not name:
            errors.append(f"物料#{i} 物料名称为必填")
        if not unit:
            errors.append(f"物料#{i} 计量单位为必填")
        try:
            u = float(usage)
            if u <= 0:
                errors.append(f"物料#{i} 用量必须为正数")
        except (TypeError, ValueError):
            errors.append(f"物料#{i} 用量必须为数值")
        try:
            y = float(yield_rate)
            if y <= 0 or y > 100:
                errors.append(f"物料#{i} 出品率须为 0-100 的正数")
        except (TypeError, ValueError):
            errors.append(f"物料#{i} 出品率为必填且须为数值")

    seen = set()
    for i, p in enumerate(data.get("processes", []), 1):
        step_no = str(p.get("step_no") or "").strip()
        name = str(p.get("name") or "").strip()

        if not step_no:
            errors.append(f"工序#{i} 工序编号为必填")
        elif step_no in seen:
            errors.append(f"工序编号 {step_no} 重复")
        else:
            seen.add(step_no)
        if not name:
            errors.append(f"工序#{i} 工序名称为必填")

        wh = p.get("work_hours")
        if wh not in (None, ""):
            try:
                if float(wh) < 0:
                    errors.append(f"工序#{i} 工时须 >= 0")
            except (TypeError, ValueError):
                errors.append(f"工序#{i} 工时格式异常，须为数值或留空")

    # V5(R3): 工序流转链完整性（阻断级）
    procs = data.get("processes", [])
    if isinstance(procs, list) and len(procs) >= 2:
        for i in range(1, len(procs)):
            prev = procs[i - 1]
            cur = procs[i]
            prev_out = str(prev.get("output") or "").strip()
            if not prev_out:
                errors.append(f"工序 {prev.get('step_no')} 的产物(output)为必填")
                continue
            cur_materials = [
                m
                for m in materials
                if str(m.get("process") or "").strip()
                == str(cur.get("step_no") or "").strip()
            ]
            names = {str(m.get("name") or "").strip() for m in cur_materials}
            if prev_out not in names:
                errors.append(
                    f"工序 {cur.get('step_no')} 的物料清单未包含上一工序 "
                    f"{prev.get('step_no')} 的产物『{prev_out}』，流转链不完整"
                )

    return errors


def derive_ingredients(data, industry=None):
    """派生配料表（仅食品行业）。

    返回 (ingredients, excluded)：
    - ingredients: 可食用物料（material_type ∈ EDIBLE），若 industry != 食品 则为空列表
    - excluded: 非食用物料（包材/其他/未分类/未填），始终保留用于 WARNING 统计

    ingredients 按 usage 降序排列（食品标签惯例）。

    V4 变更：触发条件从 category=="食品" 改为 industry=="食品"（含推断）。
    核心逻辑（过滤 EDIBLE / 排序 / 返回 excluded）完全不变。
    industry=None 时内部调用 infer_industry 推断（向后兼容旧调用）。
    """
    ingredients, excluded = [], []
    for m in data.get("materials", []):
        mt = str(m.get("material_type") or "其他").strip() or "其他"
        if mt in EDIBLE:
            ingredients.append(m)
        else:
            excluded.append(m)

    if industry is None:
        industry, _ = infer_industry(data)
    if industry != "食品":
        return [], excluded

    ingredients.sort(key=lambda x: float(x.get("usage") or 0), reverse=True)
    return ingredients, excluded


def ingredient_pct(items):
    """计算可食用物料的用量占比%（1 位小数），保证列和恰为 100.0%。

    分母 = 各物料 usage 求和；每项 round(usage / 合计 × 100, 1)。
    使用最大余数法对末位补差（0.1 粒度），使整列合计严格等于 100.0。

    Args:
        items: 可食用物料列表，每项含 "usage" 键。

    Returns:
        (pct_list, total)：pct_list 与输入 items 顺序一致；total 为用量合计。
    """
    total = sum(float(it.get("usage", 0)) for it in items) or 1.0
    raw = [float(it.get("usage", 0)) / total * 100 for it in items]
    pct = [round(x, 1) for x in raw]
    s = round(sum(pct), 1)
    diff = round(100.0 - s, 1)
    if diff != 0:
        # 按最大小数余数分配补差（0.1 粒度）
        order = sorted(range(len(raw)), key=lambda i: (raw[i] - pct[i]), reverse=True)
        step = 0.1 if diff > 0 else -0.1
        d = abs(diff)
        for i in order:
            if round(d, 1) <= 0:
                break
            pct[i] = round(pct[i] + step, 1)
            d = round(d - 0.1, 1)
    return pct, total


def check_allergen_soft(data):
    """校验 allergen：标签合法性(W1) + 名称关键词启发式(H1)，均非阻断。

    返回 WARNING 文本列表（不进入 errors，不改变退出码）：
    - W1：标签非空且任一标签 ∉ ALLERGEN_SET → 告警「过敏原标签不在八大类集合」。
    - H1：名称含致敏物关键词但其 allergen 未涵盖对应类别 → 告警
          「名称疑似含致敏物『{class}』但未在过敏原中标注，请确认」。
    """
    warnings = []
    for i, m in enumerate(data.get("materials", []), 1):
        name = str(m.get("name") or f"#{i}").strip()
        raw = str(m.get("allergen") or "").strip()
        # allergen 可能为空串，按逗号拆分得到标签集合（去空）
        tags = [t.strip() for t in raw.split(",") if t.strip()] if raw else []
        tag_set = set(tags)

        # W1：标签合法性校验（仅当标签非空且存在非法标签）
        bad = [t for t in tags if t not in ALLERGEN_SET]
        if bad:
            warnings.append(
                "WARNING: 物料『%s』的过敏原标签『%s』不在八大类集合"
                "（含麸质谷物/甲壳类/蛋类/鱼类/花生/大豆/乳/坚果/其他）内，请确认"
                % (name, "、".join(bad))
            )

        # H1：名称关键词启发式（非阻断，提示性 WARNING）
        # 名称含致敏物关键词，但对应八大类未在该物料 allergen 中标注 → 告警。
        if name:
            warned = set()
            for kw, cls in ALLERGEN_HINTS:
                if kw in name and cls not in tag_set and cls not in warned:
                    warned.add(cls)
                    warnings.append(
                        "WARNING: 物料『%s』名称疑似含致敏物『%s』但未在过敏原中标注，请确认"
                        % (name, cls)
                    )
    return warnings


def derive_components(data):
    """派生元件清单（仅电子行业）。

    返回 (components, excluded)：
    - components: 电子元件物料（排除 material_type ∈ COMPONENT_EXCLUDE，即"其他"类）
    - excluded: 被排除的物料（散热片/外壳/包装等）

    排序：按物料类型分组 → 同类型内按位号(designator)字母数字排序。
    空位号排同类型末尾（排序键用 "\uffff" 哨兵）。
    """
    components, excluded = [], []
    for m in data.get("materials", []):
        mt = str(m.get("material_type") or "其他").strip() or "其他"
        if mt in COMPONENT_EXCLUDE:
            excluded.append(m)
        else:
            components.append(m)

    # 排序：物料类型（升序）→ 位号（字母数字升序，空排末尾）
    components.sort(key=lambda x: (
        str(x.get("material_type") or ""),
        str(x.get("designator") or "\uffff"),
    ))
    return components, excluded


def derive_formula(data):
    """派生配方表（仅化工行业）。

    返回 (formula, excluded)：
    - formula: 配方原料（排除 material_type ∈ FORMULA_EXCLUDE，即"包材"类）
    - excluded: 被排除的物料（瓶子/标签等包材）

    排序：按含量(%) 降序（配方表惯例，主成分在前）。含量为空的排末尾。
    """
    formula, excluded = [], []
    for m in data.get("materials", []):
        mt = str(m.get("material_type") or "其他").strip() or "其他"
        if mt in FORMULA_EXCLUDE:
            excluded.append(m)
        else:
            formula.append(m)

    # 排序：含量(%) 降序，空值排末尾
    formula.sort(key=lambda x: (
        -(float(x.get("concentration") or 0)),
    ))
    return formula, excluded


def derive_textile(data):
    """派生面料辅料清单（仅纺织行业）。

    返回 (items, excluded)：
    - items: 纺织物料（排除 material_type ∈ TEXTILE_EXCLUDE，即"其他"类）
    - excluded: 被排除的物料（包装/其他类）

    排序：物料类型升序 → 物料名称升序（与 V4 电子/化工同构，升序稳定）。
    """
    items, excluded = [], []
    for m in data.get("materials", []):
        mt = str(m.get("material_type") or "其他").strip() or "其他"
        if mt in TEXTILE_EXCLUDE:
            excluded.append(m)
        else:
            items.append(m)

    # 排序：物料类型（升序）→ 物料名称（升序，空排末尾）
    items.sort(key=lambda x: (
        str(x.get("material_type") or ""),
        str(x.get("name") or ""),
    ))
    return items, excluded


def derive_furniture(data):
    """派生家具物料清单（仅家具行业）。

    返回 (items, excluded)：
    - items: 家具物料（排除 material_type ∈ FURNITURE_EXCLUDE，即"其他"类）
    - excluded: 被排除的物料

    排序：物料类型升序 → 物料名称升序。
    """
    items, excluded = [], []
    for m in data.get("materials", []):
        mt = str(m.get("material_type") or "其他").strip() or "其他"
        if mt in FURNITURE_EXCLUDE:
            excluded.append(m)
        else:
            items.append(m)

    # 排序：物料类型（升序）→ 物料名称（升序，空排末尾）
    items.sort(key=lambda x: (
        str(x.get("material_type") or ""),
        str(x.get("name") or ""),
    ))
    return items, excluded


def derive_mechanical(data):
    """派生机械物料清单（仅机械行业）。

    返回 (items, excluded)：
    - items: 机械物料（排除 material_type ∈ MECHANICAL_EXCLUDE，即"其他"类）
    - excluded: 被排除的物料

    排序：物料类型升序 → 物料名称升序（与 derive_textile/derive_furniture 同构）。
    注意：物料类型不进视图展示列，仅用于过滤/排序（从物料区/JSON 取）。
    """
    items, excluded = [], []
    for m in data.get("materials", []):
        mt = str(m.get("material_type") or "其他").strip() or "其他"
        if mt in MECHANICAL_EXCLUDE:
            excluded.append(m)
        else:
            items.append(m)

    # 排序：物料类型（升序）→ 物料名称（升序，空排末尾）
    items.sort(key=lambda x: (
        str(x.get("material_type") or ""),
        str(x.get("name") or ""),
    ))
    return items, excluded


def derive_packaging(data):
    """派生包装物料清单（仅包装行业）。

    返回 (items, excluded)：
    - items: 包装物料（排除 material_type ∈ PACKAGING_EXCLUDE，即"其他"类）
    - excluded: 被排除的物料

    排序：物料类型升序 → 物料名称升序（与 derive_textile/derive_furniture 同构）。
    """
    items, excluded = [], []
    for m in data.get("materials", []):
        mt = str(m.get("material_type") or "其他").strip() or "其他"
        if mt in PACKAGING_EXCLUDE:
            excluded.append(m)
        else:
            items.append(m)

    # 排序：物料类型（升序）→ 物料名称（升序，空排末尾）
    items.sort(key=lambda x: (
        str(x.get("material_type") or ""),
        str(x.get("name") or ""),
    ))
    return items, excluded


def derive_cost(data):
    """派生成本明细（跨行业，任一物料 unit_price 非空即纳入）。

    返回 (cost_items, has_cost)：
    - cost_items: unit_price 非空（≠"" 且可转 float）的物料列表（含被行业视图
      排除的"其他"/"包材"类，成本核算面向全物料）
    - has_cost: 列表非空即为 True（触发生成成本视图）

    注意：total_price 不存 JSON，渲染时按 usage × unit_price 实时计算（派生展示）。
    """
    cost_items = []
    for m in data.get("materials", []):
        up = m.get("unit_price", "")
        if up not in ("", None):
            try:
                if float(up) >= 0:
                    cost_items.append(m)
            except (TypeError, ValueError):
                pass
    return cost_items, bool(cost_items)


def check_industry_soft(data, industry):
    """行业专属软校验（W2/W3，非阻断 WARNING）。

    - W2：industry=="电子" 且物料（未排除的）未标 rohs → WARNING
    - W3：industry=="化工" 且配方原料未填 cas_number 或 ghs_hazard → WARNING
    - 含量(%) 列和校验：所有配方原料均填了 concentration → 校验列和 ≈ 100%（±5%）→ 不达标 WARNING

    Args:
        data: BOM JSON dict。
        industry: 已推断的行业字符串。

    Returns:
        warnings 字符串列表（非阻断）。
    """
    warnings = []
    if industry == "电子":
        components, _ = derive_components(data)
        for m in components:
            rohs = str(m.get("rohs") or "").strip()
            if not rohs:
                warnings.append(
                    "WARNING: 物料『%s』未标注 RoHS 合规状态，请确认"
                    % m.get("name", "")
                )
    elif industry == "化工":
        formula, _ = derive_formula(data)
        for m in formula:
            cas = str(m.get("cas_number") or "").strip()
            ghs = str(m.get("ghs_hazard") or "").strip()
            if not cas:
                warnings.append(
                    "WARNING: 物料『%s』未填写 CAS 号，请确认"
                    % m.get("name", "")
                )
            if not ghs:
                warnings.append(
                    "WARNING: 物料『%s』未填写 GHS 危险标识，请确认"
                    % m.get("name", "")
                )
        # 含量(%) 列和校验：仅当所有配方原料均填了 concentration（非空且 > 0）时才校验
        concs = [float(m.get("concentration") or 0) for m in formula]
        if concs and all(c > 0 for c in concs):
            total = sum(concs)
            if abs(total - 100.0) > 5.0:
                warnings.append(
                    "WARNING: 配方表含量(%%) 列和为 %.1f%%，偏离 100%% 超过 ±5%%，请确认"
                    % total
                )
    return warnings


def _build_meta_line(approver, effective_date, standard):
    """拼接行 5 表头区文本：仅拼接非空字段，段间用 4 个空格分隔。

    三者皆空返回空串（等价于 V2 空行，不挤占行 6）。
    """
    parts = []
    if approver:
        parts.append("审批人：" + str(approver))
    if effective_date:
        parts.append("生效日期：" + str(effective_date))
    if standard:
        parts.append("执行标准：" + str(standard))
    return "    ".join(parts)


def _build_meta_line_en(approver, effective_date, standard):
    """拼接行 5 表头区英文文本（标签英文，standard 代号保留原值）。"""
    parts = []
    if approver:
        parts.append("Approver: " + str(approver))
    if effective_date:
        parts.append("Effective Date: " + str(effective_date))
    if standard:
        parts.append("Executive Standard: " + str(standard))
    return "    ".join(parts)


# ---------------------------------------------------------------------------
# V7: 样式工厂 + 双语渲染 helper
# ---------------------------------------------------------------------------

def _make_styles():
    """构建 openpyxl 样式对象（懒加载，避免无 openpyxl 时模块导入失败）。

    Returns:
        SimpleNamespace，含 title_font / head_fill / head_font / label_font /
        cell_font / rohs_font_red / rohs_font_yellow / group_fill / border /
        group_border / center / left / title_align / pct_fmt。
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name="微软雅黑", size=16, bold=True, color="1F3864")
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    head_font = Font(name="微软雅黑", size=11, bold=True, color="1F3864")
    label_font = Font(name="微软雅黑", size=10, bold=True)
    cell_font = Font(name="微软雅黑", size=10)
    # V4: RoHS 合规标记字体（红色=不合规，黄色=待确认）
    rohs_font_red = Font(name="微软雅黑", size=10, color="FF0000")
    rohs_font_yellow = Font(name="微软雅黑", size=10, color="BF8F00")
    group_fill = PatternFill("solid", fgColor="EAF1FB")
    bar_side = Side(style="medium", color="1F3864")
    group_border = Border(left=bar_side, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # 标题单元格对齐（与 V6 一致：无 wrap_text）
    title_align = Alignment(horizontal="center", vertical="center")
    pct_fmt = '0.0"%"'
    return SimpleNamespace(
        title_font=title_font,
        head_fill=head_fill,
        head_font=head_font,
        label_font=label_font,
        cell_font=cell_font,
        rohs_font_red=rohs_font_red,
        rohs_font_yellow=rohs_font_yellow,
        group_fill=group_fill,
        border=border,
        group_border=group_border,
        center=center,
        left=left,
        title_align=title_align,
        pct_fmt=pct_fmt,
    )


def _write_head_cell(ws, r, col, text, styles):
    """写表头单元格（V6 既有样式，zh/en 共用）。"""
    cell = ws.cell(r, col, text)
    cell.font = styles.head_font
    cell.fill = styles.head_fill
    cell.alignment = styles.center
    cell.border = styles.border
    return cell


def _write_block(ws, r, marker_zh, headers_zh, lang, styles, max_col=8):
    """渲染一个区块的标题行 + 表头行（中文 / 或中英双行）。

    - lang="zh"：单表头行（与 V6 字节级一致）。
    - lang="en"：块标题「中文 (English)」合并行 + 中文表头行 + 英文表头行。

    Args:
        ws: openpyxl Worksheet。
        r: 区块起始行号（1-based）。
        marker_zh: 区块中文标题。
        headers_zh: 表头中文列表（已按列序，长度=max_col）。
        lang: "zh" / "en"。
        styles: _make_styles() 结果。
        max_col: 列数（电子 14 / 化工 13 / 其余 8），用于 marker 合并范围。

    Returns:
        下一空白行号（表头行之后）。
    """
    en_marker = ZH2EN.get(marker_zh, marker_zh)
    mrange = "A:%s" % chr(64 + max_col)
    m_start, m_end = mrange.split(":")
    if lang == "en":
        ws.merge_cells("%s%d:%s%d" % (m_start, r, m_end, r))
        ws.cell(r, 1, "%s (%s)" % (marker_zh, en_marker)).font = styles.label_font
        r += 1
        for col, h in enumerate(headers_zh, 1):
            _write_head_cell(ws, r, col, h, styles)
        r += 1
        for col, h in enumerate(headers_zh, 1):
            _write_head_cell(ws, r, col, ZH2EN.get(h, h), styles)
        r += 1
    else:
        ws.merge_cells("%s%d:%s%d" % (m_start, r, m_end, r))
        ws.cell(r, 1, marker_zh).font = styles.label_font
        r += 1
        for col, h in enumerate(headers_zh, 1):
            _write_head_cell(ws, r, col, h, styles)
        r += 1
    return r


def _translate_cell(lang, header_zh, val):
    """双语 sheet 中，物料类型列的值按 ZH2EN 显示英文（如 型材→Profile）。

    - lang="zh" 或 非「物料类型」列：原值返回。
    - lang="en" 且 列名为「物料类型」且值命中 I18N：返回英文；否则原值。
    """
    if lang == "en" and header_zh == "物料类型":
        t = ZH2EN.get(val)
        if t:
            return t
    return val


def build_workbook(data, industry=None, bilingual=False):
    """根据数据构建 BOM 表 Workbook。

    - bilingual=False（默认）：单 sheet「BOM表」（与 V6 字节级一致）。
    - bilingual=True：追加「BOM表(英)」双语 sheet（中英双行表头，结构同中文）。

    Args:
        data: BOM JSON dict。
        industry: 已推断的行业字符串；None 时内部调 infer_industry。
        bilingual: 是否追加双语 sheet。
    """
    from openpyxl import Workbook

    if industry is None:
        industry, _ = infer_industry(data)

    styles = _make_styles()
    wb = Workbook()
    zh_ws = wb.active
    zh_ws.title = "BOM表"
    _render_sheet(zh_ws, data, industry, "zh", styles)  # 与 V6 同代码路径 → 字节级不变

    if bilingual:
        en_ws = wb.create_sheet("BOM表(英)")
        _render_sheet(en_ws, data, industry, "en", styles)  # 复用同一份 data，仅表头双语化

    return wb


def _render_sheet(ws, data, industry, lang, styles):
    """承载原 build_workbook 全部渲染体（抽出，按 lang 分支表头/marker）。

    行 1 标题：lang="en" 时写 "BOM表 (BOM Table)"，否则 "BOM表"。
    行 2–5 表头区：标签（版本号/生成日期/产品名称/产品类别/全产品出品率/
        审批人/生效日期/执行标准）在 lang="en" 时翻译标签词（Executive Standard 等），
        但 standard 单元格**值保留原代号**（不翻译）。
    物料区/工序区/各派生视图/成本：均经 _write_block(..., lang) 渲染；数据行
    在 zh/en 两 sheet 完全一致（material_type 列在 en 下按 ZH2EN 显示英文）。

    Args:
        ws: openpyxl Worksheet（中文 sheet 或 双语 sheet）。
        data: BOM JSON dict。
        industry: 已推断的行业字符串。
        lang: "zh" / "en"。
        styles: _make_styles() 结果。
    """
    version = data.get("version") or "V1.0"
    date = data.get("date") or datetime.date.today().isoformat()
    product_name = str(data.get("product_name") or "").strip()
    category = str(data.get("category") or "其他").strip()
    output_rate = data.get("output_rate")
    approver = str(data.get("approver") or "").strip()
    effective_date = str(data.get("effective_date") or "").strip()
    standard = str(data.get("standard") or "").strip()

    # 行1：标题
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "BOM表" if lang == "zh" else "BOM表 (BOM Table)"
    c.font = styles.title_font
    c.alignment = styles.title_align
    ws.row_dimensions[1].height = 28

    # 行2：版本号（左） / 生成日期（右）
    ws["A2"] = ("版本号：%s" % version) if lang == "zh" else ("Version: %s" % version)
    ws["A2"].font = styles.label_font
    ws["D2"] = ("生成日期：%s" % date) if lang == "zh" else ("Date: %s" % date)
    ws["D2"].font = styles.label_font
    ws.merge_cells("A2:C2")
    ws.merge_cells("D2:H2")

    # 行3：产品名称（整行）
    ws.merge_cells("A3:H3")
    c = ws.cell(3, 1, ("产品名称：%s" % product_name) if lang == "zh"
                else ("Product Name: %s" % product_name))
    c.font = styles.label_font

    # 行4：产品类别（左） / 全产品出品率（右，0.0"%" 格式）
    ws.merge_cells("A4:C4")
    c = ws.cell(4, 1, ("产品类别：%s" % category) if lang == "zh"
                else ("Category: %s" % category))
    c.font = styles.label_font
    ws.merge_cells("D4:H4")
    c = ws.cell(4, 4, ("全产品出品率：%.1f%%" % float(output_rate)) if lang == "zh"
                else ("Overall Yield: %.1f%%" % float(output_rate)))
    c.font = styles.label_font

    # 行5：审批人 / 生效日期 / 执行标准（可选，单格合并 A5:H5；皆空则留空）
    meta_line = (_build_meta_line(approver, effective_date, standard)
                 if lang == "zh"
                 else _build_meta_line_en(approver, effective_date, standard))
    ws.merge_cells("A5:H5")
    c = ws.cell(5, 1, meta_line)
    c.font = styles.label_font
    c.alignment = styles.left

    # 行6：一、物料信息
    r = 6
    material_headers = [
        "序号", "物料名称", "单位", "用量", "出品率(%)",
        "ERP物料代码", "物料类型", "所属工序",
    ]
    r = _write_block(ws, r, "一、物料信息", material_headers, lang, styles, max_col=8)

    # 物料数据（按需分工序分组）
    materials = data.get("materials", [])
    processes = data.get("processes", [])
    valid_step_nos = {str(p.get("step_no") or "").strip() for p in processes}

    def write_material_row(m, r, seq):
        """写入一行物料数据（含首列序号），返回下一行号。"""
        raw_vals = [
            seq,
            m.get("name", ""),
            m.get("unit", ""),
            m.get("usage", ""),
            m.get("yield_rate", ""),
            m.get("erp_code", ""),
            m.get("material_type", ""),
            m.get("process", ""),
        ]
        # 列对齐：序号/单位/用量/出品率/所属工序居中；物料名称/ERP/物料类型左对齐
        aligns = [styles.center, styles.left, styles.center, styles.center,
                  styles.center, styles.left, styles.left, styles.center]
        for col, val in enumerate(raw_vals, 1):
            disp = _translate_cell(lang, material_headers[col - 1], val)
            cell = ws.cell(r, col, disp)
            cell.font = styles.cell_font
            cell.border = styles.border
            cell.alignment = aligns[col - 1]
            if col == 5:  # 出品率(%)
                cell.number_format = styles.pct_fmt
        return r + 1

    def write_group_subtitle(text, r):
        """写入分组子标题（A–H 合并，浅蓝底 + 左侧加粗色条 + 加粗），返回下一行号。"""
        ws.merge_cells("A%d:H%d" % (r, r))
        c = ws.cell(r, 1, text)
        c.font = styles.label_font
        c.fill = styles.group_fill
        c.alignment = styles.left
        c.border = styles.group_border
        return r + 1

    group_enabled = bool(processes) and any(
        str(m.get("process") or "").strip() in valid_step_nos for m in materials
    )
    seq = 0
    if group_enabled:
        attributed = {str(p.get("step_no") or "").strip(): [] for p in processes}
        unattributed = []
        for m in materials:
            sn = str(m.get("process") or "").strip()
            if sn in attributed:
                attributed[sn].append(m)
            else:
                unattributed.append(m)
        for p in processes:
            sn = str(p.get("step_no") or "").strip()
            group = attributed.get(sn, [])
            if group:
                r = write_group_subtitle("【工序 %s %s】" % (sn, p.get("name", "")), r)
                for m in group:
                    seq += 1
                    r = write_material_row(m, r, seq)
        if unattributed:
            r = write_group_subtitle("【未归属工序】", r)
            for m in unattributed:
                seq += 1
                r = write_material_row(m, r, seq)
    else:
        for m in materials:
            seq += 1
            r = write_material_row(m, r, seq)

    # 物料区末「合计用量」行：A=合计，D=全部物料 usage 求和（含包材/其他）
    total_usage = sum(float(m.get("usage") or 0) for m in materials)
    ws.cell(r, 1, "合计" if lang == "zh" else ZH2EN.get("合计", "Total")).font = styles.label_font
    ws.cell(r, 1).alignment = styles.left
    ws.cell(r, 1).border = styles.border
    tc = ws.cell(r, 4, round(total_usage, 1))
    tc.font = styles.label_font
    tc.alignment = styles.center
    tc.border = styles.border
    for col in (2, 3, 5, 6, 7, 8):
        ws.cell(r, col).border = styles.border
    r += 1
    r += 1  # 物料区后空行

    # 二、工艺工序
    process_headers = ["工序编号", "工序名称", "工序说明", "工时", "备注", "产物"]
    r = _write_block(ws, r, "二、工艺工序", process_headers, lang, styles, max_col=8)
    for p in processes:
        row_vals = [
            p.get("step_no", ""),
            p.get("name", ""),
            p.get("desc", ""),
            p.get("work_hours", ""),
            p.get("note", ""),
            p.get("output", ""),
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(r, col, val)
            cell.font = styles.cell_font
            cell.border = styles.border
            cell.alignment = styles.left if col in (3, 5) else styles.center
        r += 1

    # ===== V4: 工序区后按 industry 分支追加专属派生视图 =====

    if industry == "食品":
        # 三、配料表（沿用 V3 逻辑，触发改 industry）
        r += 1
        r = _write_block(ws, r, "三、配料表", [
            "物料名称", "物料类型", "计量单位", "用量", "出品率(%)", "用量占比%", "过敏原",
        ], lang, styles, max_col=8)
        ingredients, _ = derive_ingredients(data, industry)
        pct_list, _ = ingredient_pct(ingredients)
        for idx, m in enumerate(ingredients):
            row_vals = [
                m.get("name", ""),
                m.get("material_type", ""),
                m.get("unit", ""),
                m.get("usage", ""),
                m.get("yield_rate", ""),
                pct_list[idx],
                m.get("allergen", ""),
            ]
            # 对齐：物料名称/物料类型/过敏原左对齐；其余居中
            aligns = [styles.left, styles.left, styles.center, styles.center,
                      styles.center, styles.center, styles.left]
            for col, val in enumerate(row_vals, 1):
                disp = _translate_cell(lang, "物料类型" if col == 2 else "", val)
                cell = ws.cell(r, col, disp)
                cell.font = styles.cell_font
                cell.border = styles.border
                cell.alignment = aligns[col - 1]
                if col in (5, 6):  # 出品率(%) 与 用量占比%
                    cell.number_format = styles.pct_fmt
            r += 1

    elif industry == "电子":
        # 三、元件清单（★V4 新增，★V5 扩列至 14 列 A–N）
        r += 1
        r = _write_block(ws, r, "三、元件清单", [
            "序号", "位号(Designator)", "型号(Part#)", "封装(Footprint)", "物料名称",
            "数量", "物料类型", "RoHS", "制造商", "容差", "额定功率", "额定电压",
            "替代料", "封装温度",
        ], lang, styles, max_col=14)
        components, _ = derive_components(data)
        for idx, m in enumerate(components, 1):
            rohs_val = str(m.get("rohs") or "").strip()
            # RoHS 着色规则：否→红字，未知/空→黄字，是→默认
            if rohs_val == "否":
                rohs_font = styles.rohs_font_red
            elif rohs_val == "是":
                rohs_font = styles.cell_font
            else:
                rohs_font = styles.rohs_font_yellow

            row_vals = [
                idx,
                m.get("designator", ""),
                m.get("part_number", ""),
                m.get("footprint", ""),
                m.get("name", ""),
                m.get("usage", ""),
                m.get("material_type", ""),
                rohs_val,
                m.get("manufacturer", ""),
                m.get("tolerance", ""),
                m.get("rated_power", ""),
                m.get("rated_voltage", ""),
                m.get("alternate", ""),
                m.get("reflow_temp", ""),
            ]
            # 对齐：序号/位号/封装/数量/物料类型/RoHS/容差/额定功率/额定电压/替代料/封装温度 居中；
            # 型号/物料名称/制造商 左对齐
            aligns = [
                styles.center, styles.center, styles.left, styles.center, styles.left,
                styles.center, styles.center, styles.center, styles.left, styles.center,
                styles.center, styles.center, styles.left, styles.center,
            ]
            for col, val in enumerate(row_vals, 1):
                disp = _translate_cell(lang, "物料类型" if col == 7 else "", val)
                cell = ws.cell(r, col, disp)
                cell.border = styles.border
                cell.alignment = aligns[col - 1]
                if col == 8:  # RoHS 列使用特定字体
                    cell.font = rohs_font
                else:
                    cell.font = styles.cell_font
            r += 1

    elif industry == "化工":
        # 三、配方表（★V4 新增，★V5 扩列至 13 列 A–M）
        r += 1
        r = _write_block(ws, r, "三、配方表", [
            "序号", "物料名称", "CAS号", "含量(%)", "GHS标识", "物料类型",
            "计量单位", "用量", "纯度", "物态", "闪点", "存储条件", "危险等级",
        ], lang, styles, max_col=13)
        formula, _ = derive_formula(data)
        for idx, m in enumerate(formula, 1):
            conc_raw = m.get("concentration", "")
            # 含量(%)：数值则写入 float + 0.0"%" 格式；空则留空
            if conc_raw != "" and conc_raw is not None:
                try:
                    conc_val = float(conc_raw)
                except (TypeError, ValueError):
                    conc_val = str(conc_raw)
            else:
                conc_val = ""

            row_vals = [
                idx,
                m.get("name", ""),
                m.get("cas_number", ""),
                conc_val,
                m.get("ghs_hazard", ""),
                m.get("material_type", ""),
                m.get("unit", ""),
                m.get("usage", ""),
                m.get("purity", ""),
                m.get("physical_state", ""),
                m.get("flash_point", ""),
                m.get("storage_condition", ""),
                m.get("hazard_class", ""),
            ]
            # 对齐：序号/CAS号/含量/GHS/物料类型/计量单位/用量/纯度/物态/闪点/危险等级 居中；
            # 物料名称/存储条件 左对齐
            aligns = [
                styles.center, styles.left, styles.center, styles.center, styles.left,
                styles.center, styles.center, styles.center, styles.center, styles.center,
                styles.center, styles.left, styles.center,
            ]
            for col, val in enumerate(row_vals, 1):
                disp = _translate_cell(lang, "物料类型" if col == 6 else "", val)
                cell = ws.cell(r, col, disp)
                cell.font = styles.cell_font
                cell.border = styles.border
                cell.alignment = aligns[col - 1]
                if col == 4 and conc_val != "":  # 含量(%) 数字格式
                    cell.number_format = styles.pct_fmt
            r += 1

    elif industry == "纺织":
        # 三、面料辅料清单（★V5 新增，8 列 A–H）
        r += 1
        r = _write_block(ws, r, "三、面料辅料清单", [
            "序号", "物料名称", "物料类型", "成分比例", "纱支",
            "克重(g/m²)", "幅宽", "色号",
        ], lang, styles, max_col=8)
        textile_items, _ = derive_textile(data)
        for idx, m in enumerate(textile_items, 1):
            row_vals = [
                idx,
                m.get("name", ""),
                m.get("material_type", ""),
                m.get("composition", ""),
                m.get("yarn_count", ""),
                m.get("fabric_weight", ""),
                m.get("width", ""),
                m.get("color_no", ""),
            ]
            # 对齐：序号/物料类型/纱支/克重/幅宽/色号 居中；物料名称/成分比例 左对齐
            aligns = [styles.center, styles.left, styles.center, styles.left,
                      styles.center, styles.center, styles.center, styles.center]
            for col, val in enumerate(row_vals, 1):
                disp = _translate_cell(lang, "物料类型" if col == 3 else "", val)
                cell = ws.cell(r, col, disp)
                cell.font = styles.cell_font
                cell.border = styles.border
                cell.alignment = aligns[col - 1]
            r += 1

    elif industry == "家具":
        # 三、家具物料清单（★V5 新增，8 列 A–H）
        r += 1
        r = _write_block(ws, r, "三、家具物料清单", [
            "序号", "物料名称", "物料类型", "材质等级", "尺寸规格",
            "表面处理", "用量", "色号/花色",
        ], lang, styles, max_col=8)
        furniture_items, _ = derive_furniture(data)
        for idx, m in enumerate(furniture_items, 1):
            row_vals = [
                idx,
                m.get("name", ""),
                m.get("material_type", ""),
                m.get("material_grade", ""),
                m.get("spec_size", ""),
                m.get("surface_treatment", ""),
                m.get("usage", ""),
                m.get("color_no", ""),
            ]
            # 对齐：序号/物料类型/材质等级/用量/色号 居中；物料名称/尺寸规格/表面处理 左对齐
            aligns = [styles.center, styles.left, styles.center, styles.center,
                      styles.left, styles.left, styles.center, styles.center]
            for col, val in enumerate(row_vals, 1):
                disp = _translate_cell(lang, "物料类型" if col == 3 else "", val)
                cell = ws.cell(r, col, disp)
                cell.font = styles.cell_font
                cell.border = styles.border
                cell.alignment = aligns[col - 1]
            r += 1

    elif industry == "机械":
        # 三、机械物料清单（★V6 新增，8 列 A–H，不含「物料类型」展示列）
        r += 1
        r = _write_block(ws, r, "三、机械物料清单", [
            "序号", "物料名称", "图号", "材质", "热处理", "表面处理",
            "重量(kg)", "单重(kg/件)",
        ], lang, styles, max_col=8)
        mech_items, _ = derive_mechanical(data)
        for idx, m in enumerate(mech_items, 1):
            row_vals = [
                idx,
                m.get("name", ""),
                m.get("drawing_no", ""),
                m.get("material", ""),
                m.get("heat_treatment", ""),
                m.get("surface_treatment", ""),
                m.get("weight", ""),
                m.get("unit_weight", ""),
            ]
            # 对齐：序号/图号/材质/热处理/表面处理/重量/单重 居中；物料名称 左对齐
            aligns = [styles.center, styles.left, styles.center, styles.left,
                      styles.center, styles.center, styles.center, styles.center]
            for col, val in enumerate(row_vals, 1):
                cell = ws.cell(r, col, val)
                cell.font = styles.cell_font
                cell.border = styles.border
                cell.alignment = aligns[col - 1]
            r += 1

    elif industry == "包装":
        # 三、包装物料清单（★V6 新增，8 列 A–H，保留「物料类型」展示列）
        r += 1
        r = _write_block(ws, r, "三、包装物料清单", [
            "序号", "物料名称", "物料类型", "材质", "克重(g/m²)", "尺寸",
            "印刷工艺", "环保标识",
        ], lang, styles, max_col=8)
        pack_items, _ = derive_packaging(data)
        for idx, m in enumerate(pack_items, 1):
            row_vals = [
                idx,
                m.get("name", ""),
                m.get("material_type", ""),
                m.get("material", ""),
                m.get("basis_weight", ""),
                m.get("size", ""),
                m.get("print_process", ""),
                m.get("eco_label", ""),
            ]
            # 对齐：序号/物料类型/材质/克重/印刷工艺/环保标识 居中；物料名称/尺寸 左对齐
            aligns = [styles.center, styles.left, styles.center, styles.left,
                      styles.center, styles.left, styles.center, styles.center]
            for col, val in enumerate(row_vals, 1):
                disp = _translate_cell(lang, "物料类型" if col == 3 else "", val)
                cell = ws.cell(r, col, disp)
                cell.font = styles.cell_font
                cell.border = styles.border
                cell.alignment = aligns[col - 1]
            r += 1

    # 其他行业（通用）不生成「三、」行业专属视图

    # ===== V5: 跨行业成本明细视图（双编号） =====
    cost_items, has_cost = derive_cost(data)
    if has_cost:
        # 有行业视图（食品/电子/化工/纺织/家具）→ 四、成本明细；否则 三、成本明细
        cost_label = "四、成本明细" if industry in INDUSTRY_VIEW_SET else "三、成本明细"
        r += 1
        r = _write_block(ws, r, cost_label, [
            "序号", "物料名称", "物料类型", "用量", "单位", "单价", "币种", "总价",
        ], lang, styles, max_col=8)
        cost_total = 0.0
        for idx, m in enumerate(cost_items, 1):
            up_raw = m.get("unit_price", "")
            usage_raw = m.get("usage", "")
            try:
                up_f = float(up_raw) if up_raw not in ("", None) else 0.0
                usage_f = float(usage_raw) if usage_raw not in ("", None) else 0.0
                total = round(up_f * usage_f, 2)
            except (TypeError, ValueError):
                up_f = 0.0
                total = ""
            if isinstance(total, (int, float)):
                cost_total += total
            currency_val = str(m.get("currency") or "").strip() or DEFAULT_CURRENCY
            row_vals = [
                idx,
                m.get("name", ""),
                m.get("material_type", ""),
                m.get("usage", ""),
                m.get("unit", ""),
                up_f if isinstance(up_f, (int, float)) else "",
                currency_val,
                total,
            ]
            # 对齐：序号/物料类型/用量/单位/单价/总价 居中；物料名称 左对齐；币种 左对齐
            aligns = [styles.center, styles.left, styles.center, styles.center,
                      styles.center, styles.center, styles.left, styles.center]
            for col, val in enumerate(row_vals, 1):
                disp = _translate_cell(lang, "物料类型" if col == 3 else "", val)
                cell = ws.cell(r, col, disp)
                cell.font = styles.cell_font
                cell.border = styles.border
                cell.alignment = aligns[col - 1]
                if col in (6, 8) and isinstance(val, (int, float)):
                    cell.number_format = "0.00"  # 单价/总价 数值格式
            r += 1
        # 成本合计行：A="成本合计"，H=Σ总价（纯展示，逆向跳过）
        ws.cell(r, 1, "成本合计" if lang == "zh"
                else ZH2EN.get("成本合计", "Cost Total")).font = styles.label_font
        ws.cell(r, 1).alignment = styles.left
        ws.cell(r, 1).border = styles.border
        tc = ws.cell(r, 8, round(cost_total, 2))
        tc.font = styles.label_font
        tc.alignment = styles.center
        tc.border = styles.border
        for col in (2, 3, 4, 5, 6, 7):
            ws.cell(r, col).border = styles.border
        r += 1

    # 列宽（依据 industry 与扩列扩展）
    # 电子扩列到 14 列（A–N），化工扩列到 13 列（A–M），其余 8 列（A–H）
    if industry == "电子":
        widths = [6, 18, 18, 12, 18, 10, 13, 10, 14, 12, 14, 14, 16, 14]
    elif industry == "化工":
        widths = [6, 18, 12, 13, 14, 13, 12, 10, 10, 12, 12, 18, 14]
    elif industry == "机械":
        # V6：8 列微调（C=图号16，G/H=重量/单重 各12）
        widths = [6, 18, 16, 12, 13, 13, 12, 12]
    elif industry == "包装":
        # V6：8 列微调（E=克重12，F=尺寸18）
        widths = [6, 18, 10, 12, 12, 18, 13, 13]
    else:
        widths = [6, 18, 10, 10, 13, 16, 13, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w


# ---------------------------------------------------------------------------
# V7: B1 批量空白模板 / B2 批量生成 BOM
# ---------------------------------------------------------------------------

def _blank_data(industry):
    """构造某行业的空白 BOM 数据（不触发校验，可被 import_bom 解析为空 JSON）。"""
    if industry == "食品":
        category = "食品"
    elif industry == "通用":
        category = "其他"
    else:
        category = "工业品"
    return {
        "product_name": "",
        "category": category,
        "industry": industry,
        "output_rate": 100.0,
        "version": "V1.0",
        "date": "",
        "materials": [],
        "processes": [],
    }


def _safe_filename(product_name, date, index):
    """将产品名转为合法文件名（替换非法字符），空名回落序号。"""
    illegal = set('\\/:*?"<>|')
    cleaned = "".join(c if c not in illegal else "_" for c in product_name).strip()
    if not cleaned:
        cleaned = str(index)
    return "BOM_%s_%s.xlsx" % (cleaned, date)


def run_blank_templates(out_dir, industries=None, bilingual=False):
    """B1：为各行业生成空白 BOM 模板（template_<行业>.xlsx）。

    Args:
        out_dir: 输出目录（自动创建）。
        industries: 行业列表（None → 全部 8 行业）。
        bilingual: 是否生成含双语 sheet 的模板。
    """
    os.makedirs(out_dir, exist_ok=True)
    inds = industries if industries else ALL_INDUSTRIES
    count = 0
    for ind in inds:
        data = _blank_data(ind)
        wb = build_workbook(data, industry=ind, bilingual=bilingual)
        path = os.path.join(out_dir, "template_%s.xlsx" % ind)
        wb.save(path)
        count += 1
    print("OK: 成功生成 %d 个空白模板（含行业：%s）" % (count, "、".join(inds)))
    return count


def run_batch_generate(batch=None, batch_dir=None, out_dir=".", bilingual=False):
    """B2：批量生成 BOM（错误隔离，结束汇总成功/失败）。

    Args:
        batch: 显式逗号分隔文件路径列表（与 batch_dir 取并集）。
        batch_dir: 读取该目录下所有 *.json。
        out_dir: 输出目录（默认当前目录）。
        bilingual: 是否双语导出。

    Returns:
        失败文件数（>0 时调用方应以退出码 2 结束）。
    """
    os.makedirs(out_dir, exist_ok=True)
    files = []
    if batch_dir:
        files.extend(sorted(glob.glob(os.path.join(batch_dir, "*.json"))))
    if batch:
        files.extend([f.strip() for f in batch.split(",") if f.strip()])
    # 去重保序
    seen = set()
    unique_files = []
    for f in files:
        if f not in seen:
            seen.add(f)
            unique_files.append(f)
    files = unique_files

    success = 0
    failures = []
    for f in files:
        try:
            data = load_data(f)
        except Exception as e:  # noqa: BLE001
            failures.append((f, ["文件读取失败：%s" % e]))
            continue
        errors = validate(data)
        if errors:
            failures.append((f, errors))
            continue
        try:
            wb = build_workbook(data, bilingual=bilingual)
            product_name = str(data.get("product_name") or "").strip()
            date = str(data.get("date") or datetime.date.today().isoformat()).strip()
            fname = _safe_filename(product_name, date, success + 1)
            out_path = os.path.join(out_dir, fname)
            wb.save(out_path)
            success += 1
        except Exception as e:  # noqa: BLE001
            failures.append((f, ["生成失败：%s" % e]))
            continue

    print("批量生成完成：成功 %d / 失败 %d" % (success, len(failures)))
    for f, errs in failures:
        print(" - 失败 %s：" % f)
        for e in errs:
            print("   " + e)
    return len(failures)


def main():
    ensure_openpyxl()
    parser = argparse.ArgumentParser(description="生成 BOM 表 Excel")
    parser.add_argument("--data", help="物料/工序 JSON 文件路径（单条生成模式）")
    parser.add_argument("--out", help="输出 xlsx 路径（单条生成模式）")
    parser.add_argument("--bilingual", action="store_true",
                        help="追加「BOM表(英)」双语 sheet（中英双行表头）")
    parser.add_argument("--blank-templates", action="store_true",
                        help="B1：批量生成各行业空白模板 template_<行业>.xlsx")
    parser.add_argument("--out-dir", default=".", help="批量产物输出目录（默认当前目录）")
    parser.add_argument("--industries", default=None,
                        help="B1 限定行业（逗号分隔，默认全部 8 行业）")
    parser.add_argument("--batch-dir", default=None, help="B2：读取目录下 *.json 批量生成")
    parser.add_argument("--batch", default=None, help="B2：显式逗号分隔文件列表批量生成")
    args = parser.parse_args()

    # —— B1 批量空白模板 ——
    if args.blank_templates:
        inds = None
        if args.industries:
            inds = [x.strip() for x in args.industries.split(",") if x.strip()]
        run_blank_templates(args.out_dir, inds, args.bilingual)
        return

    # —— B2 批量生成 ——
    if args.batch_dir or args.batch:
        failures = run_batch_generate(
            batch=args.batch, batch_dir=args.batch_dir,
            out_dir=args.out_dir, bilingual=args.bilingual,
        )
        if failures > 0:
            sys.exit(2)
        return

    # —— 单条生成（V6 行为零变化） ——
    if not args.data or not args.out:
        print("USAGE_ERROR: 单条生成模式需要 --data 与 --out；"
              "批量模式使用 --blank-templates / --batch-dir / --batch")
        sys.exit(2)

    data = load_data(args.data)
    errors = validate(data)
    if errors:
        print("VALIDATION_FAILED")
        for e in errors:
            print(" - " + e)
        sys.exit(2)

    # V4: 推断 industry
    industry, v8_warnings = infer_industry(data)
    for w in v8_warnings:
        print(w)

    # 行业专属软校验 + 排除提示（非阻断 WARNING）
    if industry == "食品":
        # W1/H1 过敏原软校验
        for w in check_allergen_soft(data):
            print(w)
        # 配料表排除提示
        _, excluded = derive_ingredients(data, industry)
        if excluded:
            names = [str(m.get("name") or "") for m in excluded]
            print(
                "WARNING: 配料表已排除 %d 条非食用物料（包材/其他/未分类）：%s"
                % (len(excluded), "、".join(names))
            )
    elif industry == "电子":
        # W2 RoHS 软校验
        for w in check_industry_soft(data, industry):
            print(w)
        # 元件清单排除提示
        _, excluded = derive_components(data)
        if excluded:
            names = [str(m.get("name") or "") for m in excluded]
            print(
                "WARNING: 元件清单已排除 %d 条非元件物料（其他类）：%s"
                % (len(excluded), "、".join(names))
            )
    elif industry == "化工":
        # W3 CAS/GHS 软校验 + 含量和校验
        for w in check_industry_soft(data, industry):
            print(w)
        # 配方表排除提示
        _, excluded = derive_formula(data)
        if excluded:
            names = [str(m.get("name") or "") for m in excluded]
            print(
                "WARNING: 配方表已排除 %d 条包材物料：%s"
                % (len(excluded), "、".join(names))
            )
    elif industry == "纺织":
        # V5: 面料辅料清单排除提示（沿用 V4 排除提示模式，文案同构）
        _, excluded = derive_textile(data)
        if excluded:
            names = [str(m.get("name") or "") for m in excluded]
            print(
                "WARNING: 面料辅料清单已排除 %d 条其他类物料：%s"
                % (len(excluded), "、".join(names))
            )
    elif industry == "家具":
        # V5: 家具物料清单排除提示（沿用 V4 排除提示模式，文案同构）
        _, excluded = derive_furniture(data)
        if excluded:
            names = [str(m.get("name") or "") for m in excluded]
            print(
                "WARNING: 家具物料清单已排除 %d 条其他类物料：%s"
                % (len(excluded), "、".join(names))
            )
    elif industry == "机械":
        # V6: 机械物料清单排除提示（沿用 V4 排除提示模式，文案同构）
        _, excluded = derive_mechanical(data)
        if excluded:
            names = [str(m.get("name") or "") for m in excluded]
            print(
                "WARNING: 机械物料清单已排除 %d 条其他类物料：%s"
                % (len(excluded), "、".join(names))
            )
    elif industry == "包装":
        # V6: 包装物料清单排除提示（沿用 V4 排除提示模式，文案同构）
        _, excluded = derive_packaging(data)
        if excluded:
            names = [str(m.get("name") or "") for m in excluded]
            print(
                "WARNING: 包装物料清单已排除 %d 条其他类物料：%s"
                % (len(excluded), "、".join(names))
            )

    wb = build_workbook(data, industry, bilingual=args.bilingual)
    wb.save(args.out)
    print("OK:" + args.out)


if __name__ == "__main__":
    main()
