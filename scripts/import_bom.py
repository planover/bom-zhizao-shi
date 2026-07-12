#!/usr/bin/env python3
"""
import_bom.py - BOM智造师 配套逆向导入脚本（V5 / V4 / V3 / V2.1）

读取由 generate_bom.py 生成的 BOM 表 Excel (.xlsx)，反向解析为与正向
输入 JSON Schema 完全一致的结构化数据，便于「重新编辑已有 BOM」或闭环回写。

V5 增强（相对 V4）：
- 识别「三、面料辅料清单」/「三、家具物料清单」区块标记 → 按物料名回收
  纺织(5)/家具(4) 专属字段；识别「成本明细」关键字 → 回收 unit_price/currency。
- 电子「三、元件清单」区块扩列回收 10 字段（原 4 + 新 6）；
  化工「三、配方表」区块扩列回收 8 字段（原 3 + 新 5）。
- 推断 industry 增「三、面料辅料清单」→纺织、「三、家具物料清单」→家具。
- 物料对象专属字段扩至 28 个唯一 JSON 键（设计文档按行业叠加计为 29 概念字段，
  其中 color_no 在纺织/家具共用同一 JSON key，故唯一键为 28）。

V6 增强（相对 V5）：
- 识别「三、机械物料清单」/「三、包装物料清单」区块标记 → 按物料名回收
  机械(6)/包装(5) 专属字段；机械/包装带成本仍为「成本明细」关键字回收（沿用 V5）。
- 推断 industry 增「三、机械物料清单」→机械、「三、包装物料清单」→包装。
- 物料对象专属字段扩至 37 个唯一 JSON 键（V5 的 28 唯一键 + V6 净增 9
  唯一键；其中 material 与包装/机械同名、surface_treatment 与家具同名各只计一次）。

V4 增强（相对 V3）：
- 识别「三、元件清单」/「三、配方表」区块标记 → 按物料名回收电子/化工专属字段。
- 推断 industry：从「三、」区块标记推断（有元件清单→电子，有配方表→化工，
  有配料表→食品），无则按 category 推断。
- 输出 JSON 增 `industry` 字段；物料对象增 7 个专属字段默认空串。
- 旧 Excel（无 industry / 无专属区块）完全兼容。

V3（V2.1）能力（沿用）：
- 按列头文本映射列号（不硬编码 A–E），对旧版 5/7 列 Excel 天然兼容。
- 解析 category / output_rate / material_type / process / output /
  approver / effective_date / standard / materials[].allergen。
- 「三、配料表」仅回收「过敏原」列（按物料名称匹配回写）。
- 跳过「序号」「用量占比%」列；跳过物料区「合计用量」行。

依赖 openpyxl；若运行环境缺失则自动安装（复用 ensure_openpyxl 逻辑）。

用法:
    python3 import_bom.py --in BOM_2026-07-07.xlsx [--out data.json]
"""
import argparse
import json
import sys
import subprocess

from bom_constants import CATEGORY_TO_INDUSTRY


# V6 物料级专属字段（逆向回收后默认空串，未匹配到的保持空）
# 唯一 JSON 键计为 37 个：V5 实际 28 个（V4(7) + 纺织(5) + 家具(4) + 电子扩列(6) +
# 化工扩列(5) + 成本(2)，减去 color_no 与纺织共用重复计 1 = 28），V6 净增 9 个唯一键
# （drawing_no / material / heat_treatment / weight / unit_weight / basis_weight /
# size / print_process / eco_label）；其中 surface_treatment 已存在于 V5 家具，
# material 与包装/机械同名仅计一次。概念计数（含同名重复计）为 40，去重后 37。
_SPECIAL_FIELDS = [
    # V4 电子/化工
    "designator",
    "footprint",
    "part_number",
    "rohs",
    "cas_number",
    "concentration",
    "ghs_hazard",
    # V5 纺织
    "composition",
    "yarn_count",
    "fabric_weight",
    "width",
    "color_no",
    # V5 家具
    "material_grade",
    "spec_size",
    "surface_treatment",
    # 注：color_no 与纺织共用，不重复列出
    # V5 电子扩列
    "manufacturer",
    "tolerance",
    "rated_power",
    "rated_voltage",
    "alternate",
    "reflow_temp",
    # V5 化工扩列
    "purity",
    "physical_state",
    "flash_point",
    "storage_condition",
    "hazard_class",
    # V5 成本（入库字段；total_price 为派生不入库）
    "unit_price",
    "currency",
    # V6 机械（material 与包装同名、surface_treatment 与家具同名，均只计一次）
    "drawing_no",
    "material",
    "heat_treatment",
    "weight",
    "unit_weight",
    # V6 包装
    "basis_weight",
    "size",
    "print_process",
    "eco_label",
]


def ensure_openpyxl():
    """确保 openpyxl 可用，缺失时自动安装。"""
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"]
        )


def _find_cell_value_with_substring(ws, substring):
    """在工作表中查找包含指定子串的单元格值（按行、按列顺序，取首个命中）。

    合并单元格的值存储在其左上角（top-left）单元格，因此可直接扫描所有单元格。
    """
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is not None and substring in str(value):
                return str(value)
    return None


def _extract_after_colon(text):
    """从「键：值」文本中提取冒号后的内容；找不到冒号则返回 None。

    同时兼容中文全角冒号「：」与 ASCII 半角冒号「:」。
    """
    if text is None:
        return None
    for sep in ("：", ":"):
        if sep in text:
            return text.split(sep, 1)[1].strip()
    return None


# 行5 表头区可能拼接的多个字段 key（用于从合并单格中按 key 截取对应值）
_META_KEYS = ("审批人", "生效日期", "执行标准")


def _extract_meta_field(text, key):
    """从可能合并了多个「key：value」段的单元格文本中提取指定 key 的值。

    适用于 V3 行5 单格合并（审批人 / 生效日期 / 执行标准 拼接）场景；
    旧版无此单元格则返回空串。值取到下一个已知 key 冒号前或文本末尾。
    """
    if not text:
        return ""
    for sep in ("：", ":"):
        marker = key + sep
        idx = text.find(marker)
        if idx == -1:
            continue
        value = text[idx + len(marker):]
        cut = len(value)
        for nk in _META_KEYS:
            for nsep in ("：", ":"):
                nidx = value.find(nk + nsep)
                if nidx != -1 and nidx < cut:
                    cut = nidx
        return value[:cut].strip()
    return ""


def _to_float(value):
    """尽量转为 float；None 或空串返回空字符串，无法解析则保留原值。"""
    if value is None:
        return ""
    if isinstance(value, str) and value.strip() == "":
        return ""
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _str_or_empty(value):
    """转为去空格字符串；None 返回空字符串。"""
    if value is None:
        return ""
    return str(value).strip()


def _row_is_blank(ws, row_index, max_col=8):
    """判断一行（A–max_col）是否完全为空。"""
    for col in range(1, max_col + 1):
        if ws.cell(row_index, col).value not in (None, ""):
            return False
    return True


def _find_marker_row(ws, marker):
    """定位首列包含指定标记文本的行号（1-based）；未找到返回 None。"""
    for idx, row in enumerate(ws.iter_rows(min_col=1, max_col=1), 1):
        value = row[0].value
        if value is not None and marker in str(value):
            return idx
    return None


def _map_header(ws, header_row, max_col=8):
    """读取表头行，返回 {表头文本(去空格): 列号(1-based)} 映射。"""
    mapping = {}
    for col in range(1, max_col + 1):
        value = ws.cell(header_row, col).value
        if value is not None and str(value).strip() != "":
            mapping[str(value).strip()] = col
    return mapping


def _col_of(mapping, candidates):
    """在映射中按候选表头文本查找列号；返回首个命中，未命中返回 None。

    candidates 形如 ["单位", "计量单位"]（优先级从高到低）。
    """
    for cand in candidates:
        if cand in mapping:
            return mapping[cand]
    return None


def _infer_industry_from_blocks(ws, category):
    """从「三、」区块标记推断 industry（V4 新增，V5 扩展）。

    推断优先级：
    1. 有「三、元件清单」→ 电子
    2. 有「三、配方表」→ 化工
    3. 有「三、配料表」→ 食品
    4. 有「三、面料辅料清单」→ 纺织
    5. 有「三、家具物料清单」→ 家具
    6. 有「三、机械物料清单」→ 机械（V6 新增）
    7. 有「三、包装物料清单」→ 包装（V6 新增）
    8. 无「三、」行业区块 → 按 category 推断（CATEGORY_TO_INDUSTRY）

    注：成本明细（「三、成本明细」/「四、成本明细」）不参与 industry 推断。

    Args:
        ws: openpyxl Worksheet。
        category: 从表头区解析的产品类别字符串。

    Returns:
        industry 字符串。
    """
    if _find_marker_row(ws, "三、元件清单") is not None:
        return "电子"
    if _find_marker_row(ws, "三、配方表") is not None:
        return "化工"
    if _find_marker_row(ws, "三、配料表") is not None:
        return "食品"
    if _find_marker_row(ws, "三、面料辅料清单") is not None:
        return "纺织"
    if _find_marker_row(ws, "三、家具物料清单") is not None:
        return "家具"
    if _find_marker_row(ws, "三、机械物料清单") is not None:
        return "机械"
    if _find_marker_row(ws, "三、包装物料清单") is not None:
        return "包装"
    return CATEGORY_TO_INDUSTRY.get(category, "通用")


def _recover_block_fields(ws, marker_row, field_col_map, materials):
    """从「三、」派生区块逐行按物料名回收专属字段（V4 新增，V5 扩展）。

    通用回收函数：适用于元件清单、配方表、面料辅料清单、家具物料清单、成本明细。
    遇到空行（物料名称为空）、首列以「【」开头（分组标题）或首列含「合计」
    （成本合计行）则停止。

    Args:
        ws: openpyxl Worksheet。
        marker_row: 区块标题行号（1-based）。
        field_col_map: {物料字段名: 列号} 映射，须包含 "name" 键。
        materials: 已解析的物料列表（原地修改）。
    """
    name_c = field_col_map.get("name")
    if not name_c:
        return
    # 数值型字段：回收时转为 float（空则保留空串），保证闭环类型一致
    # V6 增 weight/unit_weight/basis_weight（机械/包装数值字段）
    float_fields = ("concentration", "fabric_weight", "unit_price",
                    "weight", "unit_weight", "basis_weight")
    rr = marker_row + 2  # 跳过标题行和表头行
    while rr <= ws.max_row:
        nm = _str_or_empty(ws.cell(rr, name_c).value)
        if not nm:
            break
        first_col = str(ws.cell(rr, 1).value or "").strip()
        if first_col.startswith("【"):
            break
        if "合计" in first_col:
            break
        for m in materials:
            if str(m.get("name") or "").strip() == nm:
                for field, col in field_col_map.items():
                    if field == "name" or col is None:
                        continue
                    raw_val = ws.cell(rr, col).value
                    if field in float_fields:
                        m[field] = _to_float(raw_val)
                    else:
                        m[field] = _str_or_empty(raw_val)
                break
        rr += 1


def parse_bom(path):
    """解析 BOM Excel 文件，返回与正向输入 JSON Schema 一致的 dict。

    若缺少必要标记（一、物料信息 / 二、工艺工序），打印 PARSE_ERROR 并以
    退出码 2 结束进程。

    V4 增强：推断 industry + 回收电子/化工专属字段 + 物料对象补全专属字段默认空串。
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print("FILE_ERROR: 无法读取 Excel 文件：%s" % e, file=sys.stderr)
        sys.exit(2)
    ws = wb["BOM表"] if "BOM表" in wb.sheetnames else wb.active

    # 标题容错：非预期标题仅告警，不中断。
    title_value = ws["A1"].value
    if title_value != "BOM表":
        print(
            "WARNING: 标题预期为 'BOM表'，实际为 '%s'，仍继续解析。"
            % title_value,
            file=sys.stderr,
        )

    # 产品级字段：扫描包含关键字的单元格并提取冒号后内容。
    version_text = _find_cell_value_with_substring(ws, "版本号")
    date_text = _find_cell_value_with_substring(ws, "生成日期")
    product_name_text = _find_cell_value_with_substring(ws, "产品名称")
    category_text = _find_cell_value_with_substring(ws, "产品类别")
    output_rate_text = _find_cell_value_with_substring(ws, "全产品出品率")
    approver_text = _find_cell_value_with_substring(ws, "审批人")
    effective_date_text = _find_cell_value_with_substring(ws, "生效日期")
    standard_text = _find_cell_value_with_substring(ws, "执行标准")

    version = _extract_after_colon(version_text) or "V1.0"
    date = _extract_after_colon(date_text) or ""
    product_name = _extract_after_colon(product_name_text) or ""
    category = _extract_after_colon(category_text) or "其他"
    approver = _extract_meta_field(approver_text, "审批人")
    effective_date = _extract_meta_field(effective_date_text, "生效日期")
    standard = _extract_meta_field(standard_text, "执行标准")

    raw_rate = _extract_after_colon(output_rate_text)
    if raw_rate is not None:
        cleaned = raw_rate.replace("%", "").strip()
        try:
            output_rate = float(cleaned)
        except (TypeError, ValueError):
            output_rate = ""
    else:
        output_rate = ""

    # 定位区标记行。
    material_row = _find_marker_row(ws, "一、物料信息")
    if material_row is None:
        print("PARSE_ERROR: 未找到物料区标记『一、物料信息』，无法解析。")
        sys.exit(2)

    process_row = _find_marker_row(ws, "二、工艺工序")
    if process_row is None:
        print("PARSE_ERROR: 未找到工序区标记『二、工艺工序』，无法解析。")
        sys.exit(2)

    # V4: 提前定位所有「三、」区块标记（用于工序区停止边界 + 专属字段回收）
    ingredient_row = _find_marker_row(ws, "三、配料表")
    component_row = _find_marker_row(ws, "三、元件清单")
    formula_row = _find_marker_row(ws, "三、配方表")
    # V5: 新增纺织/家具/成本区块定位
    textile_row = _find_marker_row(ws, "三、面料辅料清单")
    furniture_row = _find_marker_row(ws, "三、家具物料清单")
    # V6: 新增机械/包装区块定位
    mechanical_row = _find_marker_row(ws, "三、机械物料清单")
    packaging_row = _find_marker_row(ws, "三、包装物料清单")
    cost_row = _find_marker_row(ws, "成本明细")  # 关键字兼容「三、/四、成本明细」

    # 物料区表头行（标记行 +1），建立列头 → 列号映射。
    material_header_row = material_row + 1
    mat_map = _map_header(ws, material_header_row)
    # V3 首列为「序号」，物料名称在映射列（旧版为 A 列，映射同样命中）。
    mat_name_col = mat_map.get("物料名称")
    mat_unit_col = _col_of(mat_map, ["单位", "计量单位"])
    mat_usage_col = mat_map.get("用量")
    mat_yield_col = mat_map.get("出品率(%)")
    mat_erp_col = mat_map.get("ERP物料代码")
    mat_type_col = mat_map.get("物料类型")
    mat_proc_col = mat_map.get("所属工序")
    # 「序号」「用量占比%」等为派生/展示列，逆向忽略，不映射。

    # 物料数据行：从表头行下一行起，到工序标记行止；
    # 首列以「【」开头的分组子标题行跳过；首列含「合计」的合计行跳过；完全空行跳过。
    materials = []
    r = material_header_row + 1
    while r < process_row:
        col1 = ws.cell(r, 1).value  # V3=序号列 / 旧版=物料名称列
        if col1 is None or str(col1).strip() == "":
            if _row_is_blank(ws, r):
                r += 1
                continue
            break
        if str(col1).strip().startswith("【"):
            r += 1
            continue
        if "合计" in str(col1):  # 合计用量汇总行（V3），逆向跳过
            r += 1
            continue
        # 物料名称取映射列（非 A 列），缺映射则空
        name = _str_or_empty(ws.cell(r, mat_name_col).value) if mat_name_col else ""
        materials.append(
            {
                "name": name,
                "unit": _str_or_empty(ws.cell(r, mat_unit_col).value)
                if mat_unit_col
                else "",
                "usage": _to_float(ws.cell(r, mat_usage_col).value)
                if mat_usage_col
                else "",
                "yield_rate": _to_float(ws.cell(r, mat_yield_col).value)
                if mat_yield_col
                else "",
                "erp_code": _str_or_empty(ws.cell(r, mat_erp_col).value)
                if mat_erp_col
                else "",
                "material_type": _str_or_empty(ws.cell(r, mat_type_col).value)
                if mat_type_col
                else "其他",
                "process": _str_or_empty(ws.cell(r, mat_proc_col).value)
                if mat_proc_col
                else "",
                "allergen": "",  # 先置空，待配料表区块回收
            }
        )
        r += 1

    # 工序区表头行（标记行 +1），建立列头 → 列号映射。
    process_header_row = process_row + 1
    proc_map = _map_header(ws, process_header_row)
    proc_no_col = proc_map.get("工序编号")
    proc_name_col = proc_map.get("工序名称")
    proc_desc_col = proc_map.get("工序说明")
    proc_wh_col = proc_map.get("工时")
    proc_note_col = proc_map.get("备注")
    proc_out_col = proc_map.get("产物")

    # 工序数据行：从表头行下一行起，到任一「三、」区块标记或空编号止；
    # V4: 工序区停止边界扩展为所有「三、」区块（元件清单/配方表/配料表）。
    processes = []
    r = process_header_row + 1
    while r <= ws.max_row:
        if ingredient_row is not None and r >= ingredient_row:
            break
        if component_row is not None and r >= component_row:
            break
        if formula_row is not None and r >= formula_row:
            break
        if textile_row is not None and r >= textile_row:
            break
        if furniture_row is not None and r >= furniture_row:
            break
        if mechanical_row is not None and r >= mechanical_row:
            break
        if packaging_row is not None and r >= packaging_row:
            break
        if cost_row is not None and r >= cost_row:
            break
        col1 = ws.cell(r, 1).value
        if col1 is None or str(col1).strip() == "":
            if _row_is_blank(ws, r):
                r += 1
                continue
            break
        processes.append(
            {
                "step_no": str(col1).strip(),
                "name": _str_or_empty(ws.cell(r, proc_name_col).value)
                if proc_name_col
                else "",
                "desc": _str_or_empty(ws.cell(r, proc_desc_col).value)
                if proc_desc_col
                else "",
                "work_hours": _to_float(ws.cell(r, proc_wh_col).value)
                if proc_wh_col
                else "",
                "note": _str_or_empty(ws.cell(r, proc_note_col).value)
                if proc_note_col
                else "",
                "output": _str_or_empty(ws.cell(r, proc_out_col).value)
                if proc_out_col
                else "",
            }
        )
        r += 1

    # ===== V4: 专属字段回收 =====

    # 三、配料表 过敏原回收（V3 既有逻辑，仅食品；按物料名称匹配回写 material.allergen）
    if ingredient_row is not None:
        ing_map = _map_header(ws, ingredient_row + 1)
        name_c = ing_map.get("物料名称")
        alg_c = ing_map.get("过敏原")
        if name_c and alg_c:
            rr = ingredient_row + 2
            while rr <= ws.max_row:
                nm = _str_or_empty(ws.cell(rr, name_c).value)
                if not nm:
                    break
                if str(ws.cell(rr, 1).value or "").strip().startswith("【"):
                    break
                alg = _str_or_empty(ws.cell(rr, alg_c).value)
                for m in materials:
                    if str(m.get("name") or "").strip() == nm:
                        m["allergen"] = alg
                        break
                rr += 1

    # 三、元件清单 电子专属字段回收（V4 新增，V5 扩列至 10 字段）
    if component_row is not None:
        # V5: 元件清单扩列至 14 列 A–N，表头映射须覆盖到 N 列
        comp_map = _map_header(ws, component_row + 1, max_col=14)
        field_col_map = {
            "name": comp_map.get("物料名称"),
            "designator": _col_of(comp_map, ["位号(Designator)", "位号"]),
            "part_number": _col_of(comp_map, ["型号(Part#)", "型号"]),
            "footprint": _col_of(comp_map, ["封装(Footprint)", "封装"]),
            "rohs": _col_of(comp_map, ["RoHS"]),
            "manufacturer": _col_of(comp_map, ["制造商"]),
            "tolerance": _col_of(comp_map, ["容差"]),
            "rated_power": _col_of(comp_map, ["额定功率"]),
            "rated_voltage": _col_of(comp_map, ["额定电压"]),
            "alternate": _col_of(comp_map, ["替代料"]),
            "reflow_temp": _col_of(comp_map, ["封装温度"]),
        }
        _recover_block_fields(ws, component_row, field_col_map, materials)

    # 三、配方表 化工专属字段回收（V4 新增，V5 扩列至 8 字段）
    if formula_row is not None:
        # V5: 配方表扩列至 13 列 A–M，表头映射须覆盖到 M 列
        form_map = _map_header(ws, formula_row + 1, max_col=13)
        field_col_map = {
            "name": form_map.get("物料名称"),
            "cas_number": _col_of(form_map, ["CAS号", "CAS"]),
            "concentration": _col_of(form_map, ["含量(%)", "含量"]),
            "ghs_hazard": _col_of(form_map, ["GHS标识", "GHS"]),
            "purity": _col_of(form_map, ["纯度"]),
            "physical_state": _col_of(form_map, ["物态"]),
            "flash_point": _col_of(form_map, ["闪点"]),
            "storage_condition": _col_of(form_map, ["存储条件"]),
            "hazard_class": _col_of(form_map, ["危险等级"]),
        }
        _recover_block_fields(ws, formula_row, field_col_map, materials)

    # 三、面料辅料清单 纺织专属字段回收（V5 新增，5 字段）
    if textile_row is not None:
        tex_map = _map_header(ws, textile_row + 1)
        field_col_map = {
            "name": tex_map.get("物料名称"),
            "composition": _col_of(tex_map, ["成分比例"]),
            "yarn_count": _col_of(tex_map, ["纱支"]),
            "fabric_weight": _col_of(tex_map, ["克重(g/m²)", "克重"]),
            "width": _col_of(tex_map, ["幅宽"]),
            "color_no": _col_of(tex_map, ["色号"]),
        }
        _recover_block_fields(ws, textile_row, field_col_map, materials)

    # 三、家具物料清单 家具专属字段回收（V5 新增，4 字段）
    if furniture_row is not None:
        fur_map = _map_header(ws, furniture_row + 1)
        field_col_map = {
            "name": fur_map.get("物料名称"),
            "material_grade": _col_of(fur_map, ["材质等级"]),
            "spec_size": _col_of(fur_map, ["尺寸规格"]),
            "surface_treatment": _col_of(fur_map, ["表面处理"]),
            "color_no": _col_of(fur_map, ["色号/花色", "色号"]),
        }
        _recover_block_fields(ws, furniture_row, field_col_map, materials)

    # 三、机械物料清单 机械专属字段回收（V6 新增，6 字段）
    if mechanical_row is not None:
        mech_map = _map_header(ws, mechanical_row + 1)
        field_col_map = {
            "name": mech_map.get("物料名称"),
            "drawing_no": _col_of(mech_map, ["图号"]),
            "material": _col_of(mech_map, ["材质"]),
            "heat_treatment": _col_of(mech_map, ["热处理"]),
            "surface_treatment": _col_of(mech_map, ["表面处理"]),
            "weight": _col_of(mech_map, ["重量(kg)", "重量"]),
            "unit_weight": _col_of(mech_map, ["单重(kg/件)", "单重"]),
        }
        _recover_block_fields(ws, mechanical_row, field_col_map, materials)

    # 三、包装物料清单 包装专属字段回收（V6 新增，5 字段）
    # 注：material_type 已在物料区回收，此处 field_col_map 不含该键，避免重复回写
    if packaging_row is not None:
        pack_map = _map_header(ws, packaging_row + 1)
        field_col_map = {
            "name": pack_map.get("物料名称"),
            "material": _col_of(pack_map, ["材质"]),
            "basis_weight": _col_of(pack_map, ["克重(g/m²)", "克重"]),
            "size": _col_of(pack_map, ["尺寸"]),
            "print_process": _col_of(pack_map, ["印刷工艺"]),
            "eco_label": _col_of(pack_map, ["环保标识"]),
        }
        _recover_block_fields(ws, packaging_row, field_col_map, materials)

    # 成本明细 成本字段回收（V5 新增，关键字「成本明细」兼容 三/四、前缀）
    # 回收 unit_price/currency；总价(H列)为派生展示，不回收
    if cost_row is not None:
        cost_map = _map_header(ws, cost_row + 1)
        field_col_map = {
            "name": cost_map.get("物料名称"),
            "unit_price": _col_of(cost_map, ["单价"]),
            "currency": _col_of(cost_map, ["币种"]),
        }
        _recover_block_fields(ws, cost_row, field_col_map, materials)

    # V4: 物料对象补全专属字段默认空串（未回收到的）
    for m in materials:
        for f in _SPECIAL_FIELDS:
            if f not in m:
                m[f] = ""

    # V4: 推断 industry（从区块标记或 category 推断）
    industry = _infer_industry_from_blocks(ws, category)

    return {
        "product_name": product_name,
        "category": category,
        "industry": industry,
        "output_rate": output_rate,
        "version": version,
        "date": date,
        "approver": approver,
        "effective_date": effective_date,
        "standard": standard,
        "materials": materials,
        "processes": processes,
    }


def main():
    ensure_openpyxl()
    parser = argparse.ArgumentParser(description="从 BOM 表 Excel 逆向导入 JSON")
    parser.add_argument("--in", dest="in_path", required=True, help="BOM 表 xlsx 文件路径")
    parser.add_argument("--out", default=None, help="可选，导出 JSON 文件路径")
    args = parser.parse_args()

    data = parse_bom(args.in_path)
    json_str = json.dumps(data, ensure_ascii=False, indent=2)

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(json_str)
        print("OK:" + args.out)
    else:
        print("OK:" + json_str)


if __name__ == "__main__":
    main()
