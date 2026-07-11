#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BOM智造师 V5 (行业扩充 + 扩列 + 成本视图) 测试脚本（QA · 严过关）

沿用 test_bom_v4.py 的测试风格：通过 subprocess 调用 generate_bom.py /
import_bom.py，用 openpyxl 读回 xlsx 断言；断言具体、覆盖正向生成 +
逆向回收 + 列数/字段断言 + 全量回归。

覆盖 V5 全部增量行为：
  V5-T01 纺织视图：生成「三、面料辅料清单」8 列(A–H)，表头正确
  V5-T02 纺织 5 新字段(成分比例/纱支/克重/幅宽/色号)正确写入
  V5-T03 纺织排除 material_type=="其他" 的物料
  V5-T04 家具视图：生成「三、家具物料清单」8 列，表头正确
  V5-T05 家具 4 新字段(材质等级/尺寸规格/表面处理/色号花色)正确
  V5-T06 家具排除 material_type=="其他" 的物料
  V5-T07 电子扩列：生成「三、元件清单」14 列(A–N)，表头正确
  V5-T08 电子 10 字段(I–N)全部正确写入
  V5-T09 电子 RoHS 红黄字着色逻辑沿用(否红/未知空黄)
  V5-T10 化工扩列：生成「三、配方表」13 列(A–M)，表头正确
  V5-T11 化工 8 字段(I–M)全部正确写入；含量(%) 格式 0.0"%
  V5-T12 成本视图(通用)：生成「三、成本明细」；H=用量×单价；成本合计=Σ
  V5-T13 成本币种默认 人民币(CNY)
  V5-T14 电子带成本：生成「四、成本明细」(双编号) + H=用量×单价
  V5-T15 逆向 JSON 不含 total_price 键(派生不入库)
  V5-T16 行业模板预设：INDUSTRY_TEMPLATES 含电子/化工/纺织/家具/食品；
           各自含 material_types/standard/special_fields/preset_processes；
           special_fields 数量(电子10/化工8/纺织5/家具4)
  V5-T17 逆向回收-纺织：5 字段正确回填，值与原始一致
  V5-T18 逆向回收-家具：4 字段正确回填，值与原始一致
  V5-T19 逆向回收-电子：10 字段 + unit_price/currency 正确回填，一致
  V5-T20 逆向回收-化工：8 字段正确回填，值与原始一致
  V5-T21 逆向回收-成本：unit_price/currency 正确回填，一致
  V5-T22 向后兼容-电子(V4 旧示例)：仍 14 列，扩列 I–N 为空
  V5-T23 向后兼容-化工(V4 旧示例)：仍 13 列，扩列 I–M 为空
  V5-T24 向后兼容-食品(V3 旧示例)：配料表照常，无新区块
  V5-T25 最小变更核查：生成 V4 旧格式 JSON 不新增任何 WARNING 类型
  V5-T26 py_compile：三脚本编译通过 + 业务模块可 import

运行（主理人指定 venv）：
  <venv python> -m pytest tests/test_bom_v2.py tests/test_bom_v3.py \
      tests/test_bom_v4.py tests/test_bom_v5.py -q
"""

import os
import sys
import json
import shutil
import subprocess
import tempfile

from openpyxl import load_workbook

# ----------------------------------------------------------------------------
# 路径常量（沿用 test_bom_v4.py 的 SKILL_DIR 推导方式）
# ----------------------------------------------------------------------------
SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GEN = os.path.join(SKILL_DIR, "scripts", "generate_bom.py")
IMP = os.path.join(SKILL_DIR, "scripts", "import_bom.py")
CONST = os.path.join(SKILL_DIR, "scripts", "bom_constants.py")

# 主理人指定的 venv python（已装 pytest 9.1.1 + openpyxl 3.1.5）
PYTHON = r"C:\Users\姓名\.workbuddy\binaries\python\envs\default\Scripts\python.exe"

# V5 示例
SAMPLE_TEXTILE = os.path.join(SKILL_DIR, "examples", "sample_bom_v5_textile.json")
SAMPLE_FURNITURE = os.path.join(SKILL_DIR, "examples", "sample_bom_v5_furniture.json")
SAMPLE_ELEC_V5 = os.path.join(SKILL_DIR, "examples", "sample_bom_v5_electronic.json")
SAMPLE_CHEM_V5 = os.path.join(SKILL_DIR, "examples", "sample_bom_v5_chemical.json")
SAMPLE_COST = os.path.join(SKILL_DIR, "examples", "sample_bom_v5_cost.json")
# V4/V3 旧示例
SAMPLE_ELEC_V4 = os.path.join(SKILL_DIR, "examples", "sample_bom_v4_electronic.json")
SAMPLE_CHEM_V4 = os.path.join(SKILL_DIR, "examples", "sample_bom_v4_chemical.json")
SAMPLE_V3 = os.path.join(SKILL_DIR, "examples", "sample_bom_v3.json")

# 各视图专属字段集合（逆向回收后用于一致性比对）
ELEC_FIELDS = [
    "designator", "footprint", "part_number", "rohs",
    "manufacturer", "tolerance", "rated_power",
    "rated_voltage", "alternate", "reflow_temp",
]
CHEM_FIELDS = [
    "cas_number", "concentration", "ghs_hazard",
    "purity", "physical_state", "flash_point",
    "storage_condition", "hazard_class",
]
TEXTILE_FIELDS = ["composition", "yarn_count", "fabric_weight", "width", "color_no"]
FURNITURE_FIELDS = ["material_grade", "spec_size", "surface_treatment", "color_no"]

# RoHS 着色色值（与 generate_bom.py 保持一致）
ROHS_RED = "FF0000"
ROHS_YELLOW = "BF8F00"


# ----------------------------------------------------------------------------
# 运行辅助
# ----------------------------------------------------------------------------
def run_generate(data, out_path):
    """把 dict 写成临时 json，调用 generate_bom.py，返回 (rc, stdout, stderr)。"""
    tmp = out_path + ".in.json"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    p = subprocess.run(
        [PYTHON, GEN, "--data", tmp, "--out", out_path],
        capture_output=True, text=True,
    )
    return p.returncode, p.stdout, p.stderr


def run_import(xlsx_path, out_json=None):
    """调用 import_bom.py，返回 (rc, stdout, stderr, parsed_json_or_None)。"""
    args = [PYTHON, IMP, "--in", xlsx_path]
    if out_json:
        args += ["--out", out_json]
    p = subprocess.run(args, capture_output=True, text=True)
    parsed = None
    if out_json and os.path.exists(out_json):
        with open(out_json, encoding="utf-8") as f:
            parsed = json.load(f)
    return p.returncode, p.stdout, p.stderr, parsed


def load(path):
    """载入 xlsx 的活动工作表。"""
    return load_workbook(path).active


def flat_values(ws):
    """返回工作表中所有非空单元格的字符串列表。"""
    vals = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is not None and str(v).strip() != "":
                vals.append(str(v))
    return vals


def marker_row(ws, marker):
    """返回首列包含 marker 的行号（1-based），未找到返回 None。"""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and marker in str(v):
            return r
    return None


def _blank(x):
    """判断值是否「空」（None 或空串）。注意 0 不视为空。"""
    return x is None or (isinstance(x, str) and x.strip() == "")


def val_match(a, b):
    """数值容忍比较：Excel 把整数读成 float，故用 float 兼容 int/float/字符串。
    两者皆空视为相等；否则尝试数值比较，失败则按字符串比较。
    """
    if _blank(a) and _blank(b):
        return True
    try:
        fa = float(a)
        fb = float(b)
        return abs(fa - fb) < 1e-9
    except (TypeError, ValueError):
        return str(a).strip() == str(b).strip()


def read_block(ws, marker):
    """读取标记区块，返回 (headers, rows)。
    headers: 按列号顺序的表头文本列表；
    rows: 每个数据行的 {表头文本: 单元格值}。
    遇到空行/分组标题(【)/合计行 即停止。
    未找到标记返回 (None, None)。
    """
    m = marker_row(ws, marker)
    if m is None:
        return None, None
    hdr_row = m + 1
    col_of = {}
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if v is not None and str(v).strip() != "":
            col_of[str(v).strip()] = c
            headers.append(str(v).strip())
    rows = []
    r = hdr_row + 1
    while r <= ws.max_row:
        nm = ws.cell(r, 1).value
        if nm is None or str(nm).strip() == "":
            break
        if str(nm).strip().startswith("【"):
            r += 1
            continue
        if "合计" in str(nm).strip():
            break
        row = {}
        for h, c in col_of.items():
            row[h] = ws.cell(r, c).value
        rows.append(row)
        r += 1
    return headers, rows


def find_block_row(ws, marker, name, name_col):
    """在标记区块内按「物料名称」列定位某物料所在行号(1-based)。"""
    m = marker_row(ws, marker)
    assert m is not None, "未找到区块标记: %s" % marker
    r = m + 2
    while r <= ws.max_row:
        v = ws.cell(r, 1).value
        if v is None or str(v).strip() == "":
            break
        if str(ws.cell(r, name_col).value or "").strip() == name:
            return r
        r += 1
    return None


def read_cost_block(ws):
    """读取成本明细区块，返回 (title_text, rows, total_row)。
    rows: [{name, usage, unit_price, currency, total}, ...]
    total_row: {"name": "成本合计", "total": H列值} 或 None
    """
    m = marker_row(ws, "成本明细")
    assert m is not None, "未找到成本明细区块"
    hdr_row = m + 1
    col_of = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if v is not None and str(v).strip() != "":
            col_of[str(v).strip()] = c
    rows = []
    total_row = None
    r = hdr_row + 1
    while r <= ws.max_row:
        nm = ws.cell(r, 1).value
        if nm is None or str(nm).strip() == "":
            break
        if str(nm).strip().startswith("【"):
            r += 1
            continue
        if "合计" in str(nm).strip():
            total_row = {
                "name": str(nm).strip(),
                "total": ws.cell(r, col_of["总价"]).value,
            }
            r += 1
            continue
        rows.append({
            "name": str(ws.cell(r, col_of["物料名称"]).value or "").strip(),
            "usage": ws.cell(r, col_of["用量"]).value,
            "unit_price": ws.cell(r, col_of["单价"]).value,
            "currency": str(ws.cell(r, col_of["币种"]).value or "").strip(),
            "total": ws.cell(r, col_of["总价"]).value,
        })
        r += 1
    return str(ws.cell(m, 1).value or ""), rows, total_row


# ----------------------------------------------------------------------------
# V5-T01 ~ T03 纺织视图
# ----------------------------------------------------------------------------
def test_v5_t01_textile_view():
    """纺织 industry=纺织 生成「三、面料辅料清单」8 列(A–H)，表头正确。"""
    with open(SAMPLE_TEXTILE, encoding="utf-8") as f:
        data = json.load(f)
    assert data.get("industry") == "纺织", "T01 样本 industry 应=纺织"
    wd = tempfile.mkdtemp(prefix="v5t01_")
    out = os.path.join(wd, "BOM_t01.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T01 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    m = marker_row(ws, "三、面料辅料清单")
    assert m is not None, "T01 应生成『三、面料辅料清单』"
    headers, _ = read_block(ws, "三、面料辅料清单")
    assert headers is not None, "T01 应读到面料辅料清单表头"
    # 8 列(A–H)
    assert len(headers) == 8, "T01 面料辅料清单应为 8 列, 实际 %d: %s" % (len(headers), headers)
    expected = ["序号", "物料名称", "物料类型", "成分比例", "纱支",
                "克重(g/m²)", "幅宽", "色号"]
    assert headers == expected, "T01 表头不符: %s" % headers


def test_v5_t02_textile_fields():
    """纺织 5 个新字段(composition/yarn_count/fabric_weight/width/color_no)正确写入。"""
    with open(SAMPLE_TEXTILE, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t02_")
    out = os.path.join(wd, "BOM_t02.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T02 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    _, rows = read_block(ws, "三、面料辅料清单")
    assert rows is not None, "T02 应存在面料辅料清单数据行"
    by_name = {r["物料名称"]: r for r in rows}
    # 棉涤混纺布
    b = by_name.get("棉涤混纺布")
    assert b is not None, "T02 应含棉涤混纺布"
    assert b["成分比例"] == "65%棉 35%涤", "T02 成分比例不符: %s" % b["成分比例"]
    assert b["纱支"] == "40S", "T02 纱支不符: %s" % b["纱支"]
    assert val_match(b["克重(g/m²)"], 180), "T02 克重不符: %s" % b["克重(g/m²)"]
    assert val_match(b["幅宽"], 150), "T02 幅宽不符: %s" % b["幅宽"]
    assert b["色号"] == "TF-C01", "T02 色号不符: %s" % b["色号"]
    # 涤纶缝纫线（部分字段为空）
    t = by_name.get("涤纶缝纫线")
    assert t is not None, "T02 应含涤纶缝纫线"
    assert t["成分比例"] == "100%涤", "T02 成分比例不符: %s" % t["成分比例"]
    assert t["纱支"] == "40S/2", "T02 纱支不符: %s" % t["纱支"]
    assert _blank(t["克重(g/m²)"]) and _blank(t["幅宽"]), "T02 空字段应为空"


def test_v5_t03_textile_exclude_other():
    """纺织排除 material_type=='其他' 的物料（裁片/OPP包装袋）。"""
    with open(SAMPLE_TEXTILE, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t03_")
    out = os.path.join(wd, "BOM_t03.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T03 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    _, rows = read_block(ws, "三、面料辅料清单")
    names = [r["物料名称"] for r in rows]
    assert "裁片" not in names, "T03 其他类(裁片)应被排除"
    assert "OPP包装袋" not in names, "T03 其他类(OPP包装袋)应被排除"
    assert "面料辅料清单已排除" in so, "T03 应打印排除提示WARNING"
    flat = flat_values(ws)
    assert "裁片" in flat and "OPP包装袋" in flat, "T03 被排除物料应仍在物料区"


# ----------------------------------------------------------------------------
# V5-T04 ~ T06 家具视图
# ----------------------------------------------------------------------------
def test_v5_t04_furniture_view():
    """家具 industry=家具 生成「三、家具物料清单」8 列，表头正确。"""
    with open(SAMPLE_FURNITURE, encoding="utf-8") as f:
        data = json.load(f)
    assert data.get("industry") == "家具", "T04 样本 industry 应=家具"
    wd = tempfile.mkdtemp(prefix="v5t04_")
    out = os.path.join(wd, "BOM_t04.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T04 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    m = marker_row(ws, "三、家具物料清单")
    assert m is not None, "T04 应生成『三、家具物料清单』"
    headers, _ = read_block(ws, "三、家具物料清单")
    assert headers is not None, "T04 应读到家具物料清单表头"
    assert len(headers) == 8, "T04 家具物料清单应为 8 列, 实际 %d: %s" % (len(headers), headers)
    expected = ["序号", "物料名称", "物料类型", "材质等级", "尺寸规格",
                "表面处理", "用量", "色号/花色"]
    assert headers == expected, "T04 表头不符: %s" % headers


def test_v5_t05_furniture_fields():
    """家具 4 字段(材质等级/尺寸规格/表面处理/色号花色)正确写入。"""
    with open(SAMPLE_FURNITURE, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t05_")
    out = os.path.join(wd, "BOM_t05.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T05 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    _, rows = read_block(ws, "三、家具物料清单")
    assert rows is not None, "T05 应存在家具物料清单数据行"
    by_name = {r["物料名称"]: r for r in rows}
    z = by_name.get("橡胶木主框架")
    assert z is not None, "T05 应含橡胶木主框架"
    assert z["材质等级"] == "A级", "T05 材质等级不符: %s" % z["材质等级"]
    assert z["尺寸规格"] == "45×45×400mm", "T05 尺寸规格不符: %s" % z["尺寸规格"]
    assert z["表面处理"] == "开放漆", "T05 表面处理不符: %s" % z["表面处理"]
    assert z["色号/花色"] == "FZ-001-原木", "T05 色号不符: %s" % z["色号/花色"]
    # 科技木皮
    p = by_name.get("科技木皮")
    assert p is not None, "T05 应含科技木皮"
    assert p["材质等级"] == "AA级", "T05 材质等级不符: %s" % p["材质等级"]
    assert p["尺寸规格"] == "0.5mm厚", "T05 尺寸规格不符: %s" % p["尺寸规格"]
    assert p["表面处理"] == "哑光", "T05 表面处理不符: %s" % p["表面处理"]
    assert p["色号/花色"] == "FZ-003-胡桃", "T05 色号不符: %s" % p["色号/花色"]


def test_v5_t06_furniture_exclude_other():
    """家具排除 material_type=='其他' 的物料（板件/瓦楞包装纸箱）。"""
    with open(SAMPLE_FURNITURE, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t06_")
    out = os.path.join(wd, "BOM_t06.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T06 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    _, rows = read_block(ws, "三、家具物料清单")
    names = [r["物料名称"] for r in rows]
    assert "板件" not in names, "T06 其他类(板件)应被排除"
    assert "瓦楞包装纸箱" not in names, "T06 其他类(瓦楞包装纸箱)应被排除"
    assert "家具物料清单已排除" in so, "T06 应打印排除提示WARNING"
    flat = flat_values(ws)
    assert "板件" in flat and "瓦楞包装纸箱" in flat, "T06 被排除物料应仍在物料区"


# ----------------------------------------------------------------------------
# V5-T07 ~ T09 电子扩列
# ----------------------------------------------------------------------------
def test_v5_t07_electronic_expanded_header():
    """电子 industry=电子 生成「三、元件清单」14 列(A–N)，表头正确。"""
    with open(SAMPLE_ELEC_V5, encoding="utf-8") as f:
        data = json.load(f)
    assert data.get("industry") == "电子", "T07 样本 industry 应=电子"
    wd = tempfile.mkdtemp(prefix="v5t07_")
    out = os.path.join(wd, "BOM_t07.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T07 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    m = marker_row(ws, "三、元件清单")
    assert m is not None, "T07 应生成『三、元件清单』"
    headers, _ = read_block(ws, "三、元件清单")
    assert headers is not None, "T07 应读到元件清单表头"
    # 14 列(A–N)
    assert len(headers) == 14, "T07 元件清单应扩列至 14 列, 实际 %d: %s" % (len(headers), headers)
    expected = ["序号", "位号(Designator)", "型号(Part#)", "封装(Footprint)",
                "物料名称", "数量", "物料类型", "RoHS", "制造商", "容差",
                "额定功率", "额定电压", "替代料", "封装温度"]
    assert headers == expected, "T07 元件清单14列表头不符: %s" % headers


def test_v5_t08_electronic_expanded_fields():
    """电子 10 字段(I–N)全部正确写入。"""
    with open(SAMPLE_ELEC_V5, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t08_")
    out = os.path.join(wd, "BOM_t08.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T08 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    _, rows = read_block(ws, "三、元件清单")
    assert rows is not None, "T08 应存在元件清单数据行"
    by_name = {r["物料名称"]: r for r in rows}
    # 主控芯片：manufacturer/rated_power/rated_voltage/alternate/reflow_temp
    u = by_name.get("主控芯片")
    assert u is not None, "T08 应含主控芯片"
    assert u["制造商"] == "ST", "T08 制造商不符: %s" % u["制造商"]
    assert _blank(u["容差"]), "T08 主控芯片容差应空"
    assert u["额定功率"] == "0.36W", "T08 额定功率不符: %s" % u["额定功率"]
    assert u["额定电压"] == "3.6V", "T08 额定电压不符: %s" % u["额定电压"]
    assert u["替代料"] == "GD32F103C8T6", "T08 替代料不符: %s" % u["替代料"]
    assert u["封装温度"] == "260℃", "T08 封装温度不符: %s" % u["封装温度"]
    # 贴片电阻：全部 6 个扩列字段均有值
    r = by_name.get("贴片电阻")
    assert r is not None, "T08 应含贴片电阻"
    assert r["制造商"] == "Yageo", "T08 制造商不符: %s" % r["制造商"]
    assert r["容差"] == "±1%", "T08 容差不符: %s" % r["容差"]
    assert r["额定功率"] == "0.0625W", "T08 额定功率不符: %s" % r["额定功率"]
    assert r["额定电压"] == "50V", "T08 额定电压不符: %s" % r["额定电压"]
    assert r["替代料"] == "RC0402FR-0710KL", "T08 替代料不符: %s" % r["替代料"]
    assert r["封装温度"] == "260℃", "T08 封装温度不符: %s" % r["封装温度"]


def test_v5_t09_electronic_rohs_color():
    """电子 RoHS 红黄字着色逻辑沿用(否→红, 未知/空→黄)。"""
    with open(SAMPLE_ELEC_V5, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t09_")
    out = os.path.join(wd, "BOM_t09.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T09 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    # 否 → 红
    row_no = find_block_row(ws, "三、元件清单", "开关二极管", 5)
    assert row_no is not None, "T09 应含开关二极管"
    col = ws.cell(row_no, 8).font.color  # H 列 = RoHS
    assert col is not None and col.rgb.endswith(ROHS_RED), \
        "T09 rohs=否 应红色(%s), 实际 %s" % (ROHS_RED, col.rgb if col else None)
    # 空 → 黄
    row_empty = find_block_row(ws, "三、元件清单", "晶振", 5)
    assert row_empty is not None, "T09 应含晶振"
    col = ws.cell(row_empty, 8).font.color
    assert col is not None and col.rgb.endswith(ROHS_YELLOW), \
        "T09 rohs=空 应黄色(%s), 实际 %s" % (ROHS_YELLOW, col.rgb if col else None)
    # 是 → 默认(无特殊色)
    row_yes = find_block_row(ws, "三、元件清单", "贴片电阻", 5)
    assert row_yes is not None, "T09 应含贴片电阻"
    col = ws.cell(row_yes, 8).font.color
    assert col is None, \
        "T09 rohs=是 应默认(无特殊色), 实际 %s" % (col.rgb if col else None)


# ----------------------------------------------------------------------------
# V5-T10 ~ T11 化工扩列
# ----------------------------------------------------------------------------
def test_v5_t10_chemical_expanded_header():
    """化工 industry=化工 生成「三、配方表」13 列(A–M)，表头正确。"""
    with open(SAMPLE_CHEM_V5, encoding="utf-8") as f:
        data = json.load(f)
    assert data.get("industry") == "化工", "T10 样本 industry 应=化工"
    wd = tempfile.mkdtemp(prefix="v5t10_")
    out = os.path.join(wd, "BOM_t10.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T10 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    m = marker_row(ws, "三、配方表")
    assert m is not None, "T10 应生成『三、配方表』"
    headers, _ = read_block(ws, "三、配方表")
    assert headers is not None, "T10 应读到配方表表头"
    # 13 列(A–M)
    assert len(headers) == 13, "T10 配方表应扩列至 13 列, 实际 %d: %s" % (len(headers), headers)
    expected = ["序号", "物料名称", "CAS号", "含量(%)", "GHS标识",
                "物料类型", "计量单位", "用量", "纯度", "物态",
                "闪点", "存储条件", "危险等级"]
    assert headers == expected, "T10 配方表13列表头不符: %s" % headers


def test_v5_t11_chemical_expanded_fields():
    """化工 8 字段(I–M)全部正确写入；含量(%) 格式 0.0"%。"""
    with open(SAMPLE_CHEM_V5, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t11_")
    out = os.path.join(wd, "BOM_t11.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T11 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    _, rows = read_block(ws, "三、配方表")
    assert rows is not None, "T11 应存在配方表数据行"
    by_name = {r["物料名称"]: r for r in rows}
    # 去离子水
    w = by_name.get("去离子水")
    assert w is not None, "T11 应含去离子水"
    assert w["CAS号"] == "7732-18-5", "T11 CAS号不符: %s" % w["CAS号"]
    assert val_match(w["含量(%)"], 65.0), "T11 含量不符: %s" % w["含量(%)"]
    assert w["GHS标识"] == "无", "T11 GHS不符: %s" % w["GHS标识"]
    assert w["纯度"] == "≥99.9%", "T11 纯度不符: %s" % w["纯度"]
    assert w["物态"] == "液态", "T11 物态不符: %s" % w["物态"]
    assert w["闪点"] == "无", "T11 闪点不符: %s" % w["闪点"]
    assert w["存储条件"] == "常温密封", "T11 存储条件不符: %s" % w["存储条件"]
    assert w["危险等级"] == "无", "T11 危险等级不符: %s" % w["危险等级"]
    # 乙醇
    e = by_name.get("乙醇")
    assert e is not None, "T11 应含乙醇"
    assert e["CAS号"] == "64-17-5", "T11 CAS号不符: %s" % e["CAS号"]
    assert val_match(e["含量(%)"], 30.0), "T11 含量不符: %s" % e["含量(%)"]
    assert e["危险等级"] == "易燃液体 类别2", "T11 危险等级不符: %s" % e["危险等级"]
    # 含量(%) 数字格式 0.0"%" 校验
    m = marker_row(ws, "三、配方表")
    for r in range(m + 2, ws.max_row + 1):
        nm = ws.cell(r, 2).value
        if nm is None or str(nm).strip() == "":
            break
        if str(ws.cell(r, 1).value or "").strip().startswith("【"):
            continue
        fmt = ws.cell(r, 4).number_format  # D 列 = 含量(%)
        assert fmt == '0.0"%"', "T11 含量(%%)应0.0%%格式, 实际 %s (行%s)" % (fmt, r)


# ----------------------------------------------------------------------------
# V5-T12 ~ T15 成本视图
# ----------------------------------------------------------------------------
def test_v5_t12_cost_generic_view():
    """成本视图(通用)：生成「三、成本明细」；H=用量×单价；成本合计=Σ。"""
    with open(SAMPLE_COST, encoding="utf-8") as f:
        data = json.load(f)
    assert data.get("industry") == "通用", "T12 样本 industry 应=通用"
    wd = tempfile.mkdtemp(prefix="v5t12_")
    out = os.path.join(wd, "BOM_t12.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T12 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    title, rows, total_row = read_cost_block(ws)
    assert "三、成本明细" == title, "T12 通用应生成『三、成本明细』, 实际 %s" % title
    assert len(rows) == 4, "T12 成本明细应有 4 行, 实际 %d" % len(rows)
    # H 列总价 = 用量 × 单价
    expect = {
        "冷轧钢板": 2.5 * 5.2,   # 13.0
        "内六角螺栓": 8 * 0.35,    # 2.8
        "防锈漆": 0.3 * 22.0,      # 6.6
        "支架半成品": 1 * 0,         # 0
    }
    by_name = {r["name"]: r for r in rows}
    for name, exp in expect.items():
        r = by_name.get(name)
        assert r is not None, "T12 成本明细应含 %s" % name
        assert val_match(r["total"], round(exp, 2)), \
            "T12 %s 总价应为 %.2f, 实际 %s" % (name, round(exp, 2), r["total"])
        assert val_match(r["usage"], data_material_usage(data, name)), \
            "T12 %s 用量回填不一致" % name
    # 成本合计 = Σ 总价 = 22.4
    assert total_row is not None, "T12 应有成本合计行"
    assert total_row["name"] == "成本合计", "T12 合计行首列应=成本合计"
    assert val_match(total_row["total"], 22.4), \
        "T12 成本合计应为 22.4, 实际 %s" % total_row["total"]


def data_material_usage(data, name):
    """从原始 JSON 取某物料的用量(辅助)。"""
    for m in data.get("materials", []):
        if m.get("name") == name:
            return m.get("usage")
    return None


def test_v5_t13_cost_currency_default():
    """成本币种默认 人民币(CNY)（未填 currency 时）。"""
    with open(SAMPLE_COST, encoding="utf-8") as f:
        data = json.load(f)
    # 副本：去掉所有 currency 键，验证回退默认值
    import copy
    data2 = copy.deepcopy(data)
    for m in data2["materials"]:
        m.pop("currency", None)
    wd = tempfile.mkdtemp(prefix="v5t13_")
    out = os.path.join(wd, "BOM_t13.xlsx")
    rc, so, se = run_generate(data2, out)
    assert rc == 0, "T13 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    _, rows, _ = read_cost_block(ws)
    for r in rows:
        assert r["currency"] == "人民币(CNY)", \
            "T13 未填币种应回退『人民币(CNY)』, 实际 %s" % r["currency"]
    # 逆向回收同样应为默认
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T13 逆向应成功, rc=%s se2=%s" % (rc2, se2)
    for m in parsed["materials"]:
        assert m.get("currency") == "人民币(CNY)", \
            "T13 逆向币种应=人民币(CNY), 实际 %s" % m.get("currency")


def test_v5_t14_cost_electronic_dual_number():
    """电子带成本：生成「四、成本明细」(双编号) + H=用量×单价。"""
    with open(SAMPLE_ELEC_V5, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t14_")
    out = os.path.join(wd, "BOM_t14.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T14 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    # 电子已有行业视图，故成本应为「四、成本明细」
    assert marker_row(ws, "四、成本明细") is not None, \
        "T14 电子带成本应生成『四、成本明细』"
    title, rows, total_row = read_cost_block(ws)
    assert "四、成本明细" == title, "T14 标题应为『四、成本明细』, 实际 %s" % title
    # 全部 8 个物料均带 unit_price → 8 行
    assert len(rows) == 8, "T14 成本明细应有 8 行, 实际 %d" % len(rows)
    by_name = {r["name"]: r for r in rows}
    # 主控芯片 1 × 6.8 = 6.8
    u = by_name.get("主控芯片")
    assert u is not None, "T14 应含主控芯片"
    assert val_match(u["total"], round(1 * 6.8, 2)), "T14 主控芯片总价应为 6.8"
    # 贴片电阻 4 × 0.02 = 0.08
    r = by_name.get("贴片电阻")
    assert val_match(r["total"], round(4 * 0.02, 2)), "T14 贴片电阻总价应为 0.08"
    # 成本合计 = 16.21
    assert total_row is not None, "T14 应有成本合计行"
    assert val_match(total_row["total"], 16.21), \
        "T14 成本合计应为 16.21, 实际 %s" % total_row["total"]


def test_v5_t15_no_total_price_key():
    """逆向 JSON 不含 total_price 键（派生总价不入库）。"""
    with open(SAMPLE_COST, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t15_")
    out = os.path.join(wd, "BOM_t15.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T15 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T15 逆向应成功, rc=%s se2=%s" % (rc2, se2)
    for m in parsed["materials"]:
        assert "total_price" not in m, \
            "T15 逆向 JSON 物料不应含 total_price 键: %s" % m.get("name")


# ----------------------------------------------------------------------------
# V5-T16 行业模板预设
# ----------------------------------------------------------------------------
def test_v5_t16_industry_templates():
    """行业模板预设：INDUSTRY_TEMPLATES 含五大行业 + 字段结构 + special_fields 数量。"""
    sys.path.insert(0, os.path.join(SKILL_DIR, "scripts"))
    import bom_constants
    tpl = bom_constants.INDUSTRY_TEMPLATES
    # 必须包含电子/化工/纺织/家具/食品
    for k in ("电子", "化工", "纺织", "家具", "食品"):
        assert k in tpl, "T16 INDUSTRY_TEMPLATES 应含 %s" % k
    # 各自含 4 个结构字段
    for k, v in tpl.items():
        for fld in ("material_types", "standard", "special_fields", "preset_processes"):
            assert fld in v, "T16 %s 模板应含 %s" % (k, fld)
    # special_fields 数量正确
    assert len(tpl["电子"]["special_fields"]) == 10, \
        "T16 电子 special_fields 应=10, 实际 %d" % len(tpl["电子"]["special_fields"])
    assert len(tpl["化工"]["special_fields"]) == 8, \
        "T16 化工 special_fields 应=8, 实际 %d" % len(tpl["化工"]["special_fields"])
    assert len(tpl["纺织"]["special_fields"]) == 5, \
        "T16 纺织 special_fields 应=5, 实际 %d" % len(tpl["纺织"]["special_fields"])
    assert len(tpl["家具"]["special_fields"]) == 4, \
        "T16 家具 special_fields 应=4, 实际 %d" % len(tpl["家具"]["special_fields"])
    # 电子 10 字段内容断言
    assert tpl["电子"]["special_fields"] == [
        "designator", "footprint", "part_number", "rohs",
        "manufacturer", "tolerance", "rated_power",
        "rated_voltage", "alternate", "reflow_temp",
    ], "T16 电子 special_fields 内容不符"
    # 纺织 5 字段内容断言
    assert tpl["纺织"]["special_fields"] == [
        "composition", "yarn_count", "fabric_weight", "width", "color_no",
    ], "T16 纺织 special_fields 内容不符"
    # 家具 4 字段内容断言
    assert tpl["家具"]["special_fields"] == [
        "material_grade", "spec_size", "surface_treatment", "color_no",
    ], "T16 家具 special_fields 内容不符"


# ----------------------------------------------------------------------------
# V5-T17 ~ T21 逆向回收（round-trip）
# ----------------------------------------------------------------------------
def _assert_fields_recovered(parsed, orig, fields, cost=False, exclude_type=None):
    """通用：断言逆向回收字段与原始 JSON 一致。
    - fields: 需比对的专属字段集合
    - cost: 是否同时比对 unit_price/currency
    - exclude_type: 该类型物料跳过 fields 比对(被行业视图排除)
    """
    om = {m["name"]: m for m in orig["materials"]}
    dm = {m["name"]: m for m in parsed["materials"]}
    assert set(dm.keys()) == set(om.keys()), \
        "回收物料集合不一致: orig=%s back=%s" % (set(om.keys()), set(dm.keys()))
    for name, om_ in om.items():
        dm_ = dm[name]
        assert val_match(dm_.get("usage"), om_["usage"]), "用量不一致: %s" % name
        assert val_match(dm_.get("material_type"), om_.get("material_type", "")), \
            "material_type不一致: %s" % name
        if exclude_type is None or om_.get("material_type") != exclude_type:
            for f in fields:
                assert val_match(dm_.get(f), om_.get(f, "")), \
                    "字段 %s 不一致: %s (orig=%r back=%r)" % (f, name, om_.get(f, ""), dm_.get(f))
        if cost:
            assert val_match(dm_.get("unit_price"), om_.get("unit_price", "")), \
                "unit_price不一致: %s" % name
            assert val_match(dm_.get("currency"), om_.get("currency", "")), \
                "currency不一致: %s" % name
    # 每个物料均含 28 个特殊字段默认空串(由 import_bom 补全)
    from import_bom import _SPECIAL_FIELDS
    for m in parsed["materials"]:
        for f in _SPECIAL_FIELDS:
            assert f in m, "物料 %s 应含特殊字段 %s" % (m.get("name"), f)


def test_v5_t17_roundtrip_textile():
    """逆向回收-纺织：5 字段正确回填，值与原始一致。"""
    with open(SAMPLE_TEXTILE, encoding="utf-8") as f:
        orig = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t17_")
    out = os.path.join(wd, "BOM_t17.xlsx")
    rc, so, se = run_generate(orig, out)
    assert rc == 0, "T17 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T17 逆向应成功, rc=%s se2=%s" % (rc2, se2)
    assert parsed.get("industry") == "纺织", "T17 industry 应=纺织"
    # 纺织排除「其他」类("裁片"/"OPP包装袋")→ 其 5 字段不回放, 跳过比对
    _assert_fields_recovered(parsed, orig, TEXTILE_FIELDS, exclude_type="其他")


def test_v5_t18_roundtrip_furniture():
    """逆向回收-家具：4 字段正确回填，值与原始一致。"""
    with open(SAMPLE_FURNITURE, encoding="utf-8") as f:
        orig = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t18_")
    out = os.path.join(wd, "BOM_t18.xlsx")
    rc, so, se = run_generate(orig, out)
    assert rc == 0, "T18 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "BOM_t18.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T18 逆向应成功, rc=%s se2=%s" % (rc2, se2)
    assert parsed.get("industry") == "家具", "T18 industry 应=家具"
    # 家具排除「其他」类("板件"/"瓦楞包装纸箱")→ 跳过比对
    _assert_fields_recovered(parsed, orig, FURNITURE_FIELDS, exclude_type="其他")


def test_v5_t19_roundtrip_electronic():
    """逆向回收-电子：10 字段 + unit_price/currency 正确回填，一致。"""
    with open(SAMPLE_ELEC_V5, encoding="utf-8") as f:
        orig = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t19_")
    out = os.path.join(wd, "BOM_t19.xlsx")
    rc, so, se = run_generate(orig, out)
    assert rc == 0, "T19 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "BOM_t19.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T19 逆向应成功, rc=%s se2=%s" % (rc2, se2)
    assert parsed.get("industry") == "电子", "T19 industry 应=电子"
    # 电子排除「其他」类(裸PCB板/贴片完成板)→ 其元件字段不回放, 但成本字段回收
    _assert_fields_recovered(parsed, orig, ELEC_FIELDS, cost=True, exclude_type="其他")
    # 无 total_price 键
    for m in parsed["materials"]:
        assert "total_price" not in m, "T19 逆向不应含 total_price"


def test_v5_t20_roundtrip_chemical():
    """逆向回收-化工：8 字段正确回填，值与原始一致。"""
    with open(SAMPLE_CHEM_V5, encoding="utf-8") as f:
        orig = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t20_")
    out = os.path.join(wd, "BOM_t20.xlsx")
    rc, so, se = run_generate(orig, out)
    assert rc == 0, "T20 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "BOM_t20.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T20 逆向应成功, rc=%s se2=%s" % (rc2, se2)
    assert parsed.get("industry") == "化工", "T20 industry 应=化工"
    # 化工排除「包材」类(消毒液基料/喷雾瓶)→ 跳过比对
    _assert_fields_recovered(parsed, orig, CHEM_FIELDS, exclude_type="包材")


def test_v5_t21_roundtrip_cost():
    """逆向回收-成本：unit_price/currency 正确回填，一致；无 total_price。"""
    with open(SAMPLE_COST, encoding="utf-8") as f:
        orig = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t21_")
    out = os.path.join(wd, "BOM_t21.xlsx")
    rc, so, se = run_generate(orig, out)
    assert rc == 0, "T21 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "BOM_t21.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T21 逆向应成功, rc=%s se2=%s" % (rc2, se2)
    assert parsed.get("industry") == "通用", "T21 industry 应=通用"
    om = {m["name"]: m for m in orig["materials"]}
    dm = {m["name"]: m for m in parsed["materials"]}
    for name, om_ in om.items():
        dm_ = dm[name]
        assert val_match(dm_.get("unit_price"), om_.get("unit_price", "")), \
            "T21 unit_price不一致: %s" % name
        assert val_match(dm_.get("currency"), om_.get("currency", "")), \
            "T21 currency不一致: %s" % name
        assert "total_price" not in dm_, "T21 逆向不应含 total_price"
    for m in parsed["materials"]:
        assert "total_price" not in m, "T21 逆向 JSON 不应含 total_price 键"


# ----------------------------------------------------------------------------
# V5-T22 ~ T24 向后兼容
# ----------------------------------------------------------------------------
def test_v5_t22_backward_electronic():
    """向后兼容-电子(V4 旧示例)：仍 14 列，扩列 I–N 为空。"""
    with open(SAMPLE_ELEC_V4, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t22_")
    out = os.path.join(wd, "BOM_t22.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T22 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    headers, rows = read_block(ws, "三、元件清单")
    assert headers is not None, "T22 应存在元件清单"
    assert len(headers) == 14, "T22 电子仍应 14 列, 实际 %d" % len(headers)
    # V4 旧 JSON 无 manufacturer 等扩列字段 → I–N(制造商/容差/额定功率/额定电压/替代料/封装温度) 应为空
    by_name = {r["物料名称"]: r for r in rows}
    u = by_name.get("主控芯片")
    assert u is not None, "T22 应含主控芯片"
    for col in ("制造商", "容差", "额定功率", "额定电压", "替代料", "封装温度"):
        assert _blank(u[col]), "T22 旧 JSON 扩列字段 %s 应空" % col


def test_v5_t23_backward_chemical():
    """向后兼容-化工(V4 旧示例)：仍 13 列，扩列 I–M 为空。"""
    with open(SAMPLE_CHEM_V4, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t23_")
    out = os.path.join(wd, "BOM_t23.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T23 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    headers, rows = read_block(ws, "三、配方表")
    assert headers is not None, "T23 应存在配方表"
    assert len(headers) == 13, "T23 化工仍应 13 列, 实际 %d" % len(headers)
    # V4 旧 JSON 无 purity 等扩列字段 → I–M 应为空
    by_name = {r["物料名称"]: r for r in rows}
    w = by_name.get("去离子水")
    assert w is not None, "T23 应含去离子水"
    for col in ("纯度", "物态", "闪点", "存储条件", "危险等级"):
        assert _blank(w[col]), "T23 旧 JSON 扩列字段 %s 应空" % col


def test_v5_t24_backward_v3_food():
    """向后兼容-食品(V3 旧示例)：配料表照常，无 V5 新区块。"""
    with open(SAMPLE_V3, encoding="utf-8") as f:
        data = json.load(f)
    assert "industry" not in data, "T24 V3 样本应无 industry"
    wd = tempfile.mkdtemp(prefix="v5t24_")
    out = os.path.join(wd, "BOM_t24.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T24 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    # 食品(推断)仍生成配料表
    assert marker_row(ws, "三、配料表") is not None, "T24 V3 食品应生成配料表"
    # 不出现任何 V5 新区块
    assert marker_row(ws, "三、元件清单") is None, "T24 V3 不应有元件清单"
    assert marker_row(ws, "三、配方表") is None, "T24 V3 不应有配方表"
    assert marker_row(ws, "三、面料辅料清单") is None, "T24 V3 不应有面料辅料清单"
    assert marker_row(ws, "三、家具物料清单") is None, "T24 V3 不应有家具物料清单"
    assert marker_row(ws, "成本明细") is None, "T24 V3 不应有成本明细"


# ----------------------------------------------------------------------------
# V5-T25 最小变更核查
# ----------------------------------------------------------------------------
def test_v5_t25_minimum_change_warnings():
    """最小变更核查：生成 V4 旧格式 JSON 不新增任何 WARNING 类型。
    已知警告类型（V4 既有 + V5 同构提示）白名单：
    V8 / W2 / W3(CAS,GHS) / 含量和 / 排除-电子 / 排除-化工 /
    排除-纺织(同构) / 排除-家具(同构) / W1 / H1。
    """
    known = [
        "不在枚举内",            # V8
        "未标注 RoHS 合规状态",  # W2
        "未填写 CAS 号",         # W3
        "未填写 GHS 危险标识",   # W3
        "偏离 100%",            # 含量和
        "元件清单已排除",        # 排除-电子(同构)
        "配方表已排除",          # 排除-化工(同构)
        "面料辅料清单已排除",    # 排除-纺织(V5 同构提示)
        "家具物料清单已排除",    # 排除-家具(V5 同构提示)
        "过敏原标签",            # W1
        "名称疑似含致敏物",     # H1
    ]
    # 电子(V4 旧示例)
    with open(SAMPLE_ELEC_V4, encoding="utf-8") as f:
        data_e = json.load(f)
    wd = tempfile.mkdtemp(prefix="v5t25_")
    out_e = os.path.join(wd, "BOM_t25e.xlsx")
    rc, so_e, se = run_generate(data_e, out_e)
    assert rc == 0, "T25 电子生成应成功, rc=%s se=%s" % (rc, se)
    for line in so_e.splitlines():
        if "WARNING" in line:
            assert any(k in line for k in known), \
                "T25 电子出现未知 WARNING 类型: %s" % line
    # 化工(V4 旧示例)
    with open(SAMPLE_CHEM_V4, encoding="utf-8") as f:
        data_c = json.load(f)
    out_c = os.path.join(wd, "BOM_t25c.xlsx")
    rc, so_c, se = run_generate(data_c, out_c)
    assert rc == 0, "T25 化工生成应成功, rc=%s se=%s" % (rc, se)
    for line in so_c.splitlines():
        if "WARNING" in line:
            assert any(k in line for k in known), \
                "T25 化工出现未知 WARNING 类型: %s" % line


# ----------------------------------------------------------------------------
# V5-T26 py_compile
# ----------------------------------------------------------------------------
def test_v5_t26_py_compile():
    """三脚本编译通过 + 业务模块可 import（bom_constants / import_bom）。"""
    import py_compile
    for p in (GEN, IMP, CONST):
        try:
            py_compile.compile(p, doraise=True)
        except py_compile.PyCompileError as e:
            raise AssertionError("编译失败 %s: %s" % (p, e))
    # import 验证：往 sys.path 注入 scripts 目录
    sys.path.insert(0, os.path.join(SKILL_DIR, "scripts"))
    import bom_constants  # noqa: F401
    from import_bom import _SPECIAL_FIELDS  # noqa: F401
    assert len(_SPECIAL_FIELDS) == 28, \
        "T26 逆向特殊字段唯一键应为 28, 实际 %d" % len(_SPECIAL_FIELDS)
