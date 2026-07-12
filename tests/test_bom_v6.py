#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BOM 智造师 V6 增量 · 机械/包装专属视图 独立验证套件
==================================================

QA 工程师 Edward 独立编写，不依赖实现者自测结论。

覆盖 V6 增量核心能力（设计基线 incremental-design-v6.md）：
  1. 机械行业专属视图（8 列 A–H，无"物料类型"展示列 — Q2；weight/unit_weight
     两列独立 — Q1）
  2. 包装行业专属视图（8 列 A–H，含"物料类型"列）
  3. 机械/包装 正向生成 → 逆向导入闭环，专属字段正确回填
  4. 成本块双编号：机械/包装为「四、成本明细」，通用为「三、成本明细」
  5. _SPECIAL_FIELDS 唯一键精确 == 37（V5 28 + V6 净增 9）
  6. industry 推断（从"三、机械物料清单"/"三、包装物料清单"区块标记）
  7. 向后兼容：通用/无 industry 不生成机械/包装块；旧字段零改动可读

测试方式：subprocess 调用 CLI（generate_bom.py / import_bom.py）做端到端闭环，
          并对生成的 xlsx 用 openpyxl 直接读取做结构断言；
          另对纯函数 derive_mechanical/derive_packaging 与常量做单元测试。

被测脚本目录: C:/Users/姓名/.workbuddy/skills/bom-zhizao-shi/scripts
运行:
    C:\\Users\\姓名\\.workbuddy\\binaries\\python\\envs\\default\\Scripts\\python.exe test_bom_v6.py
"""

import json
import os
import shutil
import subprocess
import sys
import tempfile
import unittest

# ---------------------------------------------------------------------------
# 路径常量（与既有测试保持一致）
# ---------------------------------------------------------------------------
SKILL_DIR = r"C:\Users\姓名\.workbuddy\skills\bom-zhizao-shi"
SKILL_SCRIPTS = os.path.join(SKILL_DIR, "scripts")
GEN = os.path.join(SKILL_SCRIPTS, "generate_bom.py")
IMP = os.path.join(SKILL_SCRIPTS, "import_bom.py")
PY = r"C:\Users\姓名\.workbuddy\binaries\python\envs\default\Scripts\python.exe"
TMP = os.path.join(tempfile.gettempdir(), "bom_v6_test")
os.makedirs(TMP, exist_ok=True)

# 将脚本目录加入路径，便于直接 import 常量 / 纯函数做单测
if SKILL_SCRIPTS not in sys.path:
    sys.path.insert(0, SKILL_SCRIPTS)


def run_cli(script_path, *cli_args):
    """调用一个 CLI，返回 CompletedProcess。"""
    cmd = [PY, script_path] + list(cli_args)
    return subprocess.run(cmd, capture_output=True, text=True, cwd=TMP)


def find_marker_row(ws, marker):
    """定位首列包含 marker 的行号（1-based），未找到返回 None。"""
    for idx, row in enumerate(ws.iter_rows(min_col=1, max_col=1), 1):
        v = row[0].value
        if v is not None and marker in str(v):
            return idx
    return None


def read_block(ws, marker):
    """读取一个「三/四、」区块，返回 (headers, rows)。

    headers: 表头行 8 列值列表（A–H）。
    rows:    [{表头: 单元格值, ...}, ...]，仅含数据行（物料名非空、
            非「【」分组标题、非「合计」行），遇首个空名行即停止。
    """
    mrow = find_marker_row(ws, marker)
    if mrow is None:
        return None, []
    hrow = mrow + 1
    headers = [ws.cell(hrow, c).value for c in range(1, 9)]
    rows = []
    r = hrow + 1
    while r <= ws.max_row:
        nm = ws.cell(r, 2).value
        if nm is None or str(nm).strip() == "":
            break
        first = str(ws.cell(r, 1).value or "").strip()
        if first.startswith("【") or "合计" in first:
            break
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, 9)}
        rows.append(row)
        r += 1
    return headers, rows


def safe_rmtree(d):
    """删除临时目录；忽略环境安全删除守卫抛出的异常（不影响测试结果）。"""
    try:
        shutil.rmtree(d, ignore_errors=True)
    except BaseException:
        pass


def read_cost_block(ws):
    """读取成本明细块，返回 (label, rows)。label 为「三/四、成本明细」。"""
    mrow = find_marker_row(ws, "成本明细")
    if mrow is None:
        return None, []
    label = str(ws.cell(mrow, 1).value)
    hrow = mrow + 1
    headers = [ws.cell(hrow, c).value for c in range(1, 9)]
    rows = []
    r = hrow + 1
    while r <= ws.max_row:
        nm = ws.cell(r, 2).value
        if nm is None or str(nm).strip() == "":
            break
        first = str(ws.cell(r, 1).value or "").strip()
        if first.startswith("【") or "合计" in first:
            break
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, 9)}
        rows.append(row)
        r += 1
    return label, rows


# ---------------------------------------------------------------------------
# 共享测试数据
# ---------------------------------------------------------------------------
MECH_DATA = {
    "product_name": "减速机总成",
    "category": "工业品",
    "industry": "机械",
    "output_rate": 100,
    "version": "V1.0",
    "date": "2026-07-07",
    "materials": [
        {
            "name": "主轴", "unit": "件", "usage": 1, "yield_rate": 100,
            "erp_code": "RM-001", "material_type": "型材",
            "drawing_no": "DW-001", "material": "45钢",
            "heat_treatment": "调质", "surface_treatment": "镀锌",
            "weight": 5.0, "unit_weight": 2.5,
            "unit_price": 200.0, "currency": "",
        },
        {
            "name": "轴承", "unit": "件", "usage": 2, "yield_rate": 100,
            "erp_code": "RM-002", "material_type": "标准件",
            "drawing_no": "DW-002", "material": "GCr15",
            "heat_treatment": "淬火", "surface_treatment": "",
            "weight": 0.2, "unit_weight": 0.1,
            "unit_price": 50.0, "currency": "USD",
        },
        # material_type=="其他" → 机械视图应被排除，但基数物料区与逆向仍保留
        {
            "name": "辅材", "unit": "kg", "usage": 0.5, "yield_rate": 100,
            "erp_code": "RM-003", "material_type": "其他",
        },
    ],
    "processes": [],
}

PACK_DATA = {
    "product_name": "彩盒包装",
    "category": "工业品",
    "industry": "包装",
    "output_rate": 100,
    "version": "V1.0",
    "date": "2026-07-07",
    "materials": [
        {
            "name": "外箱", "unit": "个", "usage": 10, "yield_rate": 100,
            "erp_code": "PK-001", "material_type": "纸箱",
            "material": "牛皮纸", "basis_weight": 300,
            "size": "400x300x200", "print_process": "胶印",
            "eco_label": "FSC 认证",
            "unit_price": 3.5, "currency": "",
        },
        {
            "name": "内衬", "unit": "个", "usage": 20, "yield_rate": 100,
            "erp_code": "PK-002", "material_type": "缓冲",
            "material": "EPE", "basis_weight": 50,
            "size": "380x280x180", "print_process": "",
            "eco_label": "",
            "unit_price": 1.2, "currency": "",
        },
        {
            "name": "封箱带", "unit": "卷", "usage": 1, "yield_rate": 100,
            "erp_code": "PK-003", "material_type": "其他",
        },
    ],
    "processes": [],
}

GENERIC_DATA = {
    "product_name": "通用产品",
    "category": "工业品",
    "output_rate": 100,
    "version": "V1.0",
    "date": "2026-07-07",
    "materials": [
        {"name": "物料A", "unit": "kg", "usage": 1, "yield_rate": 100},
    ],
    "processes": [],
}


class BomV6ConstantsTest(unittest.TestCase):
    """常量 / 计数口径断言（不依赖 CLI）。"""

    def test_special_fields_is_37(self):
        import import_bom
        self.assertEqual(len(import_bom._SPECIAL_FIELDS), 37,
                         "V6 _SPECIAL_FIELDS 唯一键应精确为 37")
        self.assertEqual(len(set(import_bom._SPECIAL_FIELDS)), 37,
                         "_SPECIAL_FIELDS 不应有重复键")

    def test_special_fields_v6_keys_present(self):
        import import_bom
        s = set(import_bom._SPECIAL_FIELDS)
        # V6 净增 9 唯一键
        for k in ("drawing_no", "material", "heat_treatment",
                  "weight", "unit_weight", "basis_weight",
                  "size", "print_process", "eco_label"):
            self.assertIn(k, s, "V6 专属键 %s 应存在于 _SPECIAL_FIELDS" % k)
        # surface_treatment 由 V5 家具引入，V6 复用（只计一次）
        self.assertIn("surface_treatment", s)
        # material_type 为包装块结构列，不计入 _SPECIAL_FIELDS
        self.assertNotIn("material_type", s,
                         "material_type 为块结构列，不应计入 _SPECIAL_FIELDS")

    def test_enums_and_templates_present(self):
        import bom_constants as bc
        # MECHANICAL_TYPES / PACKAGING_TYPES 为物料子类枚举（交互引导，非行业名）
        self.assertIn("型材", bc.MECHANICAL_TYPES)
        self.assertIn("标准件", bc.MECHANICAL_TYPES)
        self.assertIn("其他", bc.MECHANICAL_TYPES)
        self.assertIn("纸箱", bc.PACKAGING_TYPES)
        self.assertIn("缓冲", bc.PACKAGING_TYPES)
        self.assertIn("其他", bc.PACKAGING_TYPES)
        # 行业总枚举含机械/包装
        self.assertIn("机械", bc.INDUSTRIES)
        self.assertIn("包装", bc.INDUSTRIES)
        self.assertEqual(bc.MECHANICAL_EXCLUDE, {"其他"})
        self.assertEqual(bc.PACKAGING_EXCLUDE, {"其他"})
        self.assertEqual(bc.INDUSTRY_STANDARD.get("机械"), "GB/T 1804-2000")
        self.assertEqual(bc.INDUSTRY_STANDARD.get("包装"), "GB/T 6543-2008")
        # 模板预设已填实（仅交互引导）
        self.assertIn("机械", bc.INDUSTRY_TEMPLATES)
        self.assertIn("包装", bc.INDUSTRY_TEMPLATES)
        self.assertTrue(bc.INDUSTRY_TEMPLATES["机械"]["special_fields"])
        self.assertTrue(bc.INDUSTRY_TEMPLATES["包装"]["special_fields"])

    def test_industry_view_set_includes_mechanical_packaging(self):
        import generate_bom as g
        self.assertIn("机械", g.INDUSTRY_VIEW_SET)
        self.assertIn("包装", g.INDUSTRY_VIEW_SET)
        # 7 行业视图集
        self.assertEqual(len(g.INDUSTRY_VIEW_SET), 7)


class BomV6DeriveUnitTest(unittest.TestCase):
    """derive_mechanical / derive_packaging 纯函数单元测试。"""

    def test_derive_mechanical_filters_other_and_sorts(self):
        import generate_bom as g
        data = {"materials": [
            {"name": "B件", "material_type": "标准件", "drawing_no": "x"},
            {"name": "A件", "material_type": "型材", "drawing_no": "y"},
            {"name": "C件", "material_type": "其他", "drawing_no": "z"},
        ]}
        items, excluded = g.derive_mechanical(data)
        self.assertEqual(len(items), 2, "material_type=='其他' 应被排除")
        self.assertEqual(len(excluded), 1)
        self.assertEqual(excluded[0]["name"], "C件")
        # 排序键 (material_type, name)：型材 < 标准件
        self.assertEqual([m["name"] for m in items], ["A件", "B件"])

    def test_derive_packaging_filters_other_and_sorts(self):
        import generate_bom as g
        data = {"materials": [
            {"name": "内衬", "material_type": "缓冲", "basis_weight": 50},
            {"name": "外箱", "material_type": "纸箱", "basis_weight": 300},
            {"name": "封箱带", "material_type": "其他", "basis_weight": 0},
        ]}
        items, excluded = g.derive_packaging(data)
        self.assertEqual(len(items), 2)
        self.assertEqual(len(excluded), 1)
        self.assertEqual(excluded[0]["name"], "封箱带")
        self.assertEqual([m["name"] for m in items], ["外箱", "内衬"])


class BomV6MechanicalTest(unittest.TestCase):
    """机械行业视图：正向结构 + 逆向回填 + 排除 + 行业推断。"""

    _dirs = []

    def _newdir(self):
        d = tempfile.mkdtemp(prefix="qa_v6_mech_", dir=TMP)
        self._dirs.append(d)
        return d

    @classmethod
    def tearDownClass(cls):
        for d in cls._dirs:
            safe_rmtree(d)

    def _gen(self, data, xlsx_path):
        d = os.path.dirname(xlsx_path)
        data_path = os.path.join(d, "data.json")
        with open(data_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        r = run_cli(GEN, "--data", data_path, "--out", xlsx_path)
        self.assertEqual(
            r.returncode, 0,
            msg="generate_bom.py 失败 (rc=%d)\nstdout=%s\nstderr=%s"
            % (r.returncode, r.stdout, r.stderr),
        )
        self.assertIn("OK:", r.stdout)
        return r

    def _imp_to_dict(self, xlsx_path):
        out_path = os.path.join(os.path.dirname(xlsx_path), "back.json")
        r = run_cli(IMP, "--in", xlsx_path, "--out", out_path)
        self.assertEqual(
            r.returncode, 0,
            msg="import_bom.py 失败 (rc=%d)\nstdout=%s\nstderr=%s"
            % (r.returncode, r.stdout, r.stderr),
        )
        self.assertIn("OK:", r.stdout)
        with open(out_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def test_01_mechanical_block_marker_and_headers(self):
        from openpyxl import load_workbook
        d = self._newdir()
        xlsx = os.path.join(d, "mech.xlsx")
        self._gen(MECH_DATA, xlsx)

        wb = load_workbook(xlsx)
        ws = wb["BOM表"]
        # 块标记存在
        self.assertIsNotNone(find_marker_row(ws, "三、机械物料清单"),
                             "应生成『三、机械物料清单』块")
        # 不应生成包装块
        self.assertIsNone(find_marker_row(ws, "三、包装物料清单"))

        headers, rows = read_block(ws, "三、机械物料清单")
        expected = ["序号", "物料名称", "图号", "材质", "热处理",
                    "表面处理", "重量(kg)", "单重(kg/件)"]
        self.assertEqual(headers, expected, "机械视图 8 列表头应精确匹配")
        # Q2：机械视图不应含"物料类型"展示列
        self.assertNotIn("物料类型", [h for h in headers if h],
                         "Q2：机械视图 8 列不应含『物料类型』展示列")
        # Q1：重量(kg) 与 单重(kg/件) 为独立两列（G/H）
        self.assertEqual(headers[6], "重量(kg)")
        self.assertEqual(headers[7], "单重(kg/件)")

    def test_02_mechanical_block_excludes_other(self):
        from openpyxl import load_workbook
        d = self._newdir()
        xlsx = os.path.join(d, "mech_excl.xlsx")
        self._gen(MECH_DATA, xlsx)

        wb = load_workbook(xlsx)
        ws = wb["BOM表"]
        _, rows = read_block(ws, "三、机械物料清单")
        # 仅 2 条非"其他"物料进视图
        self.assertEqual(len(rows), 2, "机械视图应排除 material_type=='其他' 的物料")
        names = {r.get("物料名称") for r in rows}
        self.assertNotIn("辅材", names, "『辅材』(其他类) 不应出现在机械视图")
        self.assertIn("主轴", names)
        self.assertIn("轴承", names)

    def test_03_mechanical_reverse_recovers_fields(self):
        d = self._newdir()
        xlsx = os.path.join(d, "mech_rt.xlsx")
        self._gen(MECH_DATA, xlsx)
        back = self._imp_to_dict(xlsx)

        self.assertEqual(back["industry"], "机械")
        mats = {m["name"]: m for m in back["materials"]}
        self.assertEqual(len(mats), 3)

        # 非"其他"物料：机械专属字段正确回填
        zhou = mats["主轴"]
        self.assertEqual(zhou["drawing_no"], "DW-001")
        self.assertEqual(zhou["material"], "45钢")
        self.assertEqual(zhou["heat_treatment"], "调质")
        self.assertEqual(zhou["surface_treatment"], "镀锌")
        # Q1：weight / unit_weight 独立且为 float
        self.assertIsInstance(zhou["weight"], float)
        self.assertIsInstance(zhou["unit_weight"], float)
        self.assertEqual(zhou["weight"], 5.0)
        self.assertEqual(zhou["unit_weight"], 2.5)
        self.assertNotEqual(zhou["weight"], zhou["unit_weight"])

        bearing = mats["轴承"]
        self.assertEqual(bearing["drawing_no"], "DW-002")
        self.assertEqual(bearing["heat_treatment"], "淬火")
        self.assertEqual(bearing["surface_treatment"], "")  # 空值保留空串
        self.assertEqual(bearing["weight"], 0.2)
        self.assertEqual(bearing["unit_weight"], 0.1)

        # "其他"类物料：机械字段应为默认空串（未被视图恢复）
        fu = mats["辅材"]
        self.assertEqual(fu["material_type"], "其他")
        self.assertEqual(fu["drawing_no"], "")
        self.assertEqual(fu["weight"], "")

    def test_04_infer_industry_mechanical(self):
        d = self._newdir()
        xlsx = os.path.join(d, "mech_inf.xlsx")
        self._gen(MECH_DATA, xlsx)
        back = self._imp_to_dict(xlsx)
        self.assertEqual(back["industry"], "机械",
                         "从『三、机械物料清单』区块应推断 industry=机械")


class BomV6PackagingTest(unittest.TestCase):
    """包装行业视图：正向结构 + 逆向回填 + 排除 + 行业推断。"""

    _dirs = []

    def _newdir(self):
        d = tempfile.mkdtemp(prefix="qa_v6_pack_", dir=TMP)
        self._dirs.append(d)
        return d

    @classmethod
    def tearDownClass(cls):
        for d in cls._dirs:
            safe_rmtree(d)

    def _gen(self, data, xlsx_path):
        d = os.path.dirname(xlsx_path)
        data_path = os.path.join(d, "data.json")
        with open(data_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        r = run_cli(GEN, "--data", data_path, "--out", xlsx_path)
        self.assertEqual(
            r.returncode, 0,
            msg="generate_bom.py 失败 (rc=%d)\nstdout=%s\nstderr=%s"
            % (r.returncode, r.stdout, r.stderr),
        )
        self.assertIn("OK:", r.stdout)
        return r

    def _imp_to_dict(self, xlsx_path):
        out_path = os.path.join(os.path.dirname(xlsx_path), "back.json")
        r = run_cli(IMP, "--in", xlsx_path, "--out", out_path)
        self.assertEqual(
            r.returncode, 0,
            msg="import_bom.py 失败 (rc=%d)\nstdout=%s\nstderr=%s"
            % (r.returncode, r.stdout, r.stderr),
        )
        self.assertIn("OK:", r.stdout)
        with open(out_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def test_01_packaging_block_marker_and_headers(self):
        from openpyxl import load_workbook
        d = self._newdir()
        xlsx = os.path.join(d, "pack.xlsx")
        self._gen(PACK_DATA, xlsx)

        wb = load_workbook(xlsx)
        ws = wb["BOM表"]
        self.assertIsNotNone(find_marker_row(ws, "三、包装物料清单"),
                             "应生成『三、包装物料清单』块")
        self.assertIsNone(find_marker_row(ws, "三、机械物料清单"))

        headers, rows = read_block(ws, "三、包装物料清单")
        expected = ["序号", "物料名称", "物料类型", "材质", "克重(g/m²)",
                    "尺寸", "印刷工艺", "环保标识"]
        self.assertEqual(headers, expected, "包装视图 8 列表头应精确匹配")
        # 包装视图 C 列应含"物料类型"
        self.assertEqual(headers[2], "物料类型",
                         "包装视图 C 列应为『物料类型』")

    def test_02_packaging_block_excludes_other(self):
        from openpyxl import load_workbook
        d = self._newdir()
        xlsx = os.path.join(d, "pack_excl.xlsx")
        self._gen(PACK_DATA, xlsx)

        wb = load_workbook(xlsx)
        ws = wb["BOM表"]
        _, rows = read_block(ws, "三、包装物料清单")
        self.assertEqual(len(rows), 2, "包装视图应排除 material_type=='其他' 的物料")
        names = {r.get("物料名称") for r in rows}
        self.assertNotIn("封箱带", names)
        self.assertIn("外箱", names)
        self.assertIn("内衬", names)

    def test_03_packaging_reverse_recovers_fields(self):
        d = self._newdir()
        xlsx = os.path.join(d, "pack_rt.xlsx")
        self._gen(PACK_DATA, xlsx)
        back = self._imp_to_dict(xlsx)

        self.assertEqual(back["industry"], "包装")
        mats = {m["name"]: m for m in back["materials"]}
        self.assertEqual(len(mats), 3)

        box = mats["外箱"]
        self.assertEqual(box["material_type"], "纸箱")
        self.assertEqual(box["material"], "牛皮纸")
        # basis_weight 应为 float
        self.assertIsInstance(box["basis_weight"], float)
        self.assertEqual(box["basis_weight"], 300.0)
        self.assertEqual(box["size"], "400x300x200")
        self.assertEqual(box["print_process"], "胶印")
        self.assertEqual(box["eco_label"], "FSC 认证")  # Q5 自由文本

        liner = mats["内衬"]
        self.assertEqual(liner["material_type"], "缓冲")
        self.assertEqual(liner["basis_weight"], 50.0)
        self.assertEqual(liner["print_process"], "")
        self.assertEqual(liner["eco_label"], "")

        tape = mats["封箱带"]
        self.assertEqual(tape["material_type"], "其他")
        self.assertEqual(tape["basis_weight"], "")  # 未被视图恢复

    def test_04_infer_industry_packaging(self):
        d = self._newdir()
        xlsx = os.path.join(d, "pack_inf.xlsx")
        self._gen(PACK_DATA, xlsx)
        back = self._imp_to_dict(xlsx)
        self.assertEqual(back["industry"], "包装",
                         "从『三、包装物料清单』区块应推断 industry=包装")


class BomV6CostDualNumberingTest(unittest.TestCase):
    """成本块双编号：机械/包装→四、成本明细；通用→三、成本明细。"""

    _dirs = []

    def _newdir(self):
        d = tempfile.mkdtemp(prefix="qa_v6_cost_", dir=TMP)
        self._dirs.append(d)
        return d

    @classmethod
    def tearDownClass(cls):
        for d in cls._dirs:
            safe_rmtree(d)

    def _gen(self, data, xlsx_path):
        d = os.path.dirname(xlsx_path)
        data_path = os.path.join(d, "data.json")
        with open(data_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        r = run_cli(GEN, "--data", data_path, "--out", xlsx_path)
        self.assertEqual(
            r.returncode, 0,
            msg="generate_bom.py 失败 (rc=%d)\nstdout=%s\nstderr=%s"
            % (r.returncode, r.stdout, r.stderr),
        )
        self.assertIn("OK:", r.stdout)
        return r

    def _imp_to_dict(self, xlsx_path):
        out_path = os.path.join(os.path.dirname(xlsx_path), "back.json")
        r = run_cli(IMP, "--in", xlsx_path, "--out", out_path)
        self.assertEqual(
            r.returncode, 0,
            msg="import_bom.py 失败 (rc=%d)\nstdout=%s\nstderr=%s"
            % (r.returncode, r.stdout, r.stderr),
        )
        self.assertIn("OK:", r.stdout)
        with open(out_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def test_01_mechanical_cost_label_is_four(self):
        from openpyxl import load_workbook
        d = self._newdir()
        xlsx = os.path.join(d, "cost_mech.xlsx")
        self._gen(MECH_DATA, xlsx)
        wb = load_workbook(xlsx)
        ws = wb["BOM表"]
        label, _ = read_cost_block(ws)
        self.assertEqual(label, "四、成本明细",
                         "机械行业带成本应为『四、成本明细』")

    def test_02_packaging_cost_label_is_four(self):
        from openpyxl import load_workbook
        d = self._newdir()
        xlsx = os.path.join(d, "cost_pack.xlsx")
        self._gen(PACK_DATA, xlsx)
        wb = load_workbook(xlsx)
        ws = wb["BOM表"]
        label, _ = read_cost_block(ws)
        self.assertEqual(label, "四、成本明细",
                         "包装行业带成本应为『四、成本明细』")

    def test_03_generic_cost_label_is_three(self):
        from openpyxl import load_workbook
        data = dict(GENERIC_DATA)
        data["materials"] = [dict(GENERIC_DATA["materials"][0])]
        data["materials"][0]["unit_price"] = 10.0
        d = self._newdir()
        xlsx = os.path.join(d, "cost_gen.xlsx")
        self._gen(data, xlsx)
        wb = load_workbook(xlsx)
        ws = wb["BOM表"]
        label, _ = read_cost_block(ws)
        self.assertEqual(label, "三、成本明细",
                         "通用行业带成本应为『三、成本明细』")

    def test_04_cost_reverse_recovers_unit_price_and_currency(self):
        d = self._newdir()
        xlsx = os.path.join(d, "cost_rt.xlsx")
        self._gen(MECH_DATA, xlsx)
        back = self._imp_to_dict(xlsx)
        mats = {m["name"]: m for m in back["materials"]}
        # unit_price 逆向为 float
        self.assertIsInstance(mats["主轴"]["unit_price"], float)
        self.assertEqual(mats["主轴"]["unit_price"], 200.0)
        # 未显式填 currency → 默认 人民币(CNY)
        self.assertEqual(mats["主轴"]["currency"], "人民币(CNY)")
        # 显式 currency 保留
        self.assertEqual(mats["轴承"]["currency"], "USD")
        self.assertEqual(mats["轴承"]["unit_price"], 50.0)

    def test_05_cost_total_formula_in_xlsx(self):
        from openpyxl import load_workbook
        d = self._newdir()
        xlsx = os.path.join(d, "cost_formula.xlsx")
        self._gen(MECH_DATA, xlsx)
        wb = load_workbook(xlsx)
        ws = wb["BOM表"]
        _, rows = read_cost_block(ws)
        # 总价(H) = round(用量 × 单价, 2)
        for r in rows:
            usage = float(r.get("用量") or 0)
            price = float(r.get("单价") or 0)
            expected = round(usage * price, 2)
            self.assertEqual(float(r.get("总价")), expected,
                             "成本总价应按 用量×单价 计算")


class BomV6BackwardCompatTest(unittest.TestCase):
    """向后兼容：通用 / 无 industry 不生成机械/包装块；旧字段零改动可读。"""

    _dirs = []

    def _newdir(self):
        d = tempfile.mkdtemp(prefix="qa_v6_bc_", dir=TMP)
        self._dirs.append(d)
        return d

    @classmethod
    def tearDownClass(cls):
        for d in cls._dirs:
            safe_rmtree(d)

    def _gen(self, data, xlsx_path):
        d = os.path.dirname(xlsx_path)
        data_path = os.path.join(d, "data.json")
        with open(data_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        r = run_cli(GEN, "--data", data_path, "--out", xlsx_path)
        self.assertEqual(
            r.returncode, 0,
            msg="generate_bom.py 失败 (rc=%d)\nstdout=%s\nstderr=%s"
            % (r.returncode, r.stdout, r.stderr),
        )
        self.assertIn("OK:", r.stdout)
        return r

    def _imp_to_dict(self, xlsx_path):
        out_path = os.path.join(os.path.dirname(xlsx_path), "back.json")
        r = run_cli(IMP, "--in", xlsx_path, "--out", out_path)
        self.assertEqual(
            r.returncode, 0,
            msg="import_bom.py 失败 (rc=%d)\nstdout=%s\nstderr=%s"
            % (r.returncode, r.stdout, r.stderr),
        )
        self.assertIn("OK:", r.stdout)
        with open(out_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def test_01_generic_no_mechanical_packaging_block(self):
        from openpyxl import load_workbook
        d = self._newdir()
        xlsx = os.path.join(d, "bc_gen.xlsx")
        self._gen(GENERIC_DATA, xlsx)
        wb = load_workbook(xlsx)
        ws = wb["BOM表"]
        self.assertIsNone(find_marker_row(ws, "三、机械物料清单"),
                          "通用行业不应生成机械视图块")
        self.assertIsNone(find_marker_row(ws, "三、包装物料清单"),
                          "通用行业不应生成包装视图块")
        back = self._imp_to_dict(xlsx)
        self.assertEqual(back["industry"], "通用",
                         "无 industry + category=工业品 → 推断 通用")

    def test_02_cost_block_alone_does_not_infer_mechanical_packaging(self):
        # 通用 + 成本：成本块存在，但 industry 不应被误判为机械/包装
        data = dict(GENERIC_DATA)
        data["materials"] = [dict(GENERIC_DATA["materials"][0])]
        data["materials"][0]["unit_price"] = 10.0
        d = self._newdir()
        xlsx = os.path.join(d, "bc_cost.xlsx")
        self._gen(data, xlsx)
        back = self._imp_to_dict(xlsx)
        self.assertEqual(back["industry"], "通用",
                         "仅有成本块不应推断为机械/包装")

    def test_03_mechanical_special_fields_default_empty_when_absent(self):
        # 机械行业但物料不填专属字段 → 逆向这些字段应为空串（不报错）
        data = {
            "product_name": "纯机械外壳",
            "category": "工业品",
            "industry": "机械",
            "output_rate": 100,
            "materials": [
                {"name": "壳体", "unit": "件", "usage": 1, "yield_rate": 100,
                 "material_type": "型材"},
            ],
            "processes": [],
        }
        d = self._newdir()
        xlsx = os.path.join(d, "bc_mech_empty.xlsx")
        self._gen(data, xlsx)
        back = self._imp_to_dict(xlsx)
        m = back["materials"][0]
        for k in ("drawing_no", "material", "heat_treatment",
                  "weight", "unit_weight", "surface_treatment"):
            self.assertEqual(m.get(k), "", "未填的机械字段逆向应为空串: %s" % k)


if __name__ == "__main__":
    unittest.main(verbosity=2)
