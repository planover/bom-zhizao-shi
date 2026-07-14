#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BOM 智造师 V7 增量 · 补强测试套件（QA 工程师严过关）
=========================================================================

对 tests/test_bom_v7.py（26 例）未覆盖的 V7 增量高风险点做补强。**全部为新增
独立用例，不修改任何 V2–V6 既有断言，也不改动导出行。**

补强点（对应主理人核验清单缺口）：
  1. 双语 sheet 数据行与中文 sheet 完全一致（仅 material_type 列按 ZH2EN 英文）；
     覆盖「一、物料信息」主块与「三、配料表」派生块。
  2. 行业视图双语列名正确：机械 8 列（图号→Drawing No. 等）、成本 8 列
     （单价→Unit Price / 币种→Currency / 总价→Total Price）的英文表头存在。
  3. B1 空白模板结构：列数随行业（电子 14 / 化工 13 / 其余 8）、区块标题/表头/
     空物料区/空工序区齐全，且可被 import_bom 解析为空 JSON（不崩溃）。
  4. 超长字段边界：名称/单位超长不崩溃且值完整保留（单语 + 双语）。
  5. B3 多文件 step_no 冲突聚合留痕：3 文件均含 S01/S02 → 4 条冲突备注
     （S01 在文件 1/2、1/3；S02 在文件 1/2、1/3），且不去重、不重命名。
  6. I18N 行业视图列名覆盖完整性：PRD §5.2 要求翻译的关键中文列头均在 ZH2EN 中。

测试方式：以 in-process 调用 build_workbook / run_blank_templates / _merge_boms 为主
（速度快、规避 Windows 连跑 subprocess 句柄累积崩溃），openpyxl 直接读取断言；
B3 合并经 _merge_boms 直接调用（内部走 _parse_bom_nofail 错误隔离路径）。
"""
import json
import os
import sys
import tempfile
import unittest

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# 路径常量（与既有测试保持一致）
# ---------------------------------------------------------------------------
SKILL_SCRIPTS = r"C:\Users\姓名\.workbuddy\skills\bom-zhizao-shi\scripts"
if SKILL_SCRIPTS not in sys.path:
    sys.path.insert(0, SKILL_SCRIPTS)

from bom_constants import ZH2EN  # noqa: E402
from generate_bom import build_workbook, run_blank_templates  # noqa: E402
from import_bom import _parse_bom_core, _merge_boms  # noqa: E402

TMP = os.path.join(tempfile.gettempdir(), "bom_v7_extra")
os.makedirs(TMP, exist_ok=True)


# ---------------------------------------------------------------------------
# 工具
# ---------------------------------------------------------------------------
def _find_marker_row(ws, substring):
    """定位首个首列含 substring 的行号（1-based），未找到返回 None。"""
    for idx, row in enumerate(ws.iter_rows(min_col=1, max_col=1), 1):
        v = row[0].value
        if v is not None and substring in str(v):
            return idx
    return None


def _block_data_rows(ws, marker, stop_markers):
    """从某区块提取「数据行」（首列为整数序号），跳过表头/分组副标题/合计/空行。

    Args:
        ws: Worksheet。
        marker: 区块标记子串（如 "一、物料信息"）。
        stop_markers: 区块结束边界标记子串列表（取首个命中行作为停止行）。

    Returns:
        list of [col1..colN 值]（按当前 sheet 最大列），仅数据行。
    """
    mr = _find_marker_row(ws, marker)
    if mr is None:
        return []
    stops = []
    for sm in stop_markers:
        s = _find_marker_row(ws, sm)
        if s is not None:
            stops.append(s)
    stop = min(stops) if stops else ws.max_row + 1
    rows = []
    r = mr + 1
    while r < stop:
        v2 = ws.cell(r, 2).value
        # 跳过表头行（中文 "物料名称" 或英文 "Material Name"）
        if v2 in ("物料名称", "Material Name"):
            r += 1
            continue
        v1 = ws.cell(r, 1).value
        if isinstance(v1, int):
            ncol = ws.max_column
            rows.append([ws.cell(r, c).value for c in range(1, ncol + 1)])
        r += 1
    return rows


def _compare_block_consistency(ws_zh, ws_en, marker, stop_markers, mt_col):
    """断言双语 sheet 的某区块数据行与中文 sheet 完全一致（仅 material_type 列翻译）。"""
    rows_zh = _block_data_rows(ws_zh, marker, stop_markers)
    rows_en = _block_data_rows(ws_en, marker, stop_markers)
    if len(rows_zh) != len(rows_en):
        raise AssertionError(
            "区块『%s』数据行数不一致: zh=%d en=%d" % (marker, len(rows_zh), len(rows_en)))
    for zv, ev in zip(rows_zh, rows_en):
        if len(zv) != len(ev):
            raise AssertionError("区块『%s』行长度不一致" % marker)
        for c in range(1, len(zv) + 1):
            if c == mt_col:
                continue
            if zv[c - 1] != ev[c - 1]:
                raise AssertionError(
                    "区块『%s』列 %d 不一致: %r vs %r" % (marker, c, zv[c - 1], ev[c - 1]))
        mt_zh = zv[mt_col - 1]
        mt_en = ev[mt_col - 1]
        if mt_zh in ZH2EN:
            if mt_en != ZH2EN[mt_zh]:
                raise AssertionError(
                    "区块『%s』material_type『%s』应为 %s，实际 %s"
                    % (marker, mt_zh, ZH2EN[mt_zh], mt_en))
        else:
            if mt_en != mt_zh:
                raise AssertionError(
                    "区块『%s』material_type 查不到英文应保留中文『%s』，实际 %s"
                    % (marker, mt_zh, mt_en))


# ---------------------------------------------------------------------------
# 样本数据
# ---------------------------------------------------------------------------
def _sample_food(product_name="芒果果味糖浆", date="2026-07-07"):
    return {
        "product_name": product_name,
        "category": "食品",
        "industry": "食品",
        "output_rate": 130,
        "version": "V1.0",
        "date": date,
        "approver": "张三",
        "effective_date": "2026-07-10",
        "standard": "GB 7718-2025",
        "materials": [
            {"name": "芒果原浆", "unit": "kg", "usage": 46.3, "yield_rate": 55,
             "erp_code": "RM-001", "material_type": "原料", "process": "S01",
             "allergen": "大豆,乳"},
            {"name": "白砂糖", "unit": "kg", "usage": 30.0, "yield_rate": 100,
             "erp_code": "RM-002", "material_type": "原料", "process": "S01"},
        ],
        "processes": [
            {"step_no": "S01", "name": "调配", "desc": "混合搅拌", "work_hours": 30,
             "note": "常温", "output": "芒果果味糖浆基料"},
            {"step_no": "S02", "name": "灌装", "desc": "无菌灌装", "work_hours": 20,
             "note": "", "output": "芒果果味糖浆"},
        ],
    }


def _sample_mech(product_name="支架组件", date="2026-07-08"):
    return {
        "product_name": product_name,
        "category": "工业品",
        "industry": "机械",
        "output_rate": 95,
        "version": "V1.0",
        "date": date,
        "materials": [
            {"name": "钢板", "unit": "kg", "usage": 1.0, "yield_rate": 100,
             "erp_code": "MT-001", "material_type": "主材", "process": "S01",
             "drawing_no": "DW-1", "material": "Q235", "weight": 1.0, "unit_weight": 1.0},
            {"name": "钢板件", "unit": "kg", "usage": 1.0, "yield_rate": 100,
             "erp_code": "MT-100", "material_type": "主材", "process": "S02"},
            {"name": "螺栓", "unit": "个", "usage": 10, "yield_rate": 100,
             "erp_code": "MT-002", "material_type": "辅材", "process": "S02"},
        ],
        "processes": [
            {"step_no": "S01", "name": "切割", "desc": "锯切", "work_hours": 5,
             "note": "", "output": "钢板件"},
            {"step_no": "S02", "name": "焊接", "desc": "点焊", "work_hours": 8,
             "note": "", "output": "支架组件"},
        ],
    }


# ===========================================================================
# 1 + 2. 双语 sheet 数据一致 + 行业视图双语列名
# ===========================================================================
class TestBilingualConsistency(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp(prefix="bom_v7_ext_bi_", dir=TMP)

    def _wb_pair(self, data):
        wb_zh = build_workbook(data, bilingual=False)
        wb_en = build_workbook(data, bilingual=True)
        return wb_zh["BOM表"], wb_en["BOM表(英)"]

    def test_material_block_data_identical_except_type(self):
        """一、物料信息：数据行与中文 sheet 逐列一致，仅 material_type(列7) 翻译。"""
        ws_zh, ws_en = self._wb_pair(_sample_food())
        _compare_block_consistency(
            ws_zh, ws_en, "一、物料信息", ["二、工艺工序"], mt_col=7)

    def test_ingredient_block_data_identical_except_type(self):
        """三、配料表：数据行与中文 sheet 逐列一致，仅 material_type(列2) 翻译。"""
        ws_zh, ws_en = self._wb_pair(_sample_food())
        _compare_block_consistency(
            ws_zh, ws_en, "三、配料表", ["成本明细"], mt_col=2)

    def test_mech_view_bilingual_columns(self):
        """机械视图双语列名：图号→Drawing No.、热处理→Heat Treatment 等。"""
        _, ws_en = self._wb_pair(_sample_mech())
        mrow = _find_marker_row(ws_en, "三、机械物料清单")
        self.assertIsNotNone(mrow)
        self.assertEqual(ws_en.cell(mrow, 1).value,
                         "三、机械物料清单 (III. Mechanical BOM List)")
        zh_header = [ws_en.cell(mrow + 1, c).value for c in range(1, 9)]
        en_header = [ws_en.cell(mrow + 2, c).value for c in range(1, 9)]
        self.assertIn("图号", zh_header)
        self.assertIn("Drawing No.", en_header)
        self.assertIn("热处理", zh_header)
        self.assertIn("Heat Treatment", en_header)
        self.assertIn("单重(kg/件)", zh_header)
        self.assertIn("Unit Weight(kg/pc)", en_header)

    def test_cost_view_bilingual_columns(self):
        """成本视图双语列名：单价→Unit Price、币种→Currency、总价→Total Price。"""
        data = _sample_food()
        data["materials"][0]["unit_price"] = 10.5
        data["materials"][0]["currency"] = "人民币(CNY)"
        _, ws_en = self._wb_pair(data)
        # 食品 ∈ INDUSTRY_VIEW_SET → 四、成本明细
        crow = _find_marker_row(ws_en, "成本明细")
        self.assertIsNotNone(crow)
        self.assertEqual(ws_en.cell(crow, 1).value, "四、成本明细 (IV. Cost Detail)")
        zh_header = [ws_en.cell(crow + 1, c).value for c in range(1, 9)]
        en_header = [ws_en.cell(crow + 2, c).value for c in range(1, 9)]
        self.assertIn("单价", zh_header)
        self.assertIn("Unit Price", en_header)
        self.assertIn("币种", zh_header)
        self.assertIn("Currency", en_header)
        self.assertIn("总价", zh_header)
        self.assertIn("Total Price", en_header)


# ===========================================================================
# 3. B1 空白模板结构
# ===========================================================================
class TestBlankTemplateStructure(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp(prefix="bom_v7_ext_b1_", dir=TMP)

    def test_column_counts_by_industry(self):
        """列数随行业：电子 14 / 化工 13 / 其余 8。"""
        run_blank_templates(self.d, None, bilingual=False)
        files = sorted(os.listdir(self.d))
        self.assertEqual(files, sorted("template_%s.xlsx" % i for i in
                          ["食品", "电子", "化工", "机械", "纺织", "家具", "包装", "通用"]))
        self.assertEqual(load_workbook(os.path.join(self.d, "template_电子.xlsx"))["BOM表"].max_column, 14)
        self.assertEqual(load_workbook(os.path.join(self.d, "template_化工.xlsx"))["BOM表"].max_column, 13)
        for ind in ("食品", "机械", "纺织", "家具", "包装", "通用"):
            self.assertEqual(
                load_workbook(os.path.join(self.d, "template_%s.xlsx" % ind))["BOM表"].max_column, 8,
                "行业 %s 应为 8 列" % ind)

    def test_mech_template_structure_and_importable(self):
        """机械模板：区块齐全、空物料/空工序区、可被逆向解析为空 JSON。"""
        run_blank_templates(self.d, ["机械"], bilingual=False)
        path = os.path.join(self.d, "template_机械.xlsx")
        ws = load_workbook(path)["BOM表"]
        self.assertIsNotNone(_find_marker_row(ws, "一、物料信息"))
        self.assertIsNotNone(_find_marker_row(ws, "二、工艺工序"))
        self.assertIsNotNone(_find_marker_row(ws, "三、机械物料清单"))
        # 空物料区 / 空工序区
        self.assertEqual(_block_data_rows(ws, "一、物料信息", ["二、工艺工序"]), [])
        self.assertEqual(_block_data_rows(ws, "二、工艺工序", ["三、机械物料清单"]), [])
        # 空模板可被 import 解析为空（不崩溃）
        data = _parse_bom_core(path)
        self.assertEqual(data["materials"], [])
        self.assertEqual(data["processes"], [])
        self.assertEqual(data["industry"], "机械")


# ===========================================================================
# 4. 超长字段边界
# ===========================================================================
class TestLongFieldBoundary(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp(prefix="bom_v7_ext_long_", dir=TMP)

    def test_long_name_and_unit_preserved(self):
        """超长名称/单位：不崩溃且值完整保留（单语 + 双语）。"""
        data = _sample_food()
        long_name = "超长物料名称" * 60   # 360 字符
        long_unit = "千克" * 120
        data["materials"][0]["name"] = long_name
        data["materials"][0]["unit"] = long_unit
        # 单语
        wb = build_workbook(data, bilingual=False)
        ws = wb["BOM表"]
        names = [r[1] for r in _block_data_rows(ws, "一、物料信息", ["二、工艺工序"])]
        self.assertIn(long_name, names)
        # 双语
        wb2 = build_workbook(data, bilingual=True)
        ws2 = wb2["BOM表(英)"]
        names2 = [r[1] for r in _block_data_rows(ws2, "一、物料信息", ["二、工艺工序"])]
        self.assertIn(long_name, names2)


# ===========================================================================
# 5. B3 多文件 step_no 冲突聚合留痕
# ===========================================================================
class TestMergeMultiConflict(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp(prefix="bom_v7_ext_b3_", dir=TMP)
        self.paths = []
        for i in range(3):
            p = os.path.join(self.d, "m%d.xlsx" % i)
            wb = build_workbook(_sample_mech("机械%d" % i, "2026-07-0%d" % (i + 1)))
            wb.save(p)
            self.paths.append(p)

    def test_three_file_conflict_aggregation(self):
        """3 文件均含 S01/S02：顺序拼接不去重(9 物料/6 工序)，冲突留痕 4 条，不重命名。"""
        out = os.path.join(self.d, "merged.json")
        _merge_boms(self.paths, out)
        with open(out, encoding="utf-8") as f:
            data = json.load(f)
        self.assertEqual(len(data["materials"]), 9)
        self.assertEqual(len(data["processes"]), 6)
        self.assertEqual(data["industry"], "机械")
        self.assertEqual(data["merged_from"], ["机械", "机械", "机械"])
        notes = data.get("merge_notes", [])
        self.assertTrue(any("S01" in n and "1/2" in n for n in notes), notes)
        self.assertTrue(any("S01" in n and "1/3" in n for n in notes), notes)
        self.assertTrue(any("S02" in n and "1/2" in n for n in notes), notes)
        self.assertTrue(any("S02" in n and "1/3" in n for n in notes), notes)
        # 不重命名：工序仍原序
        self.assertEqual([p["step_no"] for p in data["processes"]],
                         ["S01", "S02", "S01", "S02", "S01", "S02"])


# ===========================================================================
# 6. I18N 行业视图列名覆盖完整性（常量守卫）
# ===========================================================================
class TestI18NViewCoverage(unittest.TestCase):
    def test_industry_view_headers_translatable(self):
        """PRD §5.2 要求翻译的关键中文列头均应在 ZH2EN 中可译。"""
        expected = {
            "机械": ["图号", "材质", "热处理", "表面处理", "重量(kg)", "单重(kg/件)"],
            "包装": ["材质", "克重(g/m²)", "尺寸", "印刷工艺", "环保标识"],
            "电子": ["位号(Designator)", "型号(Part#)", "封装(Footprint)", "RoHS",
                    "制造商", "容差", "额定功率", "额定电压", "替代料", "封装温度"],
            "化工": ["CAS号", "含量(%)", "GHS标识", "计量单位", "纯度", "物态",
                    "闪点", "存储条件", "危险等级"],
            "纺织": ["成分比例", "纱支", "克重(g/m²)", "幅宽", "色号"],
            "家具": ["材质等级", "尺寸规格", "表面处理", "色号/花色"],
            "食品": ["计量单位", "用量占比%", "过敏原"],
            "成本": ["单价", "币种", "总价"],
        }
        for ind, headers in expected.items():
            for h in headers:
                self.assertIn(h, ZH2EN,
                              "行业视图列名『%s』(行业 %s) 未在 ZH2EN 中可译" % (h, ind))


if __name__ == "__main__":
    unittest.main(verbosity=2)
