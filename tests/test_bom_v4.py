#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BOM智造师 V4 (多行业扩展) 独立测试脚本（QA · 严过关）

不依赖 pytest，使用 subprocess 调用 generate_bom.py / import_bom.py，
并用 openpyxl 读回 xlsx 断言；assert + 退出码；打印 PASS/FAIL 汇总。

覆盖 V4 全部新增行为 + 原有 V1–V7 校验回归：
  V4-T01 现有食品 JSON(无industry)→推断食品→配料表照常
  V4-T02 显式 industry=电子 → 生成元件清单
  V4-T03 显式 industry=化工 → 生成配方表
  V4-T04 显式 industry=通用 → 不生成任何专属视图
  V4-T05 非食品 JSON 无 industry(工业品)→推断通用→无专属视图
  V4-T06 非法 industry(航天)→V8 WARNING+回退推断+退出码0
  V4-T07 电子示例→含「三、元件清单」区块,8列表头正确
  V4-T08 排除 material_type=其他 的物料
  V4-T09 排序正确(物料类型升序→位号字母数字升序,空位号排末尾)
  V4-T10 RoHS 着色: 否→红, 未知/空→黄(openpyxl 读 cell.font.color.rgb)
  V4-T11 物料区8列不含 designator/footprint/part_number/rohs
  V4-T12 化工示例→含「三、配方表」区块,8列表头正确
  V4-T13 排除 material_type=包材 的物料
  V4-T14 排序正确(含量降序)
  V4-T15 含量(%) 数字格式 0.0"%"
  V4-T16 物料区8列不含 cas_number/concentration/ghs_hazard
  V4-T17 W2: 电子物料未标 rohs → WARNING + 退出码0
  V4-T18 W3: 化工物料未填 cas_number/ghs_hazard → WARNING + 退出码0
  V4-T19 含量(%) 列和偏离100%超过±5% → WARNING + 退出码0
  V4-T20 含量(%) 列和=100% → 无 WARNING
  V4-T21 电子 xlsx 逆向→industry=电子, 4字段正确回收
  V4-T22 化工 xlsx 逆向→industry=化工, 3字段正确回收
  V4-T23 旧版 V3 xlsx 逆向→industry=食品, 7专属字段默认空串, 不报错
  V4-T24 逆向 JSON 不含『序号』『用量占比%』键
  V4-T25 product_name 空 → 退出码2 (V1)
  V4-T26 output_rate<=0 → 退出码2 (V3)
  V4-T27 工序流转链断裂(R3) → 退出码2
  V4-T28 电子: generate→import→关键字段(名称+designator+rohs+processes)一致
  V4-T29 化工: generate→import→关键字段(名称+cas+conc+ghs+processes)一致
  V4-T30 食品/电子/化工示例出品率显示均为带1位小数(130.0% / 100.0%)

运行：
  python3 tests/test_bom_v4.py
退出码：全部通过 0，否则 1。
"""

import os
import sys
import json
import shutil
import subprocess
import tempfile

from openpyxl import load_workbook

# ----------------------------------------------------------------------------
# 路径常量
# ----------------------------------------------------------------------------
SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GEN = os.path.join(SKILL_DIR, "scripts", "generate_bom.py")
IMP = os.path.join(SKILL_DIR, "scripts", "import_bom.py")
SAMPLE_V3 = os.path.join(SKILL_DIR, "examples", "sample_bom_v3.json")
SAMPLE_V3_XLSX = os.path.join(SKILL_DIR, "examples", "sample_bom_v3.xlsx")
SAMPLE_ELEC = os.path.join(SKILL_DIR, "examples", "sample_bom_v4_electronic.json")
SAMPLE_CHEM = os.path.join(SKILL_DIR, "examples", "sample_bom_v4_chemical.json")
PYTHON = r"C:\Users\姓名\.workbuddy\binaries\python\versions\3.13.12\python.exe"

# V4 7 个物料级专属字段
SPECIAL_FIELDS = [
    "designator", "footprint", "part_number", "rohs",
    "cas_number", "concentration", "ghs_hazard",
]

# 物料区 8 列固定表头（永不变）
MATERIAL_HEADERS = [
    "序号", "物料名称", "单位", "用量", "出品率(%)",
    "ERP物料代码", "物料类型", "所属工序",
]


# ----------------------------------------------------------------------------
# 运行辅助
# ----------------------------------------------------------------------------
def run_generate(data, out_path):
    """把 dict 写成临时 json，调用 generate_bom.py，返回 (rc, stdout, stderr)。"""
    tmp = out_path + ".in.json"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    p = subprocess.run(
        [PYTHON, GEN, "--data", tmp, "--out", out_path],
        capture_output=True, text=True,
    )
    return p.returncode, p.stdout, p.stderr


def run_import(xlsx_path, out_json=None):
    """调用 import_bom.py，返回 (rc, stdout, stderr, parsed_json_or_None)。"""
    args = [PYTHON, IMP, "--in", xlsx_path]
    if out_json:
        args += ["--out", out_json]
    p = subprocess.run(args, capture_output=True, text=True)
    parsed = None
    if out_json and os.path.exists(out_json):
        with open(out_json, encoding="utf-8") as f:
            parsed = json.load(f)
    return p.returncode, p.stdout, p.stderr, parsed


def load(path):
    return load_workbook(path).active


def flat_values(ws):
    """返回工作表中所有非空单元格的字符串列表。"""
    vals = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is not None and str(v).strip() != "":
                vals.append(str(v))
    return vals


def marker_row(ws, marker):
    """返回首列包含 marker 的行号（1-based），未找到返回 None。"""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and marker in str(v):
            return r
    return None


def material_header_row(ws):
    """定位物料表头行（某列值为『物料名称』）。"""
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value or "").strip() == "物料名称":
            return r
    return None


def total_usage_row(ws):
    """返回合计行（A 列==合计）的行号，未找到返回 None。"""
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == "合计":
            return r
    return None


def ingredient_block(ws):
    """读取『三、配料表』数据行，返回 [(name, pct, allergen), ...]。"""
    m = marker_row(ws, "三、配料表")
    if m is None:
        return []
    rows = []
    r = m + 2
    while r <= ws.max_row:
        v = ws.cell(r, 1).value
        if v is None or str(v).strip() == "":
            break
        if str(v).strip().startswith("【"):
            r += 1
            continue
        name = str(ws.cell(r, 1).value or "").strip()
        pct = ws.cell(r, 6).value
        allergen = str(ws.cell(r, 7).value or "").strip()
        rows.append((name, pct, allergen))
        r += 1
    return rows


def component_block(ws):
    """读取『三、元件清单』数据行，返回 [(seq,designator,part,footprint,name,usage,mtype,rohs), ...]。"""
    m = marker_row(ws, "三、元件清单")
    if m is None:
        return None
    rows = []
    r = m + 2
    while r <= ws.max_row:
        v = ws.cell(r, 1).value
        if v is None or str(v).strip() == "":
            break
        if str(v).strip().startswith("【"):
            r += 1
            continue
        rows.append((
            ws.cell(r, 1).value,
            str(ws.cell(r, 2).value or ""),
            str(ws.cell(r, 3).value or ""),
            str(ws.cell(r, 4).value or ""),
            str(ws.cell(r, 5).value or ""),
            ws.cell(r, 6).value,
            str(ws.cell(r, 7).value or ""),
            str(ws.cell(r, 8).value or ""),
        ))
        r += 1
    return rows


def formula_block(ws):
    """读取『三、配方表』数据行，返回 [(seq,name,cas,conc,ghs,mtype,unit,usage), ...]。"""
    m = marker_row(ws, "三、配方表")
    if m is None:
        return None
    rows = []
    r = m + 2
    while r <= ws.max_row:
        v = ws.cell(r, 1).value
        if v is None or str(v).strip() == "":
            break
        if str(v).strip().startswith("【"):
            r += 1
            continue
        rows.append((
            ws.cell(r, 1).value,
            str(ws.cell(r, 2).value or ""),
            str(ws.cell(r, 3).value or ""),
            ws.cell(r, 4).value,
            str(ws.cell(r, 5).value or ""),
            str(ws.cell(r, 6).value or ""),
            str(ws.cell(r, 7).value or ""),
            ws.cell(r, 8).value,
        ))
        r += 1
    return rows


def row4_text(ws):
    return " | ".join(str(ws.cell(4, c).value or "") for c in range(1, 9))


def material_region_values(ws):
    """返回物料区（表头行+1 至『二、工艺工序』标记行前）所有单元格字符串列表。"""
    hdr = material_header_row(ws)
    proc = marker_row(ws, "二、工艺工序")
    assert hdr is not None, "物料表头行未找到"
    assert proc is not None, "工艺工序标记行未找到"
    vals = []
    for r in range(hdr + 1, proc):
        for c in range(1, 9):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                vals.append(str(v))
    return vals


# ----------------------------------------------------------------------------
# 用例
# ----------------------------------------------------------------------------
def test_v4_t01_food_infer_ingredient():
    """现有食品 JSON(无 industry)→推断食品→配料表照常生成。"""
    with open(SAMPLE_V3, encoding="utf-8") as f:
        data = json.load(f)
    assert "industry" not in data, "T01 样本应无 industry 字段"
    wd = tempfile.mkdtemp(prefix="v4t01_")
    out = os.path.join(wd, "BOM_t01.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T01 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    assert marker_row(ws, "三、配料表") is not None, \
        "T01 食品(无industry)应推断食品并生成配料表"
    block = ingredient_block(ws)
    assert len(block) > 0, "T01 配料表应含可食用行"


def test_v4_t02_explicit_electronic():
    """显式 industry=电子 → 生成元件清单。"""
    with open(SAMPLE_ELEC, encoding="utf-8") as f:
        data = json.load(f)
    assert data.get("industry") == "电子", "T02 样本 industry 应=电子"
    wd = tempfile.mkdtemp(prefix="v4t02_")
    out = os.path.join(wd, "BOM_t02.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T02 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    assert marker_row(ws, "三、元件清单") is not None, "T02 应生成『三、元件清单』"


def test_v4_t03_explicit_chemical():
    """显式 industry=化工 → 生成配方表。"""
    with open(SAMPLE_CHEM, encoding="utf-8") as f:
        data = json.load(f)
    assert data.get("industry") == "化工", "T03 样本 industry 应=化工"
    wd = tempfile.mkdtemp(prefix="v4t03_")
    out = os.path.join(wd, "BOM_t03.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T03 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    assert marker_row(ws, "三、配方表") is not None, "T03 应生成『三、配方表』"


def test_v4_t04_generic_no_view():
    """显式 industry=通用 → 不生成任何专属视图（仅物料区+工序区+合计行）。"""
    wd = tempfile.mkdtemp(prefix="v4t04_")
    out = os.path.join(wd, "BOM_t04.xlsx")
    data = {
        "product_name": "通用品", "category": "工业品", "industry": "通用",
        "output_rate": 100,
        "materials": [
            {"name": "原料A", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S01"},
            {"name": "包材B", "unit": "个", "usage": 5, "yield_rate": 100,
             "material_type": "包材"},
        ],
        "processes": [{"step_no": "S01", "name": "p1", "output": "OUT1"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T04 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    assert marker_row(ws, "三、元件清单") is None, "T04 通用不应有元件清单"
    assert marker_row(ws, "三、配方表") is None, "T04 通用不应有配方表"
    assert marker_row(ws, "三、配料表") is None, "T04 通用不应有配料表"
    assert marker_row(ws, "二、工艺工序") is not None, "T04 应仍有工艺工序区"
    assert total_usage_row(ws) is not None, "T04 应仍有合计用量行"


def test_v4_t05_nonfood_infer_generic():
    """非食品 JSON 无 industry(工业品)→推断通用→无专属视图。"""
    wd = tempfile.mkdtemp(prefix="v4t05_")
    out = os.path.join(wd, "BOM_t05.xlsx")
    data = {
        "product_name": "工业品X", "category": "工业品", "output_rate": 100,
        "materials": [
            {"name": "原料A", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S01"},
            {"name": "包材B", "unit": "个", "usage": 5, "yield_rate": 100,
             "material_type": "包材"},
        ],
        "processes": [{"step_no": "S01", "name": "p1", "output": "OUT1"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T05 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    assert marker_row(ws, "三、元件清单") is None, "T05 通用推断不应有元件清单"
    assert marker_row(ws, "三、配方表") is None, "T05 通用推断不应有配方表"
    assert marker_row(ws, "三、配料表") is None, "T05 通用推断不应有配料表"


def test_v4_t06_illegal_industry_warning():
    """非法 industry(航天)→V8 WARNING + 回退推断 + 退出码0。"""
    wd = tempfile.mkdtemp(prefix="v4t06_")
    out = os.path.join(wd, "BOM_t06.xlsx")
    data = {
        "product_name": "非法行业测试", "category": "食品", "industry": "航天",
        "output_rate": 130,
        "materials": [
            {"name": "芒果原浆", "unit": "kg", "usage": 46.3, "yield_rate": 55,
             "material_type": "原料", "process": "S01", "allergen": "大豆,乳"},
            {"name": "白砂糖", "unit": "kg", "usage": 30, "yield_rate": 100,
             "material_type": "原料", "process": "S01"},
        ],
        "processes": [{"step_no": "S01", "name": "调配", "output": "成品"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T06 非法industry不应阻断(rc=0), rc=%s so=%s" % (rc, so)
    assert "WARNING" in so, "T06 应打印V8 WARNING, stdout=%s" % so
    assert "不在枚举内" in so, "T06 WARNING应提示不在枚举内, stdout=%s" % so
    assert "OK:" in so, "T06 仍应输出 OK:(非阻断), 实际=%s" % so
    ws = load(out)
    # 回退: category=食品 → 食品 → 配料表应生成
    assert marker_row(ws, "三、配料表") is not None, \
        "T06 回退推断为食品应生成配料表"


def test_v4_t07_electronic_header():
    """电子示例生成→含『三、元件清单』区块, 8列表头正确。"""
    with open(SAMPLE_ELEC, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t07_")
    out = os.path.join(wd, "BOM_t07.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T07 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    m = marker_row(ws, "三、元件清单")
    assert m is not None, "T07 应含『三、元件清单』"
    hdr = [str(ws.cell(m + 1, c).value or "").strip() for c in range(1, 9)]
    expected = ["序号", "位号(Designator)", "型号(Part#)", "封装(Footprint)",
                "物料名称", "数量", "物料类型", "RoHS"]
    assert hdr == expected, "T07 元件清单8列表头不符: %s" % hdr


def test_v4_t08_exclude_other():
    """排除 material_type=其他 的物料（验证元件清单不含被排除物料）。"""
    with open(SAMPLE_ELEC, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t08_")
    out = os.path.join(wd, "BOM_t08.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T08 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    rows = component_block(ws)
    assert rows is not None, "T08 应存在元件清单"
    names = [r[4] for r in rows]
    assert "裸PCB板" not in names, "T08 其他类(裸PCB板)应被排除"
    assert "贴片完成板" not in names, "T08 其他类(贴片完成板)应被排除"
    assert "元件清单已排除" in so, "T08 应打印排除提示WARNING"
    flat = flat_values(ws)
    assert "裸PCB板" in flat, "T08 被排除物料仍在物料区"


def test_v4_t09_sort():
    """排序正确: 物料类型升序→位号字母数字升序, 空位号排末尾。"""
    wd = tempfile.mkdtemp(prefix="v4t09_")
    out = os.path.join(wd, "BOM_t09.xlsx")
    data = {
        "product_name": "排序测试", "category": "工业品", "industry": "电子",
        "output_rate": 100,
        "materials": [
            {"name": "C1电容", "unit": "个", "usage": 1, "yield_rate": 100,
             "material_type": "电容", "process": "S01", "designator": "C1",
             "footprint": "0402", "part_number": "100n", "rohs": "是"},
            {"name": "C2电容", "unit": "个", "usage": 1, "yield_rate": 100,
             "material_type": "电容", "process": "S01", "designator": "C2",
             "footprint": "0402", "part_number": "200n", "rohs": "是"},
            {"name": "空号电容", "unit": "个", "usage": 1, "yield_rate": 100,
             "material_type": "电容", "process": "S01", "designator": "",
             "footprint": "", "part_number": "", "rohs": "是"},
            {"name": "R1电阻", "unit": "个", "usage": 1, "yield_rate": 100,
             "material_type": "电阻", "process": "S01", "designator": "R1",
             "footprint": "0402", "part_number": "10k", "rohs": "是"},
            {"name": "R10电阻", "unit": "个", "usage": 1, "yield_rate": 100,
             "material_type": "电阻", "process": "S01", "designator": "R10",
             "footprint": "0402", "part_number": "100k", "rohs": "是"},
            {"name": "R2电阻", "unit": "个", "usage": 1, "yield_rate": 100,
             "material_type": "电阻", "process": "S01", "designator": "R2",
             "footprint": "0402", "part_number": "20k", "rohs": "是"},
        ],
        "processes": [{"step_no": "S01", "name": "贴片", "output": "成品"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T09 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    rows = component_block(ws)
    names = [r[4] for r in rows]
    # 电容组(C1,C2,空) → 电阻组(R1,R10,R2)，空位号排末尾
    expected = ["C1电容", "C2电容", "空号电容", "R1电阻", "R10电阻", "R2电阻"]
    assert names == expected, "T09 排序不符: %s" % names


def test_v4_t10_rohs_color():
    """RoHS 着色: 否→红色, 未知/空→黄色(openpyxl 读 cell.font.color.rgb)。"""
    wd = tempfile.mkdtemp(prefix="v4t10_")
    out = os.path.join(wd, "BOM_t10.xlsx")
    data = {
        "product_name": "RoHS色测试", "category": "工业品", "industry": "电子",
        "output_rate": 100,
        "materials": [
            {"name": "红否", "unit": "个", "usage": 1, "yield_rate": 100,
             "material_type": "电阻", "process": "S01", "designator": "R1",
             "footprint": "0402", "part_number": "10k", "rohs": "否"},
            {"name": "黄未知", "unit": "个", "usage": 1, "yield_rate": 100,
             "material_type": "电容", "process": "S01", "designator": "C1",
             "footprint": "0402", "part_number": "100n", "rohs": "未知"},
            {"name": "黄空", "unit": "个", "usage": 1, "yield_rate": 100,
             "material_type": "IC", "process": "S01", "designator": "U1",
             "footprint": "QFP", "part_number": "x", "rohs": ""},
            {"name": "默认是", "unit": "个", "usage": 1, "yield_rate": 100,
             "material_type": "二极管", "process": "S01", "designator": "D1",
             "footprint": "SOD", "part_number": "y", "rohs": "是"},
        ],
        "processes": [{"step_no": "S01", "name": "p", "output": "OUT"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T10 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    m = marker_row(ws, "三、元件清单")
    assert m is not None, "T10 应含元件清单"
    # 建立 名称 -> 行号
    name2row = {}
    r = m + 2
    while r <= ws.max_row:
        nm = ws.cell(r, 5).value
        if not nm:
            break
        name2row[str(nm)] = r
        r += 1
    # 否 → 红 (FF0000)
    col = ws.cell(name2row["红否"], 8).font.color
    assert col is not None and col.rgb.endswith("FF0000"), \
        "T10 rohs=否 应红色(FF0000), 实际 %s" % (col.rgb if col else None)
    # 未知 → 黄 (BF8F00)
    col = ws.cell(name2row["黄未知"], 8).font.color
    assert col is not None and col.rgb.endswith("BF8F00"), \
        "T10 rohs=未知 应黄色(BF8F00), 实际 %s" % (col.rgb if col else None)
    # 空 → 黄
    col = ws.cell(name2row["黄空"], 8).font.color
    assert col is not None and col.rgb.endswith("BF8F00"), \
        "T10 rohs=空 应黄色(BF8F00), 实际 %s" % (col.rgb if col else None)
    # 是 → 默认(无特殊色)
    col = ws.cell(name2row["默认是"], 8).font.color
    assert col is None, \
        "T10 rohs=是 应默认(无特殊色), 实际 %s" % (col.rgb if col else None)


def test_v4_t11_material_area_no_special():
    """物料区8列不含 designator/footprint/part_number/rohs（专属字段不进物料区）。"""
    with open(SAMPLE_ELEC, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t11_")
    out = os.path.join(wd, "BOM_t11.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T11 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    hdr = [str(ws.cell(7, c).value or "").strip() for c in range(1, 9)]
    assert hdr == MATERIAL_HEADERS, "T11 物料区8列表头不符: %s" % hdr
    region = material_region_values(ws)
    joined = " ".join(region)
    assert "R1-R4" not in joined, "T11 物料区不应含 designator(如R1-R4)"
    assert "LQFP-48" not in joined, "T11 物料区不应含 footprint(LQFP-48)"
    assert "RoHS" not in joined, "T11 物料区表头不应含 RoHS"
    # 专属字段应出现在元件清单区
    flat = flat_values(ws)
    assert "LQFP-48" in flat, "T11 专属字段应出现在元件清单区"
    assert "R1-R4" in flat, "T11 designator应出现在元件清单区"


def test_v4_t12_chemical_header():
    """化工示例生成→含『三、配方表』区块, 8列表头正确。"""
    with open(SAMPLE_CHEM, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t12_")
    out = os.path.join(wd, "BOM_t12.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T12 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    m = marker_row(ws, "三、配方表")
    assert m is not None, "T12 应含『三、配方表』"
    hdr = [str(ws.cell(m + 1, c).value or "").strip() for c in range(1, 9)]
    expected = ["序号", "物料名称", "CAS号", "含量(%)", "GHS标识",
                "物料类型", "计量单位", "用量"]
    assert hdr == expected, "T12 配方表8列表头不符: %s" % hdr


def test_v4_t13_exclude_baocai():
    """排除 material_type=包材 的物料。"""
    with open(SAMPLE_CHEM, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t13_")
    out = os.path.join(wd, "BOM_t13.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T13 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    rows = formula_block(ws)
    assert rows is not None, "T13 应存在配方表"
    names = [r[1] for r in rows]
    assert "消毒液基料" not in names, "T13 包材(消毒液基料)应排除"
    assert "喷雾瓶" not in names, "T13 包材(喷雾瓶)应排除"
    assert "配方表已排除" in so, "T13 应打印排除提示WARNING"
    flat = flat_values(ws)
    assert "消毒液基料" in flat, "T13 被排除物料仍在物料区"


def test_v4_t14_formula_sort():
    """排序正确: 含量(%) 降序。"""
    with open(SAMPLE_CHEM, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t14_")
    out = os.path.join(wd, "BOM_t14.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T14 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    rows = formula_block(ws)
    names = [r[1] for r in rows]
    # 65,30,3,2 → 降序
    expected = ["去离子水", "乙醇", "甘油", "薄荷香精"]
    assert names == expected, "T14 配方排序(含量降序)不符: %s" % names


def test_v4_t15_conc_format():
    """含量(%) 数字格式 0.0"%" (openpyxl number_format)。"""
    with open(SAMPLE_CHEM, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t15_")
    out = os.path.join(wd, "BOM_t15.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T15 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    m = marker_row(ws, "三、配方表")
    assert m is not None, "T15 应含配方表"
    for r in range(m + 2, ws.max_row + 1):
        nm = ws.cell(r, 2).value
        if not nm:
            break
        fmt = ws.cell(r, 4).number_format
        assert fmt == '0.0"%"', \
            "T15 含量(%)应0.0%%格式, 实际 %s (行%s)" % (fmt, r)
        # 含量值应为数值(非字符串)
        assert isinstance(ws.cell(r, 4).value, (int, float)), \
            "T15 含量(%)应为数值, 实际 %r" % ws.cell(r, 4).value


def test_v4_t16_chem_material_area_no_special():
    """物料区8列不含 cas_number/concentration/ghs_hazard。"""
    with open(SAMPLE_CHEM, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t16_")
    out = os.path.join(wd, "BOM_t16.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T16 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    hdr = [str(ws.cell(7, c).value or "").strip() for c in range(1, 9)]
    assert hdr == MATERIAL_HEADERS, "T16 物料区8列表头不符: %s" % hdr
    region = material_region_values(ws)
    joined = " ".join(region)
    assert "7732-18-5" not in joined, "T16 物料区不应含 cas_number"
    assert "CAS号" not in joined, "T16 物料区表头不应含 CAS号"
    assert "GHS" not in joined, "T16 物料区不应含 GHS"
    flat = flat_values(ws)
    assert "7732-18-5" in flat, "T16 专属字段应出现在配方表区"


def test_v4_t17_w2_rohs_warning():
    """W2: 电子物料未标 rohs → WARNING 打印 + 退出码0(非阻断)。"""
    with open(SAMPLE_ELEC, encoding="utf-8") as f:
        data = json.load(f)
    # 样本内『晶振』rohs 为空 → 触发 W2
    wd = tempfile.mkdtemp(prefix="v4t17_")
    out = os.path.join(wd, "BOM_t17.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T17 W2不应阻断(rc=0), rc=%s so=%s se=%s" % (rc, so, se)
    assert "WARNING" in so, "T17 应打印WARNING, stdout=%s" % so
    assert "未标注 RoHS" in so, "T17 应含 RoHS 未标注 WARNING, stdout=%s" % so
    assert "OK:" in so, "T17 仍应输出 OK:(非阻断), 实际=%s" % so


def test_v4_t18_w3_cas_ghs_warning():
    """W3: 化工物料未填 cas_number/ghs_hazard → WARNING + 退出码0。"""
    wd = tempfile.mkdtemp(prefix="v4t18_")
    out = os.path.join(wd, "BOM_t18.xlsx")
    data = {
        "product_name": "W3测试", "category": "日化化妆品", "industry": "化工",
        "output_rate": 100,
        "materials": [
            {"name": "主料A", "unit": "kg", "usage": 70, "yield_rate": 100,
             "material_type": "主料", "process": "S01", "cas_number": "",
             "concentration": 70.0, "ghs_hazard": "易燃"},
            {"name": "溶剂B", "unit": "kg", "usage": 30, "yield_rate": 100,
             "material_type": "溶剂", "process": "S01", "cas_number": "123-45-6",
             "concentration": 30.0, "ghs_hazard": ""},
        ],
        "processes": [{"step_no": "S01", "name": "p", "output": "OUT"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T18 W3不应阻断(rc=0), rc=%s so=%s se=%s" % (rc, so, se)
    assert "WARNING" in so, "T18 应打印WARNING, stdout=%s" % so
    assert "未填写 CAS 号" in so, "T18 应含 CAS 未填 WARNING, stdout=%s" % so
    assert "未填写 GHS 危险标识" in so, "T18 应含 GHS 未填 WARNING, stdout=%s" % so
    assert "OK:" in so, "T18 仍应输出 OK:(非阻断), 实际=%s" % so


def test_v4_t19_conc_sum_deviation():
    """含量(%) 列和偏离100%超过±5% → WARNING + 退出码0。"""
    wd = tempfile.mkdtemp(prefix="v4t19_")
    out = os.path.join(wd, "BOM_t19.xlsx")
    data = {
        "product_name": "含量和偏离", "category": "日化化妆品", "industry": "化工",
        "output_rate": 100,
        "materials": [
            {"name": "主料A", "unit": "kg", "usage": 50, "yield_rate": 100,
             "material_type": "主料", "process": "S01", "cas_number": "1-1-1",
             "concentration": 50.0, "ghs_hazard": "x"},
            {"name": "溶剂B", "unit": "kg", "usage": 30, "yield_rate": 100,
             "material_type": "溶剂", "process": "S01", "cas_number": "2-2-2",
             "concentration": 30.0, "ghs_hazard": "y"},
        ],
        "processes": [{"step_no": "S01", "name": "p", "output": "OUT"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T19 含量和校验不应阻断(rc=0), rc=%s so=%s" % (rc, so)
    assert "WARNING" in so, "T19 应打印WARNING, stdout=%s" % so
    assert "偏离 100%" in so, "T19 应含含量列和偏离 WARNING, stdout=%s" % so
    assert "OK:" in so, "T19 仍应输出 OK:, 实际=%s" % so


def test_v4_t20_conc_sum_100_no_warning():
    """含量(%) 列和=100% → 无 WARNING。"""
    wd = tempfile.mkdtemp(prefix="v4t20_")
    out = os.path.join(wd, "BOM_t20.xlsx")
    data = {
        "product_name": "含量和达标", "category": "日化化妆品", "industry": "化工",
        "output_rate": 100,
        "materials": [
            {"name": "主料A", "unit": "kg", "usage": 60, "yield_rate": 100,
             "material_type": "主料", "process": "S01", "cas_number": "1-1-1",
             "concentration": 60.0, "ghs_hazard": "x"},
            {"name": "溶剂B", "unit": "kg", "usage": 40, "yield_rate": 100,
             "material_type": "溶剂", "process": "S01", "cas_number": "2-2-2",
             "concentration": 40.0, "ghs_hazard": "y"},
        ],
        "processes": [{"step_no": "S01", "name": "p", "output": "OUT"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T20 生成应成功(rc=0), rc=%s so=%s se=%s" % (rc, so, se)
    assert "WARNING" not in so, "T20 含量和=100%应无WARNING, stdout=%s" % so
    assert "OK:" in so, "T20 应输出 OK:, 实际=%s" % so


def test_v4_t21_elec_reverse():
    """电子 xlsx 逆向→JSON 含 industry=电子, designator/footprint/part_number/rohs 正确回收。"""
    with open(SAMPLE_ELEC, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t21_")
    out = os.path.join(wd, "BOM_t21.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T21 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T21 逆向应成功, rc=%s so2=%s se2=%s" % (rc2, so2, se2)
    assert parsed.get("industry") == "电子", \
        "T21 industry应=电子, 实际=%s" % parsed.get("industry")
    mats = {m["name"]: m for m in parsed["materials"]}
    mc = mats.get("主控芯片")
    assert mc is not None, "T21 应含主控芯片"
    assert mc.get("designator") == "U1", "T21 designator 应=U1, 实际=%s" % mc.get("designator")
    assert mc.get("footprint") == "LQFP-48", "T21 footprint 应=LQFP-48, 实际=%s" % mc.get("footprint")
    assert mc.get("part_number") == "STM32F103C8T6", "T21 part_number 不符"
    assert mc.get("rohs") == "是", "T21 rohs 应=是, 实际=%s" % mc.get("rohs")
    # 空 rohs 回收为 ""
    jx = mats.get("晶振")
    assert jx.get("rohs") == "", "T21 晶振 rohs 应回收空串, 实际=%r" % jx.get("rohs")
    # 7 个专属字段均存在且默认空串(对未回收者)
    for m in parsed["materials"]:
        for f in SPECIAL_FIELDS:
            assert f in m, "T21 物料 %s 应含专属字段 %s" % (m.get("name"), f)


def test_v4_t22_chem_reverse():
    """化工 xlsx 逆向→JSON 含 industry=化工, cas_number/concentration/ghs_hazard 正确回收。"""
    with open(SAMPLE_CHEM, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t22_")
    out = os.path.join(wd, "BOM_t22.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T22 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T22 逆向应成功, rc=%s so2=%s se2=%s" % (rc2, so2, se2)
    assert parsed.get("industry") == "化工", \
        "T22 industry应=化工, 实际=%s" % parsed.get("industry")
    mats = {m["name"]: m for m in parsed["materials"]}
    water = mats.get("去离子水")
    assert water is not None, "T22 应含去离子水"
    assert water.get("cas_number") == "7732-18-5", "T22 cas_number 不符: %s" % water.get("cas_number")
    assert water.get("concentration") == 65.0, "T22 concentration 应=65.0, 实际=%r" % water.get("concentration")
    assert water.get("ghs_hazard") == "无", "T22 ghs_hazard 应=无, 实际=%s" % water.get("ghs_hazard")
    # 包材物料专属字段为默认空串
    pk = mats.get("喷雾瓶")
    assert pk.get("cas_number") == "", "T22 包材 cas 应空"
    assert pk.get("concentration") == "", "T22 包材 concentration 应空"


def test_v4_t23_v3_reverse_food():
    """旧版 V3 xlsx 逆向→industry 推断为食品, 7 个专属字段默认空串, 不报错。"""
    assert os.path.exists(SAMPLE_V3_XLSX), "T23 旧版示例 sample_bom_v3.xlsx 缺失"
    wd = tempfile.mkdtemp(prefix="v4t23_")
    back = os.path.join(wd, "back.json")
    rc, so, se, parsed = run_import(SAMPLE_V3_XLSX, back)
    assert rc == 0, "T23 旧版V3 xlsx 逆向应成功(rc=0), rc=%s so=%s se=%s" % (rc, so, se)
    assert parsed.get("industry") == "食品", \
        "T23 industry应推断为食品, 实际=%s" % parsed.get("industry")
    for m in parsed["materials"]:
        for f in SPECIAL_FIELDS:
            assert f in m and m[f] == "", \
                "T23 物料 %s 专属字段 %s 应默认空串, 实际=%r" % (m.get("name"), f, m.get(f))


def test_v4_t24_reverse_no_derived_keys():
    """逆向 JSON 不含『序号』『用量占比%』键(派生展示列不回写)。"""
    with open(SAMPLE_ELEC, encoding="utf-8") as f:
        data = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t24_")
    out = os.path.join(wd, "BOM_t24.xlsx")
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T24 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T24 逆向应成功, rc=%s so2=%s se2=%s" % (rc2, so2, se2)
    assert "序号" not in parsed, "T24 顶层不应含『序号』"
    assert "用量占比%" not in parsed, "T24 顶层不应含『用量占比%』"
    for m in parsed["materials"]:
        assert "序号" not in m, "T24 物料不应含『序号』"
        assert "用量占比%" not in m, "T24 物料不应含『用量占比%』"
    for p in parsed["processes"]:
        assert "序号" not in p, "T24 工序不应含『序号』"


def test_v4_t25_empty_product_name():
    """product_name 空 → 退出码2 (V1 校验)。"""
    wd = tempfile.mkdtemp(prefix="v4t25_")
    out = os.path.join(wd, "BOM_t25.xlsx")
    rc, so, se = run_generate({
        "product_name": "", "category": "食品", "output_rate": 130,
        "materials": [{"name": "A", "unit": "kg", "usage": 10,
                       "yield_rate": 100, "material_type": "原料"}],
        "processes": [],
    }, out)
    assert rc == 2, "T25 产品名空应退出2, 实际 rc=%s" % rc
    assert "VALIDATION_FAILED" in so, "T25 应含 VALIDATION_FAILED, 实际=%s" % so


def test_v4_t26_output_rate_zero():
    """output_rate<=0 → 退出码2 (V3 校验)。"""
    wd = tempfile.mkdtemp(prefix="v4t26_")
    out = os.path.join(wd, "BOM_t26.xlsx")
    rc, so, se = run_generate({
        "product_name": "X", "category": "食品", "output_rate": 0,
        "materials": [{"name": "A", "unit": "kg", "usage": 10,
                       "yield_rate": 100, "material_type": "原料"}],
        "processes": [],
    }, out)
    assert rc == 2, "T26 output_rate=0 应退出2, 实际 rc=%s" % rc
    assert "VALIDATION_FAILED" in so, "T26 应含 VALIDATION_FAILED, 实际=%s" % so


def test_v4_t27_chain_broken():
    """工序流转链断裂(R3) → 退出码2。"""
    wd = tempfile.mkdtemp(prefix="v4t27_")
    out = os.path.join(wd, "BOM_t27.xlsx")
    rc, so, se = run_generate({
        "product_name": "X", "category": "食品", "output_rate": 130,
        "materials": [
            {"name": "A", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S01"},
            {"name": "B", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S02"},
        ],
        "processes": [
            {"step_no": "S01", "name": "p1", "output": "OUT1"},
            {"step_no": "S02", "name": "p2", "output": "OUT2"},
        ],
    }, out)
    assert rc == 2, "T27 R3断链应退出2, 实际 rc=%s" % rc
    assert "流转链不完整" in so, "T27 应提示『流转链不完整』, 实际=%s" % so


def test_v4_t28_elec_closed_loop():
    """电子: generate→import→关键字段(名称+designator+rohs+processes)一致。"""
    with open(SAMPLE_ELEC, encoding="utf-8") as f:
        orig = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t28_")
    out = os.path.join(wd, "BOM_t28.xlsx")
    rc, so, se = run_generate(orig, out)
    assert rc == 0, "T28 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T28 逆向应成功, rc=%s so2=%s se2=%s" % (rc2, so2, se2)
    assert parsed.get("product_name") == orig["product_name"], "T28 product_name 不一致"
    assert parsed.get("industry") == "电子", "T28 industry 不一致"
    assert parsed.get("category") == orig["category"], "T28 category 不一致"
    assert parsed.get("output_rate") == float(orig["output_rate"]), "T28 output_rate 不一致"
    om = {m["name"]: m for m in orig["materials"]}
    dm = {m["name"]: m for m in parsed["materials"]}
    assert set(dm.keys()) == set(om.keys()), \
        "T28 物料集合不一致: orig=%s back=%s" % (set(om.keys()), set(dm.keys()))
    for name, om_ in om.items():
        dm_ = dm[name]
        assert float(dm_.get("usage", 0)) == float(om_["usage"]), \
            "T28 物料 %s usage 不一致" % name
        assert dm_.get("material_type") == om_.get("material_type", ""), \
            "T28 物料 %s material_type 不一致" % name
        assert dm_.get("process") == om_.get("process", ""), \
            "T28 物料 %s process 不一致" % name
        # 元件清单仅含非其他类; 其他类专属字段不可回放, 跳过
        if om_.get("material_type") != "其他":
            assert dm_.get("designator") == om_.get("designator", ""), \
                "T28 物料 %s designator 不一致" % name
            assert dm_.get("rohs") == om_.get("rohs", ""), \
                "T28 物料 %s rohs 不一致" % name
            assert dm_.get("part_number") == om_.get("part_number", ""), \
                "T28 物料 %s part_number 不一致" % name
            assert dm_.get("footprint") == om_.get("footprint", ""), \
                "T28 物料 %s footprint 不一致" % name
    op = {p["step_no"]: p for p in orig["processes"]}
    dp = {p["step_no"]: p for p in parsed["processes"]}
    assert set(dp.keys()) == set(op.keys()), "T28 工序集合不一致"
    for sn, op_ in op.items():
        assert dp[sn].get("output") == op_.get("output", ""), \
            "T28 工序 %s output 不一致" % sn
    assert "序号" not in parsed, "T28 JSON 不应含『序号』"


def test_v4_t29_chem_closed_loop():
    """化工: generate→import→关键字段(名称+cas+conc+ghs+processes)一致。"""
    with open(SAMPLE_CHEM, encoding="utf-8") as f:
        orig = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t29_")
    out = os.path.join(wd, "BOM_t29.xlsx")
    rc, so, se = run_generate(orig, out)
    assert rc == 0, "T29 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, parsed = run_import(out, back)
    assert rc2 == 0, "T29 逆向应成功, rc=%s so2=%s se2=%s" % (rc2, so2, se2)
    assert parsed.get("product_name") == orig["product_name"], "T29 product_name 不一致"
    assert parsed.get("industry") == "化工", "T29 industry 不一致"
    assert parsed.get("category") == orig["category"], "T29 category 不一致"
    assert parsed.get("output_rate") == float(orig["output_rate"]), "T29 output_rate 不一致"
    om = {m["name"]: m for m in orig["materials"]}
    dm = {m["name"]: m for m in parsed["materials"]}
    assert set(dm.keys()) == set(om.keys()), \
        "T29 物料集合不一致: orig=%s back=%s" % (set(om.keys()), set(dm.keys()))
    for name, om_ in om.items():
        dm_ = dm[name]
        assert float(dm_.get("usage", 0)) == float(om_["usage"]), \
            "T29 物料 %s usage 不一致" % name
        assert dm_.get("material_type") == om_.get("material_type", ""), \
            "T29 物料 %s material_type 不一致" % name
        assert dm_.get("process") == om_.get("process", ""), \
            "T29 物料 %s process 不一致" % name
        # 配方表仅含非包材类; 包材类专属字段不可回放, 跳过
        if om_.get("material_type") != "包材":
            assert dm_.get("cas_number") == om_.get("cas_number", ""), \
                "T29 物料 %s cas_number 不一致" % name
            assert float(dm_.get("concentration", 0) or 0) == float(om_.get("concentration", 0) or 0), \
                "T29 物料 %s concentration 不一致" % name
            assert dm_.get("ghs_hazard") == om_.get("ghs_hazard", ""), \
                "T29 物料 %s ghs_hazard 不一致" % name
    op = {p["step_no"]: p for p in orig["processes"]}
    dp = {p["step_no"]: p for p in parsed["processes"]}
    assert set(dp.keys()) == set(op.keys()), "T29 工序集合不一致"
    for sn, op_ in op.items():
        assert dp[sn].get("output") == op_.get("output", ""), \
            "T29 工序 %s output 不一致" % sn
    assert "序号" not in parsed, "T29 JSON 不应含『序号』"


def test_v4_t30_output_rate_format():
    """食品/电子/化工示例出品率显示均为带1位小数格式(130.0% / 100.0%)。"""
    # 食品示例 (output_rate=130)
    with open(SAMPLE_V3, encoding="utf-8") as f:
        v3 = json.load(f)
    wd = tempfile.mkdtemp(prefix="v4t30_")
    out1 = os.path.join(wd, "BOM_t30_food.xlsx")
    rc, so, se = run_generate(v3, out1)
    assert rc == 0, "T30 食品生成应成功, rc=%s se=%s" % (rc, se)
    ws1 = load(out1)
    r4_1 = row4_text(ws1)
    assert "全产品出品率：130.0%" in r4_1, \
        "T30 食品应含『全产品出品率：130.0%』, 实际=%s" % r4_1
    assert "全产品出品率：130%" not in r4_1, \
        "T30 食品不应为裸『130%』(应带1位小数)"
    # 电子示例 (output_rate=100)
    with open(SAMPLE_ELEC, encoding="utf-8") as f:
        elec = json.load(f)
    out2 = os.path.join(wd, "BOM_t30_elec.xlsx")
    rc, so, se = run_generate(elec, out2)
    assert rc == 0, "T30 电子生成应成功, rc=%s se=%s" % (rc, se)
    ws2 = load(out2)
    r4_2 = row4_text(ws2)
    assert "全产品出品率：100.0%" in r4_2, \
        "T30 电子应含『全产品出品率：100.0%』, 实际=%s" % r4_2
    # 化工示例 (output_rate=100)
    with open(SAMPLE_CHEM, encoding="utf-8") as f:
        chem = json.load(f)
    out3 = os.path.join(wd, "BOM_t30_chem.xlsx")
    rc, so, se = run_generate(chem, out3)
    assert rc == 0, "T30 化工生成应成功, rc=%s se=%s" % (rc, se)
    ws3 = load(out3)
    r4_3 = row4_text(ws3)
    assert "全产品出品率：100.0%" in r4_3, \
        "T30 化工应含『全产品出品率：100.0%』, 实际=%s" % r4_3


# ----------------------------------------------------------------------------
# 运行器
# ----------------------------------------------------------------------------
TESTS = [
    ("V4-T01", "食品无industry→推断食品→配料表", test_v4_t01_food_infer_ingredient),
    ("V4-T02", "显式industry=电子→元件清单", test_v4_t02_explicit_electronic),
    ("V4-T03", "显式industry=化工→配方表", test_v4_t03_explicit_chemical),
    ("V4-T04", "显式industry=通用→无专属视图", test_v4_t04_generic_no_view),
    ("V4-T05", "非食品无industry→推断通用→无视图", test_v4_t05_nonfood_infer_generic),
    ("V4-T06", "非法industry→V8 WARNING+回退+rc0", test_v4_t06_illegal_industry_warning),
    ("V4-T07", "电子示例→元件清单8列表头", test_v4_t07_electronic_header),
    ("V4-T08", "排除其他类物料", test_v4_t08_exclude_other),
    ("V4-T09", "元件排序(类型→位号,空末)", test_v4_t09_sort),
    ("V4-T10", "RoHS着色(否红/未知空黄)", test_v4_t10_rohs_color),
    ("V4-T11", "物料区8列不含电子专属字段", test_v4_t11_material_area_no_special),
    ("V4-T12", "化工示例→配方表8列表头", test_v4_t12_chemical_header),
    ("V4-T13", "排除包材类物料", test_v4_t13_exclude_baocai),
    ("V4-T14", "配方排序(含量降序)", test_v4_t14_formula_sort),
    ("V4-T15", "含量(%)数字格式0.0%", test_v4_t15_conc_format),
    ("V4-T16", "物料区8列不含化工专属字段", test_v4_t16_chem_material_area_no_special),
    ("V4-T17", "W2:电子未标rohs→WARNING", test_v4_t17_w2_rohs_warning),
    ("V4-T18", "W3:化工未填cas/ghs→WARNING", test_v4_t18_w3_cas_ghs_warning),
    ("V4-T19", "含量和偏离→WARNING", test_v4_t19_conc_sum_deviation),
    ("V4-T20", "含量和=100%→无WARNING", test_v4_t20_conc_sum_100_no_warning),
    ("V4-T21", "电子xlsx逆向→industry+4字段", test_v4_t21_elec_reverse),
    ("V4-T22", "化工xlsx逆向→industry+3字段", test_v4_t22_chem_reverse),
    ("V4-T23", "旧V3 xlsx逆向→食品+7字段空", test_v4_t23_v3_reverse_food),
    ("V4-T24", "逆向JSON不含序号/用量占比%", test_v4_t24_reverse_no_derived_keys),
    ("V4-T25", "product_name空→rc2", test_v4_t25_empty_product_name),
    ("V4-T26", "output_rate<=0→rc2", test_v4_t26_output_rate_zero),
    ("V4-T27", "工序链断裂→rc2", test_v4_t27_chain_broken),
    ("V4-T28", "电子闭环一致性", test_v4_t28_elec_closed_loop),
    ("V4-T29", "化工闭环一致性", test_v4_t29_chem_closed_loop),
    ("V4-T30", "出品率1位小数格式", test_v4_t30_output_rate_format),
]


def main():
    print("==================================================")
    print(" BOM智造师 V4 (多行业扩展) 独立测试 (QA 严过关)")
    print("==================================================")
    results = []
    for tid, name, fn in TESTS:
        try:
            fn()
            results.append((tid, name, True, "PASS"))
            print("%-8s %-32s PASS" % (tid, name))
        except AssertionError as e:
            results.append((tid, name, False, str(e)))
            print("%-8s %-32s FAIL  %s" % (tid, name, str(e)))
        except Exception as e:  # noqa: BLE001
            results.append((tid, name, False, "EXCEPTION: %s" % e))
            print("%-8s %-32s FAIL  EXCEPTION: %s" % (tid, name, e))

    total = len(results)
    passed = sum(1 for r in results if r[2])
    failed = total - passed
    print("--------------------------------------------------")
    print("总计: %d   通过: %d   失败: %d" % (total, passed, failed))
    print("IS_PASS: %s" % ("YES" if failed == 0 else "NO"))
    if failed:
        print("失败用例:")
        for tid, name, ok, detail in results:
            if not ok:
                print("  [%s] %s -> %s" % (tid, name, detail))
    print("--------------------------------------------------")
    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
