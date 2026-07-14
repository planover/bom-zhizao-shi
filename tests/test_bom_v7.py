#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BOM 智造师 V7 增量 · 双语 / 空白模板 / 批量生成 / 逆向合并 独立验证套件
=========================================================================

QA 工程师 Edward 独立编写，不依赖实现者自测结论，且**不修改任何 V2–V6 既有断言**。

覆盖 V7 增量核心能力（设计基线 incremental-design-v7.md）：
  P0-A  双语 BOM（--bilingual）：主表「BOM表」与 V6 逐字节一致；
        追加「BOM表(英)」双语 sheet（中英双行表头、编码不翻译、物料类型英文）。
  P0-B1 空白模板批量（--blank-templates）：生成 template_<行业>.xlsx。
  P0-B2 批量生成（--batch-dir / --batch）：错误隔离、失败退出码 2、命名规则。
  P0-B3 逆向合并（import_bom.py --in nargs="+" --merge）：顺序 extend、
        step_no 冲突留痕、industry 首位优先、merged_from/merge_notes、单文件失败跳过。
  边界：空数据 / 超长字段 / 非法 industry 回落 / 缺必填 V1–V3 / 异常 JSON /
        1000+ 物料性能基线。
  约束：import_bom.py 不引用 I18N/ZH2EN；industry 枚举不变；不新增依赖。

测试方式：subprocess 调用 CLI 做端到端闭环 + openpyxl 直接读取做结构断言；
          另对纯函数 / 常量做单元测试。

被测脚本目录: C:/Users/姓名/.workbuddy/skills/bom-zhizao-shi/scripts
运行:
    C:\\Users\\姓名\\.workbuddy\\binaries\\python\\envs\\default\\Scripts\\python.exe -m pytest tests/test_bom_v7.py -v
"""

import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import unittest

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# 路径常量（与既有测试保持一致）
# ---------------------------------------------------------------------------
SKILL_DIR = r"C:\Users\姓名\.workbuddy\skills\bom-zhizao-shi"
SKILL_SCRIPTS = os.path.join(SKILL_DIR, "scripts")
GEN = os.path.join(SKILL_SCRIPTS, "generate_bom.py")
IMP = os.path.join(SKILL_SCRIPTS, "import_bom.py")
PY = r"C:\Users\姓名\.workbuddy\binaries\python\envs\default\Scripts\python.exe"
ALL_INDUSTRIES = ["食品", "电子", "化工", "机械", "纺织", "家具", "包装", "通用"]

if SKILL_SCRIPTS not in sys.path:
    sys.path.insert(0, SKILL_SCRIPTS)

from bom_constants import I18N, ZH2EN, INDUSTRIES  # noqa: E402
from generate_bom import ALL_INDUSTRIES  # noqa: E402

TMP = os.path.join(tempfile.gettempdir(), "bom_v7_test")
os.makedirs(TMP, exist_ok=True)


def run_cli(script_path, *cli_args):
    """调用一个 CLI，返回 CompletedProcess。"""
    cmd = [PY, script_path] + list(cli_args)
    return subprocess.run(cmd, capture_output=True, text=True, cwd=TMP)


def find_marker_row(ws, substring):
    """定位首列包含 substring 的行号（1-based），未找到返回 None。"""
    for idx, row in enumerate(ws.iter_rows(min_col=1, max_col=1), 1):
        v = row[0].value
        if v is not None and substring in str(v):
            return idx
    return None


def _cell_sig(cell):
    """提取单元格的可比签名（值/数字格式/字体/对齐/填充）。"""
    f = cell.font
    a = cell.alignment
    fill = cell.fill
    fg = None
    if fill is not None and getattr(fill, "patternType", None) is not None:
        try:
            fg = fill.fgColor.rgb
        except Exception:
            fg = None
    color = None
    if getattr(f, "color", None) is not None:
        try:
            color = f.color.rgb
        except Exception:
            color = None
    return {
        "v": cell.value,
        "nf": cell.number_format,
        "fname": getattr(f, "name", None),
        "fsz": getattr(f, "sz", None),
        "fb": getattr(f, "b", None),
        "fcolor": color,
        "h": getattr(a, "horizontal", None),
        "vrt": getattr(a, "vertical", None),
        "wrap": getattr(a, "wrap_text", None),
        "fg": fg,
    }


def assert_sheets_byte_equal(self, ws1, ws2, label=""):
    """断言两个工作表逐字节一致（值/字体/对齐/合并/列宽/行高）。"""
    self.assertEqual(ws1.max_row, ws2.max_row, "%s: max_row 不同" % label)
    self.assertEqual(ws1.max_column, ws2.max_column, "%s: max_column 不同" % label)
    for r in range(1, ws1.max_row + 1):
        for c in range(1, ws1.max_column + 1):
            s1 = _cell_sig(ws1.cell(r, c))
            s2 = _cell_sig(ws2.cell(r, c))
            self.assertEqual(s1, s2, "%s: 单元 (%d,%d) 不一致: %r vs %r" % (label, r, c, s1, s2))
    self.assertEqual(
        sorted(str(m) for m in ws1.merged_cells.ranges),
        sorted(str(m) for m in ws2.merged_cells.ranges),
        "%s: 合并区域不同" % label,
    )
    for col, dim in ws1.column_dimensions.items():
        self.assertEqual(dim.width, ws2.column_dimensions[col].width,
                         "%s: 列 %s 宽度不同" % (label, col))
    for r, dim in ws1.row_dimensions.items():
        self.assertEqual(dim.height, ws2.row_dimensions[r].height,
                         "%s: 行 %d 高度不同" % (label, r))


# ---------------------------------------------------------------------------
# 样本数据
# ---------------------------------------------------------------------------
def _sample_food(product_name="芒果果味糖浆", date="2026-07-07"):
    return {
        "product_name": product_name,
        "category": "食品",
        "industry": "食品",
        "output_rate": 130,
        "version": "V1.0",
        "date": date,
        "approver": "张三",
        "effective_date": "2026-07-10",
        "standard": "GB 7718-2025",
        "materials": [
            {"name": "芒果原浆", "unit": "kg", "usage": 46.3, "yield_rate": 55,
             "erp_code": "RM-001", "material_type": "原料", "process": "S01",
             "allergen": "大豆,乳"},
            {"name": "白砂糖", "unit": "kg", "usage": 30.0, "yield_rate": 100,
             "erp_code": "RM-002", "material_type": "原料", "process": "S01"},
            {"name": "芒果果味糖浆基料", "unit": "kg", "usage": 70.0, "yield_rate": 98,
             "erp_code": "RM-100", "material_type": "原料", "process": "S02"},
            {"name": "PE 瓶", "unit": "个", "usage": 100, "yield_rate": 100,
             "erp_code": "PK-001", "material_type": "包材", "process": ""},
        ],
        "processes": [
            {"step_no": "S01", "name": "调配", "desc": "混合搅拌", "work_hours": 30,
             "note": "常温", "output": "芒果果味糖浆基料"},
            {"step_no": "S02", "name": "灌装", "desc": "无菌灌装", "work_hours": 20,
             "note": "", "output": "芒果果味糖浆"},
        ],
    }


def _sample_mech(product_name="支架组件", date="2026-07-08"):
    return {
        "product_name": product_name,
        "category": "工业品",
        "industry": "机械",
        "output_rate": 95,
        "version": "V1.0",
        "date": date,
        "materials": [
            {"name": "钢板", "unit": "kg", "usage": 1.0, "yield_rate": 100,
             "erp_code": "MT-001", "material_type": "主材", "process": "S01",
             "drawing_no": "DW-1", "material": "Q235", "weight": 1.0, "unit_weight": 1.0},
            {"name": "钢板件", "unit": "kg", "usage": 1.0, "yield_rate": 100,
             "erp_code": "MT-100", "material_type": "主材", "process": "S02"},
            {"name": "螺栓", "unit": "个", "usage": 10, "yield_rate": 100,
             "erp_code": "MT-002", "material_type": "辅材", "process": "S02"},
        ],
        "processes": [
            {"step_no": "S01", "name": "切割", "desc": "锯切", "work_hours": 5,
             "note": "", "output": "钢板件"},
            {"step_no": "S02", "name": "焊接", "desc": "点焊", "work_hours": 8,
             "note": "", "output": "支架组件"},
        ],
    }


# ===========================================================================
# P0-A 双语 BOM
# ===========================================================================
class TestBilingual(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp(prefix="bom_v7_bi_", dir=TMP)

    def _gen(self, data, bilingual):
        p = os.path.join(self.d, "bom.json")
        with open(p, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        out = os.path.join(self.d, "out_bi.xlsx" if bilingual else "out_zh.xlsx")
        rc = run_cli(GEN, "--data", p, "--out", out,
                     *((("--bilingual",) if bilingual else ())))
        if rc.returncode != 0:
            self.fail("generate rc=%d\nSTDOUT=%r\nSTDERR=%r"
                      % (rc.returncode, rc.stdout, rc.stderr))
        return out

    def test_zh_sheet_byte_identical_to_non_bilingual(self):
        """双语模式下主表「BOM表」必须与单语模式逐字节一致。"""
        zh = self._gen(_sample_food(), bilingual=False)
        bi = self._gen(_sample_food(), bilingual=True)
        wb_zh = load_workbook(zh)
        wb_bi = load_workbook(bi)
        self.assertIn("BOM表", wb_zh.sheetnames)
        self.assertIn("BOM表", wb_bi.sheetnames)
        assert_sheets_byte_equal(self, wb_zh["BOM表"], wb_bi["BOM表"], "BOM表")

    def test_bilingual_appends_english_sheet(self):
        bi = self._gen(_sample_food(), bilingual=True)
        wb = load_workbook(bi)
        self.assertEqual(wb.sheetnames, ["BOM表", "BOM表(英)"])
        ws = wb["BOM表(英)"]
        self.assertEqual(ws["A1"].value, "BOM表 (BOM Table)")

    def test_bilingual_block_title_and_dual_header(self):
        bi = self._gen(_sample_food(), bilingual=True)
        ws = load_workbook(bi)["BOM表(英)"]
        mrow = find_marker_row(ws, "一、物料信息")
        self.assertIsNotNone(mrow)
        # 区块标题合并行：中文 (English)
        self.assertEqual(ws.cell(mrow, 1).value, "一、物料信息 (I. Material Information)")
        # 下行中文表头，再下行英文表头
        zh_header = [ws.cell(mrow + 1, c).value for c in range(1, 9)]
        en_header = [ws.cell(mrow + 2, c).value for c in range(1, 9)]
        self.assertIn("物料名称", zh_header)
        self.assertIn("Material Name", en_header)
        self.assertIn("物料类型", zh_header)
        self.assertIn("Material Type", en_header)

    def test_bilingual_material_type_translated_not_code(self):
        """英文表物料类型按 ZH2EN 显示英文；standard 编码不翻译。"""
        bi = self._gen(_sample_food(), bilingual=True)
        ws = load_workbook(bi)["BOM表(英)"]
        mrow = find_marker_row(ws, "一、物料信息")
        # 定位 Material Type 列（英文表头行 mrow+2）
        en_header = [ws.cell(mrow + 2, c).value for c in range(1, 9)]
        mt_col = en_header.index("Material Type") + 1
        # 首个数据行：跳过「中文/英文表头」与「【工序 …】」分组副标题，
        # 第一个物料名（芒果原浆）所在行即为数据行。
        first_type = None
        for r in range(mrow + 3, ws.max_row + 1):
            if ws.cell(r, 2).value == "芒果原浆":
                first_type = ws.cell(r, mt_col).value
                break
        self.assertEqual(first_type, ZH2EN.get("原料"), "物料类型应翻译为英文")
        # standard 编码不翻译：英文表执行标准行标签译为 Executive Standard，
        # 但代号原样保留（见 test_bilingual_standard_code_not_translated）。
        self.assertIsNotNone(ZH2EN.get("执行标准"))

    def test_bilingual_standard_code_not_translated(self):
        """编码类字段（standard 代号）在双语表保留原值（拼接于行 5 元信息单元格）。"""
        bi = self._gen(_sample_food(), bilingual=True)
        wb = load_workbook(bi)
        ws = wb["BOM表(英)"]
        # 行 5 元信息单元格形如：
        #   'Approver: 张三    Effective Date: 2026-07-10    Executive Standard: GB 7718-2025'
        # 代号 GB 7718-2025 作为子串保留（标签已译为英文，编码不翻译）。
        found = False
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "GB 7718-2025" in v:
                    found = True
        self.assertTrue(found, "standard 代号应在双语表保留（子串匹配）")


# ===========================================================================
# P0-B1 空白模板批量
# ===========================================================================
class TestBlankTemplates(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp(prefix="bom_v7_b1_", dir=TMP)

    def test_default_8_templates(self):
        rc = run_cli(GEN, "--blank-templates", "--out-dir", self.d)
        self.assertEqual(rc.returncode, 0, rc.stderr)
        files = sorted(os.listdir(self.d))
        expected = sorted("template_%s.xlsx" % i for i in ALL_INDUSTRIES)
        self.assertEqual(files, expected)
        # 每个模板可打开且含「BOM表」与「一、物料信息」标记
        for fn in files:
            wb = load_workbook(os.path.join(self.d, fn))
            self.assertIn("BOM表", wb.sheetnames)
            self.assertIsNotNone(find_marker_row(wb["BOM表"], "一、物料信息"))

    def test_subset_industries(self):
        rc = run_cli(GEN, "--blank-templates", "--out-dir", self.d,
                     "--industries", "食品,电子")
        self.assertEqual(rc.returncode, 0, rc.stderr)
        files = sorted(os.listdir(self.d))
        self.assertEqual(files, ["template_电子.xlsx", "template_食品.xlsx"])

    def test_blank_templates_bilingual(self):
        rc = run_cli(GEN, "--blank-templates", "--out-dir", self.d,
                     "--industries", "食品", "--bilingual")
        self.assertEqual(rc.returncode, 0, rc.stderr)
        wb = load_workbook(os.path.join(self.d, "template_食品.xlsx"))
        self.assertEqual(wb.sheetnames, ["BOM表", "BOM表(英)"])


# ===========================================================================
# P0-B2 批量生成
# ===========================================================================
class TestBatchGenerate(unittest.TestCase):
    def setUp(self):
        self.indir = tempfile.mkdtemp(prefix="bom_v7_b2in_", dir=TMP)
        self.outdir = tempfile.mkdtemp(prefix="bom_v7_b2out_", dir=TMP)

    def _write(self, name, data):
        p = os.path.join(self.indir, name)
        with open(p, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        return p

    def test_all_success_naming(self):
        self._write("a.json", _sample_food("芒果果味糖浆", "2026-07-07"))
        self._write("b.json", _sample_mech("支架组件", "2026-07-08"))
        rc = run_cli(GEN, "--batch-dir", self.indir, "--out-dir", self.outdir)
        self.assertEqual(rc.returncode, 0, rc.stderr)
        outs = sorted(os.listdir(self.outdir))
        self.assertIn("BOM_芒果果味糖浆_2026-07-07.xlsx", outs)
        self.assertIn("BOM_支架组件_2026-07-08.xlsx", outs)
        self.assertIn("成功 2 / 失败 0", rc.stdout)

    def test_error_isolation_exit_code_2(self):
        """单文件异常不中断其余文件；任一失败整体退出码 2。"""
        self._write("ok.json", _sample_food("芒果果味糖浆", "2026-07-07"))
        # 非法 JSON
        with open(os.path.join(self.indir, "bad.json"), "w", encoding="utf-8") as f:
            f.write("{ this is not json ")
        # 缺必填（无 product_name）
        self._write("missing.json", {"category": "食品", "output_rate": 100})
        rc = run_cli(GEN, "--batch-dir", self.indir, "--out-dir", self.outdir)
        self.assertEqual(rc.returncode, 2, "有失败应退出码 2")
        self.assertIn("成功 1 / 失败 2", rc.stdout)
        # 合法文件仍生成
        self.assertTrue(os.path.exists(
            os.path.join(self.outdir, "BOM_芒果果味糖浆_2026-07-07.xlsx")))

    def test_explicit_batch_list(self):
        pa = self._write("a.json", _sample_food("芒果果味糖浆", "2026-07-07"))
        pb = self._write("b.json", _sample_mech("支架组件", "2026-07-08"))
        rc = run_cli(GEN, "--batch", "%s,%s" % (pa, pb), "--out-dir", self.outdir)
        self.assertEqual(rc.returncode, 0, rc.stderr)
        self.assertEqual(len(os.listdir(self.outdir)), 2)

    def test_illegal_filename_chars_replaced(self):
        """产品名含非法文件名字符 → 替换为 _。"""
        self._write("x.json", _sample_food('A/B:C*?', "2026-07-07"))
        rc = run_cli(GEN, "--batch-dir", self.indir, "--out-dir", self.outdir)
        self.assertEqual(rc.returncode, 0, rc.stderr)
        outs = os.listdir(self.outdir)
        self.assertTrue(any(o.startswith("BOM_A_B_C___") and o.endswith(".xlsx")
                            for o in outs), outs)


# ===========================================================================
# P0-B3 逆向合并
# ===========================================================================
class TestMergeImport(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp(prefix="bom_v7_b3_", dir=TMP)
        # 准备可合并的 xlsx
        self.food_xlsx = os.path.join(self.d, "food.xlsx")
        self.mech_xlsx = os.path.join(self.d, "mech.xlsx")
        run_cli(GEN, "--data", self._dump(_sample_food("食品A", "2026-07-01")),
                "--out", self.food_xlsx)
        run_cli(GEN, "--data", self._dump(_sample_mech("机械B", "2026-07-02")),
                "--out", self.mech_xlsx)

    def _dump(self, data):
        p = os.path.join(self.d, "tmp_%s.json" % id(data))
        with open(p, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        return p

    def _merge(self, *paths, out="merged.json"):
        outp = os.path.join(self.d, out)
        rc = run_cli(IMP, "--in", *paths, "--merge", "--out", outp)
        return rc, outp

    def test_single_no_merge_no_extra_fields(self):
        """单 --in 无 --merge：V6 行为，无 merged_from/merge_notes。"""
        outp = os.path.join(self.d, "single.json")
        rc = run_cli(IMP, "--in", self.food_xlsx, "--out", outp)
        self.assertEqual(rc.returncode, 0, rc.stderr)
        data = json.load(open(outp, encoding="utf-8"))
        self.assertNotIn("merged_from", data)
        self.assertNotIn("merge_notes", data)
        self.assertEqual(data["industry"], "食品")

    def test_merge_two_files(self):
        rc, outp = self._merge(self.food_xlsx, self.mech_xlsx)
        self.assertEqual(rc.returncode, 0, rc.stderr)
        data = json.load(open(outp, encoding="utf-8"))
        self.assertEqual(data["industry"], "食品")  # 首个非空
        self.assertEqual(data["merged_from"], ["食品", "机械"])
        self.assertEqual(len(data["materials"]), 4 + 3)
        self.assertEqual(len(data["processes"]), 2 + 2)

    def test_merge_step_no_conflict_notes(self):
        """同 step_no 跨文件 → merge_notes 记录冲突，不重命名。"""
        # 两个机械文件，均含 S01/S02
        m1 = self.mech_xlsx
        m2 = os.path.join(self.d, "mech2.xlsx")
        run_cli(GEN, "--data", self._dump(_sample_mech("机械C", "2026-07-03")),
                "--out", m2)
        rc, outp = self._merge(m1, m2)
        self.assertEqual(rc.returncode, 0, rc.stderr)
        data = json.load(open(outp, encoding="utf-8"))
        notes = data.get("merge_notes", [])
        self.assertTrue(any("S01" in n and "重复" in n for n in notes), notes)
        self.assertTrue(any("S02" in n and "重复" in n for n in notes), notes)
        # 不重命名：仍保留 4 个工序（2+2）
        self.assertEqual(len(data["processes"]), 4)

    def test_merge_single_file_failure_skipped(self):
        """单文件解析失败跳过并记 merge_notes，不影响其余。"""
        bad = os.path.join(self.d, "bad.xlsx")
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = "garbage"; wb.save(bad)
        rc, outp = self._merge(self.food_xlsx, bad)
        self.assertEqual(rc.returncode, 0, rc.stderr)
        data = json.load(open(outp, encoding="utf-8"))
        self.assertEqual(data["industry"], "食品")
        self.assertEqual(len(data["materials"]), 4)
        self.assertTrue(any("跳过" in n for n in data.get("merge_notes", [])),
                        data.get("merge_notes"))

    def test_merge_all_fail(self):
        """全部失败 → MERGE_FAILED 退出码 2。"""
        bad = os.path.join(self.d, "bad.xlsx")
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = "garbage"; wb.save(bad)
        bad2 = os.path.join(self.d, "bad2.xlsx")
        wb2 = Workbook(); ws2 = wb2.active; ws2["A1"] = "garbage"; wb2.save(bad2)
        rc, outp = self._merge(bad, bad2)
        self.assertEqual(rc.returncode, 2, "全失败应退出码 2")
        self.assertIn("MERGE_FAILED", rc.stderr)

    def test_import_does_not_depend_on_i18n(self):
        """约束：import_bom.py 不得引用 I18N/ZH2EN（零回归）。"""
        src = open(IMP, encoding="utf-8").read()
        self.assertNotIn("I18N", src)
        self.assertNotIn("ZH2EN", src)


# ===========================================================================
# 边界 / 约束
# ===========================================================================
class TestBoundary(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp(prefix="bom_v7_bd_", dir=TMP)

    def test_empty_data_blank_template_opens(self):
        """空数据（空白模板）可正常生成与打开（B1）。"""
        rc = run_cli(GEN, "--blank-templates", "--out-dir", self.d,
                     "--industries", "通用")
        self.assertEqual(rc.returncode, 0, rc.stderr)
        wb = load_workbook(os.path.join(self.d, "template_通用.xlsx"))
        self.assertIn("BOM表", wb.sheetnames)

    def test_illegal_industry_falls_back(self):
        """非法 industry → 软告警 + 回落通用，仍成功生成。"""
        data = _sample_food()
        data["industry"] = "不存在的行业XYZ"
        p = os.path.join(self.d, "bad_ind.json")
        with open(p, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        out = os.path.join(self.d, "o.xlsx")
        rc = run_cli(GEN, "--data", p, "--out", out)
        self.assertEqual(rc.returncode, 0, "非法 industry 应回落而非失败")
        self.assertIn("WARNING", rc.stdout)  # V8 软校验
        self.assertTrue(os.path.exists(out))

    def test_missing_required_validation_failed(self):
        """缺必填（product_name/category/output_rate）→ VALIDATION_FAILED 退出码 2。"""
        for partial in [
            {"category": "食品", "output_rate": 100},          # 缺 product_name
            {"product_name": "X", "output_rate": 100},         # 缺 category
            {"product_name": "X", "category": "食品"},         # 缺 output_rate
        ]:
            p = os.path.join(self.d, "miss.json")
            with open(p, "w", encoding="utf-8") as f:
                json.dump(partial, f, ensure_ascii=False)
            rc = run_cli(GEN, "--data", p, "--out", os.path.join(self.d, "o.xlsx"))
            self.assertEqual(rc.returncode, 2, "缺必填应退出码 2: %r" % partial)
            self.assertIn("VALIDATION_FAILED", rc.stdout)

    def test_abnormal_json_in_batch_skipped(self):
        """batch 中异常 JSON 文件被跳过（错误隔离）。"""
        indir = tempfile.mkdtemp(prefix="bom_v7_ab_", dir=TMP)
        outdir = tempfile.mkdtemp(prefix="bom_v7_abo_", dir=TMP)
        with open(os.path.join(indir, "bad.json"), "w", encoding="utf-8") as f:
            f.write("not json <<<<")
        with open(os.path.join(indir, "ok.json"), "w", encoding="utf-8") as f:
            json.dump(_sample_food("芒果果味糖浆", "2026-07-07"), f, ensure_ascii=False)
        rc = run_cli(GEN, "--batch-dir", indir, "--out-dir", outdir)
        self.assertEqual(rc.returncode, 2)
        self.assertTrue(os.path.exists(
            os.path.join(outdir, "BOM_芒果果味糖浆_2026-07-07.xlsx")))

    def test_1000_materials_perf_baseline(self):
        """1000+ 物料性能基线：单语 + 双语均应在合理时间内完成。"""
        mats = [{"name": "M%d" % i, "unit": "kg", "usage": 1.0,
                 "yield_rate": 100, "erp_code": "RM-%d" % i,
                 "material_type": "原料", "process": ""}
                for i in range(1000)]
        data = _sample_food()
        data["materials"] = mats
        data["processes"] = []  # 纯物料 BOM，避免流转链校验开销
        p = os.path.join(self.d, "big.json")
        with open(p, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        out = os.path.join(self.d, "big.xlsx")
        t0 = time.time()
        rc = run_cli(GEN, "--data", p, "--out", out)
        dt = time.time() - t0
        self.assertEqual(rc.returncode, 0, rc.stderr)
        self.assertLess(dt, 30.0, "1000 物料单语生成应 < 30s，实际 %.1fs" % dt)
        # 双语同样基线
        out_bi = os.path.join(self.d, "big_bi.xlsx")
        t1 = time.time()
        rc = run_cli(GEN, "--data", p, "--out", out_bi, "--bilingual")
        dt_bi = time.time() - t1
        self.assertEqual(rc.returncode, 0, rc.stderr)
        self.assertLess(dt_bi, 45.0, "1000 物料双语生成应 < 45s，实际 %.1fs" % dt_bi)


# ===========================================================================
# 常量 / 约束单测
# ===========================================================================
class TestConstants(unittest.TestCase):
    def test_i18n_bidirectional(self):
        self.assertEqual(len(I18N), len(ZH2EN), "I18N/ZH2EN 应为可逆双射")
        for k, v in I18N.items():
            self.assertEqual(ZH2EN[v], k)

    def test_all_industries_8(self):
        self.assertEqual(sorted(ALL_INDUSTRIES), sorted(INDUSTRIES))
        self.assertEqual(len(ALL_INDUSTRIES), 8)

    def test_material_type_mapping_present(self):
        # 样本用到的中文 material_type 必须可查英文
        for zh in ["原料", "主材", "辅材", "包材"]:
            self.assertIn(zh, ZH2EN, "物料类型 %s 应可翻译" % zh)


if __name__ == "__main__":
    unittest.main(verbosity=2)
