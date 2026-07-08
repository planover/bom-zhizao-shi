#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BOM智造师 V2 独立测试脚本（QA · 严过关）

不依赖 pytest，使用 subprocess 调用 generate_bom.py / import_bom.py，
并用 openpyxl 读回 xlsx 断言；assert + 退出码；打印 PASS/FAIL 汇总。

覆盖用例：
  T01 正向happy（食品多工序）
  T02 逆向闭环（import → JSON → 再 generate）
  T03 R3 流转链断链（阻断）
  T04 产品名称空
  T05 类别非法（非法枚举 / 缺失）
  T06 出品率（负向 -5/0；正向 130>100 成功）
  T07 物料出品率越界（120 失败 / 100 成功）
  T08 非食品不生成配料表
  T09 纯物料无工序（平铺、无分组子标题）
  T10 旧版 5 列 Excel 兼容（手造最小旧格式）
  T11 未分类物料（空/其他）排除出配料表
  T12 SVG/文档存在性与内容

运行：
  python3 tests/test_bom_v2.py
退出码：全部通过 0，否则 1。
"""

import os
import sys
import json
import subprocess
import tempfile

from openpyxl import load_workbook, Workbook
import xml.etree.ElementTree as ET

# ----------------------------------------------------------------------------
# 路径常量
# ----------------------------------------------------------------------------
SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GEN = os.path.join(SKILL_DIR, "scripts", "generate_bom.py")
IMP = os.path.join(SKILL_DIR, "scripts", "import_bom.py")
SAMPLE = os.path.join(SKILL_DIR, "examples", "sample_bom_v2.json")
SVG = os.path.join(SKILL_DIR, "references", "bom-demo.svg")
README = os.path.join(SKILL_DIR, "README.md")
CHANGELOG = os.path.join(SKILL_DIR, "CHANGELOG.md")
PYTHON = r"C:\Users\姓名\.workbuddy\binaries\python\versions\3.13.12\python.exe"


# ----------------------------------------------------------------------------
# 运行辅助
# ----------------------------------------------------------------------------
def run_generate(data, out_path):
    """把 dict 写成临时 json，调用 generate_bom.py，返回 (rc, stdout, stderr)。"""
    tmp = out_path + ".in.json"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    p = subprocess.run(
        [PYTHON, GEN, "--data", tmp, "--out", out_path],
        capture_output=True, text=True,
    )
    return p.returncode, p.stdout, p.stderr


def run_import(xlsx_path, out_json=None):
    """调用 import_bom.py，返回 (rc, stdout, stderr, parsed_json_or_None)。"""
    args = [PYTHON, IMP, "--in", xlsx_path]
    if out_json:
        args += ["--out", out_json]
    p = subprocess.run(args, capture_output=True, text=True)
    parsed = None
    if out_json and os.path.exists(out_json):
        with open(out_json, encoding="utf-8") as f:
            parsed = json.load(f)
    return p.returncode, p.stdout, p.stderr, parsed


def flat_values(ws):
    """返回工作表中所有非空单元格的字符串列表。"""
    vals = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is not None and str(v).strip() != "":
                vals.append(str(v))
    return vals


def marker_row(ws, marker):
    """返回首列包含 marker 的行号（1-based），未找到返回 None。"""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and marker in str(v):
            return r
    return None


def get_block_names(ws, marker):
    """读取『marker』区块（如 三、配料表）的数据行首列名称列表。"""
    m = marker_row(ws, marker)
    if m is None:
        return []
    names = []
    r = m + 2  # m+1 为表头行，数据从 m+2 起
    while r <= ws.max_row:
        v = ws.cell(r, 1).value
        if v is None or str(v).strip() == "":
            break
        if str(v).strip().startswith("【"):
            r += 1
            continue
        names.append(str(v).strip())
        r += 1
    return names


# ----------------------------------------------------------------------------
# 用例
# ----------------------------------------------------------------------------
def test_t01():
    wd = tempfile.mkdtemp(prefix="bomqa_t01_")
    out = os.path.join(wd, "BOM_食品多工序.xlsx")
    with open(SAMPLE, encoding="utf-8") as f:
        sample = json.load(f)
    rc, so, se = run_generate(sample, out)
    assert rc == 0, "T01 生成应成功(退出0), 实际 rc=%s, stderr=%s" % (rc, se)
    assert "OK:" in so, "T01 stdout 应含 OK:, 实际=%s" % so

    wb = load_workbook(out)
    ws = wb.active

    # V3 物料区为 8 列 A–H（含首列「序号」）
    headers = [str(ws.cell(7, c).value or "").strip() for c in range(1, 9)]
    assert headers == ["序号", "物料名称", "单位", "用量", "出品率(%)",
                       "ERP物料代码", "物料类型", "所属工序"], \
        "T01 物料表头8列不符(V3): %s" % headers
    assert ws.max_column >= 8, "T01 max_column 应>=8(V3), 实际=%s" % ws.max_column

    # 行4：产品类别 + 全产品出品率
    row4 = " | ".join(str(ws.cell(4, c).value or "") for c in range(1, 8))
    assert "产品类别：食品" in row4, "T01 行4 应含『产品类别：食品』, 实际=%s" % row4
    exp_rate = "全产品出品率：130.0%"
    assert exp_rate in row4, \
        "T01 行4 应含『%s』(设计格式 0.0%%), 实际=%s" % (exp_rate, row4)

    flat = flat_values(ws)
    assert "【工序 S01 调配】" in flat, "T01 物料区应含『【工序 S01 调配】』"
    assert "【工序 S02 灌装】" in flat, "T01 物料区应含『【工序 S02 灌装】』"
    assert "【未归属工序】" in flat, "T01 物料区应含『【未归属工序』(PE瓶)"
    assert "PE 瓶" in flat, "T01 物料区应含 PE瓶"

    # 三、配料表
    ing = get_block_names(ws, "三、配料表")
    assert len(ing) == 4, "T01 配料表行数应为4, 实际=%s" % ing
    assert ing == ["芒果果味糖浆基料", "芒果原浆", "白砂糖", "柠檬酸"], \
        "T01 配料表应按 usage 降序[70.0,46.3,30.0,0.5], 实际=%s" % ing
    assert "PE 瓶" not in ing, "T01 配料表不应含 PE瓶(包材)"


def test_t02():
    wd = tempfile.mkdtemp(prefix="bomqa_t02_")
    out = os.path.join(wd, "BOM_t02.xlsx")
    with open(SAMPLE, encoding="utf-8") as f:
        sample = json.load(f)
    rc, so, se = run_generate(sample, out)
    assert rc == 0, "T02 生成应成功, rc=%s se=%s" % (rc, se)

    back = os.path.join(wd, "back.json")
    rc2, so2, se2, data = run_import(out, back)
    assert rc2 == 0, "T02 逆向解析应成功, rc=%s so2=%s se2=%s" % (rc2, so2, se2)
    assert data is not None, "T02 应解析出 JSON"

    assert data.get("category") == "食品", "T02 category 应=食品, 实际=%s" % data.get("category")
    assert data.get("output_rate") == 130.0, "T02 output_rate 应=130.0, 实际=%r" % data.get("output_rate")

    mats = {m["name"]: m for m in data.get("materials", [])}
    assert "芒果原浆" in mats and \
        mats["芒果原浆"].get("material_type") == "原料" and \
        mats["芒果原浆"].get("process") == "S01", \
        "T02 芒果原浆 material_type/process 应正确, 实际=%s" % mats.get("芒果原浆")
    assert "PE 瓶" in mats and \
        mats["PE 瓶"].get("material_type") == "包材" and \
        mats["PE 瓶"].get("process") == "", \
        "T02 PE瓶 material_type=包材/process=空, 实际=%s" % mats.get("PE 瓶")

    procs = {p["step_no"]: p for p in data.get("processes", [])}
    assert procs.get("S01", {}).get("output") == "芒果果味糖浆基料", \
        "T02 S01.output 应=芒果果味糖浆基料, 实际=%s" % procs.get("S01")
    assert procs.get("S02", {}).get("output") == "芒果果味糖浆", \
        "T02 S02.output 应=芒果果味糖浆, 实际=%s" % procs.get("S02")

    # JSON 不含配料表字段（派生不回写）
    assert "ingredients" not in data, "T02 JSON 不应含配料表字段(派生不回写)"

    # 再次喂回 generate 成功（闭环）
    out2 = os.path.join(wd, "BOM_t02_round2.xlsx")
    rc3, so3, se3 = run_generate(data, out2)
    assert rc3 == 0, "T02 闭环重生成应成功, rc=%s so3=%s se3=%s" % (rc3, so3, se3)
    assert "OK:" in so3, "T02 闭环重生成应含 OK:, 实际=%s" % so3


def test_t03():
    wd = tempfile.mkdtemp(prefix="bomqa_t03_")
    out = os.path.join(wd, "BOM_t03.xlsx")
    data = {
        "product_name": "测试", "category": "食品", "output_rate": 130,
        "materials": [
            {"name": "A", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S01"},
            {"name": "B", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S02"},
        ],
        "processes": [
            {"step_no": "S01", "name": "p1", "output": "OUT1"},
            {"step_no": "S02", "name": "p2", "output": "OUT2"},
        ],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 2, "T03 R3断链应退出码2, 实际 rc=%s" % rc
    assert "VALIDATION_FAILED" in so, "T03 stdout 应含 VALIDATION_FAILED, 实际=%s" % so
    assert "流转链不完整" in so, "T03 应提示『流转链不完整』, 实际=%s" % so


def test_t04():
    wd = tempfile.mkdtemp(prefix="bomqa_t04_")
    out = os.path.join(wd, "BOM_t04.xlsx")
    data = {
        "product_name": "", "category": "食品", "output_rate": 130,
        "materials": [{"name": "A", "unit": "kg", "usage": 10,
                       "yield_rate": 100, "material_type": "原料"}],
        "processes": [],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 2, "T04 产品名称为空应退出码2, 实际 rc=%s" % rc
    assert "VALIDATION_FAILED" in so, "T04 应含 VALIDATION_FAILED, 实际=%s" % so
    assert "产品名称为必填，且不可为空" in so, \
        "T04 应含『产品名称为必填，且不可为空』, 实际=%s" % so


def test_t05():
    wd = tempfile.mkdtemp(prefix="bomqa_t05_")
    # 非法枚举
    out = os.path.join(wd, "BOM_t05.xlsx")
    data = {
        "product_name": "X", "category": "饮料", "output_rate": 130,
        "materials": [{"name": "A", "unit": "kg", "usage": 10,
                       "yield_rate": 100, "material_type": "原料"}],
        "processes": [],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 2, "T05 类别非法应退出码2, 实际 rc=%s" % rc
    assert "VALIDATION_FAILED" in so, "T05 应含 VALIDATION_FAILED, 实际=%s" % so
    assert "产品类别为必填，且须为：食品/工业品/日化化妆品/医药/其他" in so, \
        "T05 应含类别枚举文案, 实际=%s" % so
    # 缺失类别
    out2 = os.path.join(wd, "BOM_t05b.xlsx")
    data2 = {
        "product_name": "X", "output_rate": 130,
        "materials": [{"name": "A", "unit": "kg", "usage": 10,
                       "yield_rate": 100, "material_type": "原料"}],
        "processes": [],
    }
    rc2, so2, se2 = run_generate(data2, out2)
    assert rc2 == 2, "T05 类别缺失应退出码2, 实际 rc=%s" % rc2
    assert "产品类别为必填，且须为：食品/工业品/日化化妆品/医药/其他" in so2, \
        "T05 缺失类别应含类别文案, 实际=%s" % so2


def test_t06():
    base = {
        "product_name": "X", "category": "工业品",
        "materials": [{"name": "A", "unit": "kg", "usage": 10,
                       "yield_rate": 100, "material_type": "原料"}],
        "processes": [],
    }
    wd = tempfile.mkdtemp(prefix="bomqa_t06_")
    for bad in (-5, 0):
        d = dict(base)
        d["output_rate"] = bad
        out = os.path.join(wd, "BOM_t06_%s.xlsx" % bad)
        rc, so, se = run_generate(d, out)
        assert rc == 2, "T06 output_rate=%s 应退出码2, 实际 rc=%s" % (bad, rc)
        assert "VALIDATION_FAILED" in so, "T06 output_rate=%s 应含 VALIDATION_FAILED, 实际=%s" % (bad, so)
        assert "全产品出品率(output_rate)" in so, \
            "T06 output_rate=%s 应含出品率文案, 实际=%s" % (bad, so)
    # 正向：130 (>100) 成功
    d = dict(base)
    d["output_rate"] = 130
    out = os.path.join(wd, "BOM_t06_130.xlsx")
    rc, so, se = run_generate(d, out)
    assert rc == 0, "T06 output_rate=130(>100) 应成功, 实际 rc=%s so=%s se=%s" % (rc, so, se)
    assert "OK:" in so, "T06 output_rate=130 应含 OK:, 实际=%s" % so


def test_t07():
    wd = tempfile.mkdtemp(prefix="bomqa_t07_")
    # 负向：yield_rate=120
    out = os.path.join(wd, "BOM_t07_120.xlsx")
    d = {
        "product_name": "X", "category": "工业品", "output_rate": 100,
        "materials": [{"name": "A", "unit": "kg", "usage": 10,
                       "yield_rate": 120, "material_type": "原料"}],
        "processes": [],
    }
    rc, so, se = run_generate(d, out)
    assert rc == 2, "T07 yield_rate=120 应退出码2, 实际 rc=%s" % rc
    assert "VALIDATION_FAILED" in so, "T07 应含 VALIDATION_FAILED, 实际=%s" % so
    assert "出品率须为 0-100" in so, "T07 应含『出品率须为 0-100』, 实际=%s" % so
    # 正向：yield_rate=100 成功
    out2 = os.path.join(wd, "BOM_t07_100.xlsx")
    d2 = {
        "product_name": "X", "category": "工业品", "output_rate": 100,
        "materials": [{"name": "A", "unit": "kg", "usage": 10,
                       "yield_rate": 100, "material_type": "原料"}],
        "processes": [],
    }
    rc2, so2, se2 = run_generate(d2, out2)
    assert rc2 == 0, "T07 yield_rate=100 应成功, 实际 rc=%s so=%s" % (rc2, so2)
    assert "OK:" in so2, "T07 yield_rate=100 应含 OK:, 实际=%s" % so2


def test_t08():
    wd = tempfile.mkdtemp(prefix="bomqa_t08_")
    out = os.path.join(wd, "BOM_t08.xlsx")
    data = {
        "product_name": "工业品X", "category": "工业品", "output_rate": 100,
        "materials": [
            {"name": "原料A", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S01"},
            {"name": "包材B", "unit": "个", "usage": 5, "yield_rate": 100,
             "material_type": "包材"},
        ],
        "processes": [{"step_no": "S01", "name": "p1", "output": "OUT1"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T08 工业品生成应成功, 实际 rc=%s so=%s se=%s" % (rc, so, se)
    assert "OK:" in so, "T08 应含 OK:, 实际=%s" % so
    wb = load_workbook(out)
    ws = wb.active
    flat = flat_values(ws)
    assert "三、配料表" not in flat, "T08 非食品不应有『三、配料表』区块"
    assert "WARNING" not in so, "T08 非食品不应有配料表 WARNING, 实际 stdout=%s" % so


def test_t09():
    wd = tempfile.mkdtemp(prefix="bomqa_t09_")
    out = os.path.join(wd, "BOM_t09.xlsx")
    data = {
        "product_name": "纯物料X", "category": "工业品", "output_rate": 100,
        "materials": [
            {"name": "原料A", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料"},
            {"name": "包材B", "unit": "个", "usage": 5, "yield_rate": 100,
             "material_type": "包材"},
        ],
        "processes": [],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T09 纯物料(无工序)应成功, 实际 rc=%s so=%s se=%s" % (rc, so, se)
    assert "OK:" in so, "T09 应含 OK:, 实际=%s" % so
    wb = load_workbook(out)
    ws = wb.active
    flat = flat_values(ws)
    assert "【" not in " ".join(flat), "T09 物料区不应有分组子标题(应平铺), 实际含【"


def test_t10():
    # 手造最小旧版 5 列 Excel（无产品名称/类别/出品率行/无物料类型/所属工序/产物列）
    wd = tempfile.mkdtemp(prefix="bomqa_t10_")
    old = os.path.join(wd, "old5.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM表"
    ws["A1"] = "BOM表"
    ws["A2"] = "版本号：V1.0"
    ws["D2"] = "生成日期：2024-01-01"
    ws["A3"] = "一、物料信息"
    ws["A4"] = "物料名称"; ws["B4"] = "单位"; ws["C4"] = "用量"
    ws["D4"] = "出品率(%)"; ws["E4"] = "ERP物料代码"
    ws["A5"] = "旧物料"; ws["B5"] = "kg"; ws["C5"] = 5
    ws["D5"] = 100; ws["E5"] = "RM-OLD"
    ws["A6"] = "二、工艺工序"
    ws["A7"] = "工序编号"; ws["B7"] = "工序名称"; ws["C7"] = "工序说明"
    ws["D7"] = "工时"; ws["E7"] = "备注"
    ws["A8"] = "S01"; ws["B8"] = "旧工序"; ws["C8"] = "说明"
    ws["D8"] = 10; ws["E8"] = "备注"
    wb.save(old)

    out_json = os.path.join(wd, "back.json")
    rc, so, se, data = run_import(old, out_json)
    assert (rc != 2) or ("PARSE_ERROR" not in so and "FILE_ERROR" not in so), \
        "T10 旧版解析不应崩溃/报PARSE_ERROR/FILE_ERROR, rc=%s so=%s se=%s" % (rc, so, se)
    assert data is not None, "T10 应解析出 JSON"
    assert data.get("category") == "其他", "T10 category 默认应=其他, 实际=%s" % data.get("category")
    assert data.get("output_rate") == "", "T10 output_rate 默认应=空串, 实际=%r" % data.get("output_rate")
    mats = data.get("materials", [])
    assert len(mats) == 1, "T10 应解析出1条物料, 实际=%s" % mats
    assert mats[0].get("material_type") == "其他", \
        "T10 物料 material_type 默认=其他, 实际=%s" % mats[0].get("material_type")
    assert mats[0].get("process") == "", "T10 物料 process 默认=空, 实际=%r" % mats[0].get("process")
    procs = data.get("processes", [])
    assert len(procs) == 1, "T10 应解析出1条工序, 实际=%s" % procs
    assert procs[0].get("output") == "", "T10 工序 output 默认=空, 实际=%r" % procs[0].get("output")


def test_t11():
    wd = tempfile.mkdtemp(prefix="bomqa_t11_")
    out = os.path.join(wd, "BOM_t11.xlsx")
    data = {
        "product_name": "测试食品", "category": "食品", "output_rate": 100,
        "materials": [
            {"name": "主料A", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S01"},
            {"name": "空类型B", "unit": "kg", "usage": 5, "yield_rate": 100,
             "material_type": "", "process": "S01"},
            {"name": "其他C", "unit": "kg", "usage": 3, "yield_rate": 100,
             "material_type": "其他", "process": "S01"},
        ],
        "processes": [{"step_no": "S01", "name": "工序1", "output": "产物X"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T11 生成应成功, rc=%s so=%s se=%s" % (rc, so, se)
    wb = load_workbook(out)
    ws = wb.active
    ing = get_block_names(ws, "三、配料表")
    assert ing == ["主料A"], \
        "T11 配料表应只含可食用『主料A』(空/其他被排除), 实际=%s" % ing
    assert "空类型B" not in ing and "其他C" not in ing, \
        "T11 空/其他类型不应进配料表"


def test_t12():
    try:
        ET.parse(SVG)
    except Exception as e:
        raise AssertionError("T12 bom-demo.svg 不是合法 XML: %s" % e)
    with open(README, encoding="utf-8") as f:
        rd = f.read()
    assert "bom-demo.svg" in rd, "T12 README 应引用 bom-demo.svg"
    assert "已知限制" in rd, "T12 README 应含『已知限制』小节"
    with open(CHANGELOG, encoding="utf-8") as f:
        cl = f.read()
    assert "V2" in cl, "T12 CHANGELOG 应含 V2 记录"


# ----------------------------------------------------------------------------
# 运行器
# ----------------------------------------------------------------------------
TESTS = [
    ("T01", "正向happy(食品多工序)", test_t01),
    ("T02", "逆向闭环", test_t02),
    ("T03", "R3流转链断链(阻断)", test_t03),
    ("T04", "产品名称空", test_t04),
    ("T05", "类别非法", test_t05),
    ("T06", "出品率(-5/0 vs 130>100)", test_t06),
    ("T07", "物料出品率越界(120/100)", test_t07),
    ("T08", "非食品不生成配料表", test_t08),
    ("T09", "纯物料无工序(平铺)", test_t09),
    ("T10", "旧版5列Excel兼容", test_t10),
    ("T11", "未分类物料排除", test_t11),
    ("T12", "SVG/文档存在性", test_t12),
]


def main():
    print("==================================================")
    print(" BOM智造师 V2 独立测试 (QA 严过关)")
    print("==================================================")
    results = []
    for tid, name, fn in TESTS:
        try:
            fn()
            results.append((tid, name, True, "PASS"))
            print("%-4s %-26s PASS" % (tid, name))
        except AssertionError as e:
            results.append((tid, name, False, str(e)))
            print("%-4s %-26s FAIL  %s" % (tid, name, str(e)))
        except Exception as e:  # noqa: BLE001
            results.append((tid, name, False, "EXCEPTION: %s" % e))
            print("%-4s %-26s FAIL  EXCEPTION: %s" % (tid, name, e))

    total = len(results)
    passed = sum(1 for r in results if r[2])
    failed = total - passed
    print("--------------------------------------------------")
    print("总计: %d   通过: %d   失败: %d" % (total, passed, failed))
    print("IS_PASS: %s" % ("YES" if failed == 0 else "NO"))
    if failed:
        print("失败用例:")
        for tid, name, ok, detail in results:
            if not ok:
                print("  [%s] %s -> %s" % (tid, name, detail))
    print("--------------------------------------------------")
    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
