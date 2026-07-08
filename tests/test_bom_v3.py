#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BOM智造师 V3 (V2.1) 独立测试脚本（QA · 严过关）

不依赖 pytest，使用 subprocess 调用 generate_bom.py / import_bom.py，
并用 openpyxl 读回 xlsx 断言；assert + 退出码；打印 PASS/FAIL 汇总。

覆盖 V3（8 列 A–H）全部新增行为 + 原有 V1–V5 校验回归：
  V3-T00 8 列表头（含「序号」）
  V3-T01 占比% 列和=100.0（最大余数补差，sample_v3）
  V3-T02 占比% 分母=可食用用量合计（排除包材/其他）
  V3-T03 序号连续、跨工序组不重置（sample_v3 → [1,2,3,4,5]）
  V3-T04 配料表仅食品生成；非食品仍含「合计用量」行
  V3-T05 过敏原列展示 + 逆向回收（芒果原浆 allergen==大豆,乳）
  V3-T06 审批人/生效日期/执行标准 显示 + 逆向
  V3-T07 合计行（D=全部 usage 求和 246.8）；逆向跳过合计行
  V3-T08 旧版 7 列 Excel 兼容（sample_bom_v2.xlsx，退出 0，默认空）
  V3-T09 过敏原软校验 WARNING（非法标签，非阻断，退出 0）
  V3-T10 全产品出品率 130.0%（含 1 位小数，非 130%）
  V3-T11 稳定排序（usage 降序；同用量保持输入顺序确定性）
  V3-T12 V1–V5 校验仍生效（空名/R3断链/output_rate≤0 → 退出 2）
  V3-T13 闭环一致性（generate→import 关键字段一致；JSON 不含序号/用量占比%）

运行：
  python3 tests/test_bom_v3.py
退出码：全部通过 0，否则 1。
"""

import os
import sys
import json
import shutil
import subprocess
import tempfile

from openpyxl import load_workbook

# ----------------------------------------------------------------------------
# 路径常量
# ----------------------------------------------------------------------------
SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GEN = os.path.join(SKILL_DIR, "scripts", "generate_bom.py")
IMP = os.path.join(SKILL_DIR, "scripts", "import_bom.py")
SAMPLE_V3 = os.path.join(SKILL_DIR, "examples", "sample_bom_v3.json")
SAMPLE_V3_XLSX = os.path.join(SKILL_DIR, "examples", "sample_bom_v3.xlsx")
SAMPLE_V2_XLSX = os.path.join(SKILL_DIR, "examples", "sample_bom_v2.xlsx")
PYTHON = r"C:\Users\姓名\.workbuddy\binaries\python\versions\3.13.12\python.exe"

SAMPLE_V3_TOTAL_USAGE = 246.8  # 46.3+30.0+70.0+0.5+100(PE瓶)


# ----------------------------------------------------------------------------
# 运行辅助
# ----------------------------------------------------------------------------
def run_generate(data, out_path):
    """把 dict 写成临时 json，调用 generate_bom.py，返回 (rc, stdout, stderr)。"""
    tmp = out_path + ".in.json"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    p = subprocess.run(
        [PYTHON, GEN, "--data", tmp, "--out", out_path],
        capture_output=True, text=True,
    )
    return p.returncode, p.stdout, p.stderr


def run_import(xlsx_path, out_json=None):
    """调用 import_bom.py，返回 (rc, stdout, stderr, parsed_json_or_None)。"""
    args = [PYTHON, IMP, "--in", xlsx_path]
    if out_json:
        args += ["--out", out_json]
    p = subprocess.run(args, capture_output=True, text=True)
    parsed = None
    if out_json and os.path.exists(out_json):
        with open(out_json, encoding="utf-8") as f:
            parsed = json.load(f)
    return p.returncode, p.stdout, p.stderr, parsed


def load(path):
    return load_workbook(path).active


def flat_values(ws):
    """返回工作表中所有非空单元格的字符串列表。"""
    vals = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is not None and str(v).strip() != "":
                vals.append(str(v))
    return vals


def marker_row(ws, marker):
    """返回首列包含 marker 的行号（1-based），未找到返回 None。"""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and marker in str(v):
            return r
    return None


def material_header_row(ws):
    """定位物料表头行（某列值为『物料名称』）。"""
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value or "").strip() == "物料名称":
            return r
    return None


def material_seqs(ws):
    """读取物料区序号列（A 列）序列，跳过分组子标题与合计行。"""
    hdr = material_header_row(ws)
    assert hdr is not None, "物料表头行未找到（无『物料名称』）"
    seqs = []
    r = hdr + 1
    while r <= ws.max_row:
        a = ws.cell(r, 1).value
        if a is None:
            break
        s = str(a).strip()
        if s.startswith("【"):
            r += 1
            continue
        if s == "合计":
            break
        try:
            seqs.append(int(float(a)))
        except (TypeError, ValueError):
            break
        r += 1
    return seqs


def total_usage_row(ws):
    """返回合计行（A 列==合计）的行号，未找到返回 None。"""
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == "合计":
            return r
    return None


def ingredient_block(ws):
    """读取『三、配料表』数据行，返回 [(name, pct, allergen), ...]。"""
    m = marker_row(ws, "三、配料表")
    if m is None:
        return []
    rows = []
    r = m + 2  # m+1 表头，m+2 起为数据
    while r <= ws.max_row:
        v = ws.cell(r, 1).value
        if v is None or str(v).strip() == "":
            break
        if str(v).strip().startswith("【"):
            r += 1
            continue
        name = str(ws.cell(r, 1).value or "").strip()
        pct = ws.cell(r, 6).value  # F 列 用量占比%
        allergen = str(ws.cell(r, 7).value or "").strip()  # G 列 过敏原
        rows.append((name, pct, allergen))
        r += 1
    return rows


def row4_text(ws):
    return " | ".join(str(ws.cell(4, c).value or "") for c in range(1, 9))


# ----------------------------------------------------------------------------
# 用例
# ----------------------------------------------------------------------------
def test_v3_t00_header8():
    """物料表头应为 8 列（含首列『序号』）。"""
    wd = tempfile.mkdtemp(prefix="v3t00_")
    out = os.path.join(wd, "BOM_t00.xlsx")
    with open(SAMPLE_V3, encoding="utf-8") as f:
        sample = json.load(f)
    rc, so, se = run_generate(sample, out)
    assert rc == 0, "T00 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    hdr = material_header_row(ws)
    headers = [str(ws.cell(hdr, c).value or "").strip() for c in range(1, 9)]
    assert headers == ["序号", "物料名称", "单位", "用量", "出品率(%)",
                       "ERP物料代码", "物料类型", "所属工序"], \
        "T00 8 列物料表头不符: %s" % headers
    assert ws.max_column == 8, "T00 max_column 应=8, 实际=%s" % ws.max_column


def test_v3_t01_pct_sum_100():
    """占比% 列和恰为 100.0（最大余数补差）。"""
    wd = tempfile.mkdtemp(prefix="v3t01_")
    out = os.path.join(wd, "BOM_t01.xlsx")
    with open(SAMPLE_V3, encoding="utf-8") as f:
        sample = json.load(f)
    rc, so, se = run_generate(sample, out)
    assert rc == 0, "T01 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    block = ingredient_block(ws)
    assert len(block) == 4, "T01 配料表应 4 行(可食用), 实际=%s" % block
    pcts = [b[1] for b in block]
    # 实际期望（usage 降序 + 最大余数补差）:
    # 70.0->47.7, 46.3->31.5, 30.0->20.4, 0.5->0.4
    names = [b[0] for b in block]
    assert names == ["芒果果味糖浆基料", "芒果原浆", "白砂糖", "柠檬酸"], \
        "T01 配料表应按 usage 降序, 实际=%s" % names
    assert pcts == [47.7, 31.5, 20.4, 0.4], \
        "T01 占比% 应 [47.7,31.5,20.4,0.4], 实际=%s" % pcts
    s = sum(float(p) for p in pcts)
    assert abs(s - 100.0) < 1e-6, "T01 占比% 列和应=100.0, 实际=%s" % s
    assert round(s, 1) == 100.0, "T01 占比% 列和(1位)应=100.0, 实际=%s" % round(s, 1)


def test_v3_t02_pct_denominator_edible():
    """占比% 分母=可食用用量合计（排除包材/其他）。"""
    wd = tempfile.mkdtemp(prefix="v3t02_")
    out = os.path.join(wd, "BOM_t02.xlsx")
    data = {
        "product_name": "占比分母测试", "category": "食品", "output_rate": 130,
        "materials": [
            {"name": "A酱", "unit": "kg", "usage": 70, "yield_rate": 98,
             "material_type": "原料", "process": ""},
            {"name": "B糖", "unit": "kg", "usage": 30, "yield_rate": 100,
             "material_type": "原料", "process": ""},
            {"name": "C瓶", "unit": "个", "usage": 100, "yield_rate": 100,
             "material_type": "包材", "process": ""},
        ],
        "processes": [],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T02 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    block = ingredient_block(ws)
    names = [b[0] for b in block]
    # 分母应为可食用合计=70+30=100；C瓶(包材)排除
    assert "C瓶" not in names, "T02 包材不应进配料表, 实际=%s" % names
    assert names == ["A酱", "B糖"], "T02 配料表应仅含可食用[A酱,B糖], 实际=%s" % names
    pcts = [float(b[1]) for b in block]
    assert abs(sum(pcts) - 100.0) < 1e-6, "T02 占比% 和应=100.0(分母=可食用), 实际=%s" % sum(pcts)


def test_v3_t03_seq_continuous():
    """序号连续、跨工序分组不重置 → [1,2,3,4,5]。"""
    wd = tempfile.mkdtemp(prefix="v3t03_")
    out = os.path.join(wd, "BOM_t03.xlsx")
    with open(SAMPLE_V3, encoding="utf-8") as f:
        sample = json.load(f)
    rc, so, se = run_generate(sample, out)
    assert rc == 0, "T03 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    seqs = material_seqs(ws)
    assert seqs == [1, 2, 3, 4, 5], "T03 序号应连续跨组不重置=[1..5], 实际=%s" % seqs


def test_v3_t04_ingredient_only_food():
    """配料表仅食品生成；非食品仍含『合计用量』行。"""
    # 非食品
    wd = tempfile.mkdtemp(prefix="v3t04_")
    out = os.path.join(wd, "BOM_t04_nonfood.xlsx")
    data = {
        "product_name": "工业品X", "category": "工业品", "output_rate": 100,
        "materials": [
            {"name": "原料A", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S01"},
            {"name": "包材B", "unit": "个", "usage": 5, "yield_rate": 100,
             "material_type": "包材"},
        ],
        "processes": [{"step_no": "S01", "name": "p1", "output": "OUT1"}],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T04 非食品生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    assert marker_row(ws, "三、配料表") is None, "T04 非食品不应有配料表"
    tr = total_usage_row(ws)
    assert tr is not None, "T04 非食品物料区仍应有『合计用量』行"
    assert ws.cell(tr, 4).value == 15.0, \
        "T04 合计用量 D 应=10+5=15.0, 实际=%s" % ws.cell(tr, 4).value
    # 食品
    out2 = os.path.join(wd, "BOM_t04_food.xlsx")
    data2 = dict(data)
    data2["category"] = "食品"
    rc2, so2, se2 = run_generate(data2, out2)
    assert rc2 == 0, "T04 食品生成应成功, rc=%s se=%s" % (rc2, se2)
    ws2 = load(out2)
    assert marker_row(ws2, "三、配料表") is not None, "T04 食品应有配料表"


def test_v3_t05_allergen_display_and_recover():
    """过敏原列展示 + 逆向回收（芒果原浆 allergen==大豆,乳）。"""
    wd = tempfile.mkdtemp(prefix="v3t05_")
    out = os.path.join(wd, "BOM_t05.xlsx")
    with open(SAMPLE_V3, encoding="utf-8") as f:
        sample = json.load(f)
    rc, so, se = run_generate(sample, out)
    assert rc == 0, "T05 生成应成功, rc=%s se=%s" % (rc, se)
    # 正向：配料表 芒果原浆 过敏原列
    ws = load(out)
    block = ingredient_block(ws)
    alg_map = {name: allergen for name, _, allergen in block}
    assert alg_map.get("芒果原浆") == "大豆,乳", \
        "T05 正向配料表 芒果原浆 过敏原应=大豆,乳, 实际=%s" % alg_map.get("芒果原浆")
    # 逆向：import xlsx → JSON 回收
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, data = run_import(out, back)
    assert rc2 == 0, "T05 逆向解析应成功, rc=%s se=%s" % (rc2, se2)
    mats = {m["name"]: m for m in data.get("materials", [])}
    assert mats.get("芒果原浆", {}).get("allergen") == "大豆,乳", \
        "T05 逆向 芒果原浆 allergen 应=大豆,乳, 实际=%s" % mats.get("芒果原浆")


def test_v3_t06_meta_display_and_recover():
    """审批人/生效日期/执行标准 显示 + 逆向。"""
    wd = tempfile.mkdtemp(prefix="v3t06_")
    out = os.path.join(wd, "BOM_t06.xlsx")
    with open(SAMPLE_V3, encoding="utf-8") as f:
        sample = json.load(f)
    rc, so, se = run_generate(sample, out)
    assert rc == 0, "T06 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    # 行5 合并单格
    meta = str(ws.cell(5, 1).value or "")
    assert "审批人：张三" in meta, "T06 行5 应含『审批人：张三』, 实际=%s" % meta
    assert "生效日期：2026-07-10" in meta, "T06 行5 应含『生效日期：2026-07-10』, 实际=%s" % meta
    assert "执行标准：GB 7718-2025" in meta, "T06 行5 应含『执行标准：GB 7718-2025』, 实际=%s" % meta
    # 逆向
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, data = run_import(out, back)
    assert rc2 == 0, "T06 逆向应成功, rc=%s se=%s" % (rc2, se2)
    assert data.get("approver") == "张三", "T06 逆向 approver 应=张三, 实际=%r" % data.get("approver")
    assert data.get("effective_date") == "2026-07-10", \
        "T06 逆向 effective_date 应=2026-07-10, 实际=%r" % data.get("effective_date")
    assert data.get("standard") == "GB 7718-2025", \
        "T06 逆向 standard 应=GB 7718-2025, 实际=%r" % data.get("standard")


def test_v3_t07_total_row_and_skip():
    """合计行 D=全部 usage 求和；逆向跳过合计行。"""
    wd = tempfile.mkdtemp(prefix="v3t07_")
    out = os.path.join(wd, "BOM_t07.xlsx")
    with open(SAMPLE_V3, encoding="utf-8") as f:
        sample = json.load(f)
    rc, so, se = run_generate(sample, out)
    assert rc == 0, "T07 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    tr = total_usage_row(ws)
    assert tr is not None, "T07 应有合计行"
    assert ws.cell(tr, 4).value == SAMPLE_V3_TOTAL_USAGE, \
        "T07 合计 D 应=%s, 实际=%s" % (SAMPLE_V3_TOTAL_USAGE, ws.cell(tr, 4).value)
    # 逆向：物料数应=5（不含合计行），且无名为合计的物料
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, data = run_import(out, back)
    assert rc2 == 0, "T07 逆向应成功, rc=%s se=%s" % (rc2, se2)
    mats = data.get("materials", [])
    assert len(mats) == 5, "T07 逆向物料应=5(跳过合计行), 实际=%s" % len(mats)
    assert all(str(m.get("name") or "") != "合计" for m in mats), \
        "T07 逆向不应含名为『合计』的物料"


def test_v3_t08_old7_compat():
    """旧版 7 列 Excel 兼容（sample_bom_v2.xlsx）：退出 0，默认空。"""
    assert os.path.exists(SAMPLE_V2_XLSX), "T08 旧版示例 sample_bom_v2.xlsx 缺失"
    wd = tempfile.mkdtemp(prefix="v3t08_")
    back = os.path.join(wd, "back.json")
    rc, so, se, data = run_import(SAMPLE_V2_XLSX, back)
    assert rc == 0, "T08 旧版 7 列解析应成功(rc=0), rc=%s so=%s se=%s" % (rc, so, se)
    assert data is not None, "T08 应解析出 JSON"
    assert data.get("approver") == "", "T08 approver 默认应=空, 实际=%r" % data.get("approver")
    assert data.get("effective_date") == "", "T08 effective_date 默认应=空, 实际=%r" % data.get("effective_date")
    assert data.get("standard") == "", "T08 standard 默认应=空, 实际=%r" % data.get("standard")
    for m in data.get("materials", []):
        assert m.get("allergen") == "", \
            "T08 旧版物料 allergen 默认应=空, 物料=%s 实际=%r" % (m.get("name"), m.get("allergen"))
    # 旧版无『序号』列 → 物料名称应正确被映射读取
    mats = {m["name"]: m for m in data.get("materials", [])}
    assert "芒果原浆" in mats, "T08 旧版应解析出『芒果原浆』"


def test_v3_t09_allergen_warning():
    """过敏原软校验 WARNING：非法标签 → 打印 WARNING 且退出码 0（非阻断）。"""
    wd = tempfile.mkdtemp(prefix="v3t09_")
    out = os.path.join(wd, "BOM_t09.xlsx")
    data = {
        "product_name": "告警测试", "category": "食品", "output_rate": 130,
        "materials": [
            {"name": "问题料", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "",
             "allergen": "树莓辣条"},  # 不在八大类集合
        ],
        "processes": [],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T09 非法标签不应阻断(rc=0), 实际 rc=%s so=%s se=%s" % (rc, so, se)
    assert "WARNING" in so, "T09 应打印 WARNING(非法过敏原标签), stdout=%s" % so
    assert "树莓辣条" in so, "T09 WARNING 应包含非法标签『树莓辣条』, stdout=%s" % so
    assert "OK:" in so, "T09 仍应输出 OK:(非阻断), 实际=%s" % so


def test_v3_t10_output_rate_format():
    """全产品出品率显示 130.0%（1 位小数，非 130%）。"""
    wd = tempfile.mkdtemp(prefix="v3t10_")
    out = os.path.join(wd, "BOM_t10.xlsx")
    with open(SAMPLE_V3, encoding="utf-8") as f:
        sample = json.load(f)
    rc, so, se = run_generate(sample, out)
    assert rc == 0, f"T10 生成应成功, rc={rc} se={se}"
    ws = load(out)
    r4 = row4_text(ws)
    assert "全产品出品率：130.0%" in r4, \
        f"T10 行4 应含『全产品出品率：130.0%』(1位小数), 实际={r4}"
    assert "全产品出品率：130%" not in r4, \
        f"T10 行4 不应为裸『全产品出品率：130%』(应带1位小数), 实际={r4}"


def test_v3_t11_stable_sort():
    """稳定排序：按 usage 降序；同用量保持确定性（输入顺序）。"""
    # 3 条不同用量 → usage 降序
    wd = tempfile.mkdtemp(prefix="v3t11_")
    out = os.path.join(wd, "BOM_t11.xlsx")
    data = {
        "product_name": "排序测试", "category": "食品", "output_rate": 130,
        "materials": [
            {"name": "X大", "unit": "kg", "usage": 30, "yield_rate": 100,
             "material_type": "原料", "process": ""},
            {"name": "Y小", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": ""},
            {"name": "Z中", "unit": "kg", "usage": 20, "yield_rate": 100,
             "material_type": "原料", "process": ""},
        ],
        "processes": [],
    }
    rc, so, se = run_generate(data, out)
    assert rc == 0, "T11 生成应成功, rc=%s se=%s" % (rc, se)
    ws = load(out)
    block = ingredient_block(ws)
    names = [b[0] for b in block]
    assert names == ["X大", "Z中", "Y小"], \
        "T11 应按 usage 降序 [30,20,10], 实际=%s" % names

    # 2 条同用量 → 确定性（稳定排序保持输入顺序），相邻且相对顺序守恒
    out2 = os.path.join(wd, "BOM_t11b.xlsx")
    data2 = {
        "product_name": "排序测试2", "category": "食品", "output_rate": 130,
        "materials": [
            {"name": "Z糖", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": ""},
            {"name": "A盐", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": ""},
        ],
        "processes": [],
    }
    rc2, so2, se2 = run_generate(data2, out2)
    assert rc2 == 0, "T11b 生成应成功, rc=%s se=%s" % (rc2, se2)
    ws2 = load(out2)
    block2 = ingredient_block(ws2)
    names2 = [b[0] for b in block2]
    assert names2 == ["Z糖", "A盐"], \
        "T11b 同用量应保持确定性顺序(实测=输入顺序), 实际=%s" % names2


def test_v3_t12_validation_regression():
    """V1–V5 校验仍生效：空名 / R3断链 / output_rate≤0 → 退出 2。"""
    wd = tempfile.mkdtemp(prefix="v3t12_")
    # V1: 产品名称为空
    out = os.path.join(wd, "BOM_t12_empty.xlsx")
    rc, so, se = run_generate({
        "product_name": "", "category": "食品", "output_rate": 130,
        "materials": [{"name": "A", "unit": "kg", "usage": 10,
                       "yield_rate": 100, "material_type": "原料"}],
        "processes": [],
    }, out)
    assert rc == 2, "T12 产品名空应退出2, 实际 rc=%s" % rc
    assert "VALIDATION_FAILED" in so, "T12 应含 VALIDATION_FAILED, 实际=%s" % so

    # V3: output_rate <= 0
    out2 = os.path.join(wd, "BOM_t12_rate.xlsx")
    rc2, so2, se2 = run_generate({
        "product_name": "X", "category": "食品", "output_rate": 0,
        "materials": [{"name": "A", "unit": "kg", "usage": 10,
                       "yield_rate": 100, "material_type": "原料"}],
        "processes": [],
    }, out2)
    assert rc2 == 2, "T12 output_rate=0 应退出2, 实际 rc=%s" % rc2

    # V5/R3: 工序流转链断裂
    out3 = os.path.join(wd, "BOM_t12_chain.xlsx")
    rc3, so3, se3 = run_generate({
        "product_name": "X", "category": "食品", "output_rate": 130,
        "materials": [
            {"name": "A", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S01"},
            {"name": "B", "unit": "kg", "usage": 10, "yield_rate": 100,
             "material_type": "原料", "process": "S02"},
        ],
        "processes": [
            {"step_no": "S01", "name": "p1", "output": "OUT1"},
            {"step_no": "S02", "name": "p2", "output": "OUT2"},
        ],
    }, out3)
    assert rc3 == 2, "T12 R3断链应退出2, 实际 rc=%s" % rc3
    assert "流转链不完整" in so3, "T12 应提示『流转链不完整』, 实际=%s" % so3


def test_v3_t13_closed_loop():
    """闭环一致性：generate→import 关键字段一致；JSON 不含序号/用量占比%。"""
    wd = tempfile.mkdtemp(prefix="v3t13_")
    out = os.path.join(wd, "BOM_t13.xlsx")
    with open(SAMPLE_V3, encoding="utf-8") as f:
        orig = json.load(f)
    rc, so, se = run_generate(orig, out)
    assert rc == 0, "T13 生成应成功, rc=%s se=%s" % (rc, se)
    back = os.path.join(wd, "back.json")
    rc2, so2, se2, data = run_import(out, back)
    assert rc2 == 0, "T13 逆向应成功, rc=%s se=%s" % (rc2, se2)
    assert data is not None, "T13 应解析出 JSON"

    # 顶层关键字段
    assert data.get("product_name") == orig["product_name"], "T13 product_name 不一致"
    assert data.get("category") == orig["category"], "T13 category 不一致"
    assert data.get("output_rate") == float(orig["output_rate"]), "T13 output_rate 不一致"
    assert data.get("approver") == orig.get("approver", ""), "T13 approver 不一致"
    assert data.get("effective_date") == orig.get("effective_date", ""), "T13 effective_date 不一致"
    assert data.get("standard") == orig.get("standard", ""), "T13 standard 不一致"

    # 物料关键字段（按名匹配）
    om = {m["name"]: m for m in orig["materials"]}
    dm = {m["name"]: m for m in data["materials"]}
    assert set(dm.keys()) == set(om.keys()), \
        "T13 物料集合不一致: orig=%s back=%s" % (set(om.keys()), set(dm.keys()))
    for name, om_ in om.items():
        dm_ = dm[name]
        assert float(dm_.get("usage", 0)) == float(om_["usage"]), \
            "T13 物料 %s usage 不一致" % name
        assert dm_.get("material_type") == om_.get("material_type"), \
            "T13 物料 %s material_type 不一致" % name
        assert dm_.get("process") == om_.get("process", ""), \
            "T13 物料 %s process 不一致" % name
        assert dm_.get("allergen") == om_.get("allergen", ""), \
            "T13 物料 %s allergen 不一致" % name

    # 工序 output
    op = {p["step_no"]: p for p in orig["processes"]}
    dp = {p["step_no"]: p for p in data["processes"]}
    assert set(dp.keys()) == set(op.keys()), "T13 工序集合不一致"
    for sn, op_ in op.items():
        assert dp[sn].get("output") == op_.get("output", ""), \
            "T13 工序 %s output 不一致" % sn

    # JSON 不应含派生展示键
    assert "序号" not in data, "T13 JSON 不应含『序号』"
    assert "用量占比%" not in data, "T13 JSON 不应含『用量占比%』"
    for m in data["materials"]:
        assert "序号" not in m, "T13 物料不应含『序号』"
        assert "用量占比%" not in m, "T13 物料不应含『用量占比%』"


def test_v3_t14a_keyword_warning():
    """V3-T14a: 名称含致敏物关键词(牛奶)但 allergen 空 → H1 WARNING, 退出码 0(非阻断)。"""
    wd = tempfile.mkdtemp(prefix="v3t14a_")
    try:
        out = os.path.join(wd, "BOM_t14a.xlsx")
        with open(SAMPLE_V3, encoding="utf-8") as f:
            sample = json.load(f)
        # 复制 sample_v3 并追加一个名称含「牛奶」且 allergen 为空的物料
        sample["materials"].append({
            "name": "牛奶粉", "unit": "kg", "usage": 20, "yield_rate": 100,
            "material_type": "原料", "process": "S01",
        })
        rc, so, se = run_generate(sample, out)
        assert rc == 0, "T14a 应非阻断(rc=0), rc=%s so=%s se=%s" % (rc, so, se)
        expected = ("WARNING: 物料『牛奶粉』名称疑似含致敏物『乳』"
                    "但未在过敏原中标注，请确认")
        assert expected in so, "T14a 应打印 H1 WARNING(乳), stdout=%s" % so
    finally:
        shutil.rmtree(wd, ignore_errors=True)


def test_v3_t14b_keyword_annotated_no_warning():
    """V3-T14b: 名称含关键词(纯牛奶)且 allergen 已正确标注(乳) → 不产生 H1 WARNING。"""
    wd = tempfile.mkdtemp(prefix="v3t14b_")
    try:
        out = os.path.join(wd, "BOM_t14b.xlsx")
        data = {
            "product_name": "关键词已标注", "category": "食品", "output_rate": 130,
            "materials": [
                {"name": "纯牛奶", "unit": "kg", "usage": 10, "yield_rate": 100,
                 "material_type": "原料", "process": "", "allergen": "乳"},
            ],
            "processes": [],
        }
        rc, so, se = run_generate(data, out)
        assert rc == 0, "T14b 应成功(rc=0), rc=%s so=%s se=%s" % (rc, so, se)
        assert "名称疑似含致敏物" not in so, \
            "T14b 已正确标注不应产生 H1 WARNING, stdout=%s" % so
    finally:
        shutil.rmtree(wd, ignore_errors=True)


def test_v3_t14c_keyword_classes():
    """V3-T14c: 名称含 蛋/花生/大豆/麸质/坚果/虾/鱼 其一且未标注 → 对应类 WARNING 出现。"""
    wd = tempfile.mkdtemp(prefix="v3t14c_")
    try:
        out = os.path.join(wd, "BOM_t14c.xlsx")
        data = {
            "product_name": "关键词多类", "category": "食品", "output_rate": 130,
            "materials": [
                {"name": "鸡蛋液", "unit": "kg", "usage": 1, "yield_rate": 100,
                 "material_type": "原料", "process": ""},
                {"name": "花生碎", "unit": "kg", "usage": 1, "yield_rate": 100,
                 "material_type": "原料", "process": ""},
                {"name": "大豆粉", "unit": "kg", "usage": 1, "yield_rate": 100,
                 "material_type": "原料", "process": ""},
                {"name": "全麦面粉", "unit": "kg", "usage": 1, "yield_rate": 100,
                 "material_type": "原料", "process": ""},
                {"name": "核桃仁", "unit": "kg", "usage": 1, "yield_rate": 100,
                 "material_type": "原料", "process": ""},
                {"name": "鲜虾仁", "unit": "kg", "usage": 1, "yield_rate": 100,
                 "material_type": "原料", "process": ""},
                {"name": "鱼肉糜", "unit": "kg", "usage": 1, "yield_rate": 100,
                 "material_type": "原料", "process": ""},
            ],
            "processes": [],
        }
        rc, so, se = run_generate(data, out)
        assert rc == 0, "T14c 应非阻断(rc=0), rc=%s so=%s se=%s" % (rc, so, se)
        expected_classes = ["蛋类", "花生", "大豆", "含麸质谷物",
                            "坚果", "甲壳类", "鱼类"]
        for cls in expected_classes:
            frag = "名称疑似含致敏物『%s』但未在过敏原中标注" % cls
            assert frag in so, "T14c 应含类『%s』的 H1 WARNING, stdout=%s" % (cls, so)
    finally:
        shutil.rmtree(wd, ignore_errors=True)


# ----------------------------------------------------------------------------
# 运行器
# ----------------------------------------------------------------------------
TESTS = [
    ("V3-T00", "8列表头(含序号)", test_v3_t00_header8),
    ("V3-T01", "占比%列和=100.0(补差)", test_v3_t01_pct_sum_100),
    ("V3-T02", "占比%分母=可食用合计", test_v3_t02_pct_denominator_edible),
    ("V3-T03", "序号连续跨组不重置", test_v3_t03_seq_continuous),
    ("V3-T04", "配料表仅食品+非食品合计行", test_v3_t04_ingredient_only_food),
    ("V3-T05", "过敏原展示+逆向回收", test_v3_t05_allergen_display_and_recover),
    ("V3-T06", "审批人/日期/标准显示+逆向", test_v3_t06_meta_display_and_recover),
    ("V3-T07", "合计行+逆向跳过", test_v3_t07_total_row_and_skip),
    ("V3-T08", "旧版7列兼容", test_v3_t08_old7_compat),
    ("V3-T09", "过敏原软校验WARNING", test_v3_t09_allergen_warning),
    ("V3-T10", "全产品出品率130.0%", test_v3_t10_output_rate_format),
    ("V3-T11", "稳定排序", test_v3_t11_stable_sort),
    ("V3-T12", "V1-V5校验回归", test_v3_t12_validation_regression),
    ("V3-T13", "闭环一致性", test_v3_t13_closed_loop),
    ("V3-T14a", "关键词告警(未标注)", test_v3_t14a_keyword_warning),
    ("V3-T14b", "关键词已标注无告警", test_v3_t14b_keyword_annotated_no_warning),
    ("V3-T14c", "关键词多类告警", test_v3_t14c_keyword_classes),
]


def main():
    print("==================================================")
    print(" BOM智造师 V3 (V2.1) 独立测试 (QA 严过关)")
    print("==================================================")
    results = []
    for tid, name, fn in TESTS:
        try:
            fn()
            results.append((tid, name, True, "PASS"))
            print("%-8s %-30s PASS" % (tid, name))
        except AssertionError as e:
            results.append((tid, name, False, str(e)))
            print("%-8s %-30s FAIL  %s" % (tid, name, str(e)))
        except Exception as e:  # noqa: BLE001
            results.append((tid, name, False, "EXCEPTION: %s" % e))
            print("%-8s %-30s FAIL  EXCEPTION: %s" % (tid, name, e))

    total = len(results)
    passed = sum(1 for r in results if r[2])
    failed = total - passed
    print("--------------------------------------------------")
    print("总计: %d   通过: %d   失败: %d" % (total, passed, failed))
    print("IS_PASS: %s" % ("YES" if failed == 0 else "NO"))
    if failed:
        print("失败用例:")
        for tid, name, ok, detail in results:
            if not ok:
                print("  [%s] %s -> %s" % (tid, name, detail))
    print("--------------------------------------------------")
    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
